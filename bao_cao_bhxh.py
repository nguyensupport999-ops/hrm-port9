"""
Module xuất báo cáo trích nộp BHXH — dùng chung cho DN & HKD.
Mẫu: DANH SÁCH LAO ĐỘNG NỘP BẢO HIỂM
"""

import io
import datetime
import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import base64

def _auto_download_excel(file_data: bytes, filename: str):
    """Tự động kích hoạt tải file Excel ngay khi vừa tạo xong — không cần bấm thêm nút Tải."""
    b64 = base64.b64encode(file_data).decode()
    dl_html = f"""
    <html><body>
    <a id="auto_dl_link" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}"></a>
    <script>document.getElementById('auto_dl_link').click();</script>
    </body></html>
    """
    st.components.v1.html(dl_html, height=0, width=0)

def _safe_float(val):
    """Chuyển giá trị (có thể là str/None) sang float an toàn."""
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _safe_date_format(val, fmt="%d/%m/%Y"):
    """Format ngày an toàn."""
    if val is None:
        return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime(fmt)
    if isinstance(val, str) and len(val) >= 10:
        try:
            return datetime.datetime.strptime(val[:10], "%Y-%m-%d").strftime(fmt)
        except Exception:
            return str(val)
    return str(val)


def _extract_so_hd(so_hdld):
    """Trích xuất số thứ tự từ số HĐLĐ (ví dụ: 05/2026/HĐLĐ-CHL -> 5)"""
    if not so_hdld:
        return 0
    try:
        import re
        match = re.search(r'^(\d+)/', str(so_hdld))
        if match:
            return int(match.group(1))
        return 0
    except:
        return 0


def _is_hop_dong_thu_viec(so_hdld):
    """Kiểm tra xem số HĐLĐ có phải là hợp đồng thử việc không"""
    if not so_hdld:
        return False
    try:
        # Kiểm tra nếu số HĐ có chứa 'HĐTV' hoặc 'HDTV'
        if 'HĐTV' in str(so_hdld) or 'HDTV' in str(so_hdld):
            return True
        return False
    except:
        return False


def xuat_bao_cao_trich_nop_bhxh(db_engine, tu_ngay=None, den_ngay=None):
    """Xuất báo cáo trích nộp BHXH.
    Trả về (excel_bytes, so_nv).
    """
    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        # Query lấy tất cả nhân viên (bao gồm NGHI_VIEC) 
        # và LOẠI BỎ: so_hdld = NULL, so_hdld rỗng, hợp đồng thử việc
        c.execute("""
            SELECT
                ho_ten,
                luong_bao_hiem,
                so_hdld,
                ma_so_bhxh,
                so_cccd,
                chuc_danh_nghe,
                phong_ban_lam_viec,
                thang_bat_dau_bh,
                thang_ket_thuc_bh,
                ngay_ket_thuc,
                ngay_cap_cccd,
                noi_dang_ky_kcb,
                ghi_chu,
                phuong_an_dieu_chinh,
                thang_phuong_an,
                trang_thai_bhxh,
                trang_thai
            FROM nhan_vien
            WHERE trang_thai IN ('DANG_LAM', 'THU_VIEC', 'NGHI_VIEC')
              AND so_hdld IS NOT NULL
              AND so_hdld != ''
              AND so_hdld NOT LIKE '%HĐTV%'
              AND so_hdld NOT LIKE '%HDTV%'
            ORDER BY 
                CAST(SUBSTRING(so_hdld FROM '^([0-9]+)/') AS INTEGER) ASC
        """)
        col_names = [desc[0] for desc in c.description]
        rows = [dict(zip(col_names, r)) for r in c.fetchall()]
    except Exception as e:
        st.warning(f"Lỗi khi truy vấn dữ liệu: {e}")
        # Fallback query
        try:
            conn.close()
        except Exception:
            pass
        conn = db_engine.get_connection()
        c = conn.cursor()
        try:
            c.execute("""
                SELECT 
                    ho_ten, 
                    luong_bao_hiem, 
                    chuc_danh_nghe,
                    ma_so_bhxh,
                    so_cccd,
                    so_hdld,
                    thang_bat_dau_bh,
                    thang_ket_thuc_bh,
                    ngay_ket_thuc,
                    ngay_cap_cccd,
                    noi_dang_ky_kcb,
                    ghi_chu,
                    phuong_an_dieu_chinh,
                    thang_phuong_an,
                    trang_thai_bhxh,
                    trang_thai,
                    phong_ban_lam_viec
                FROM nhan_vien
                WHERE trang_thai IN ('DANG_LAM', 'THU_VIEC', 'NGHI_VIEC')
                  AND so_hdld IS NOT NULL
                  AND so_hdld != ''
                  AND so_hdld NOT LIKE '%HĐTV%'
                  AND so_hdld NOT LIKE '%HDTV%'
                ORDER BY 
                    CAST(SUBSTRING(so_hdld FROM '^([0-9]+)/') AS INTEGER) ASC
            """)
            col_names = [desc[0] for desc in c.description]
            rows = [dict(zip(col_names, r)) for r in c.fetchall()]
        except Exception as e2:
            st.error(f"Lỗi fallback: {e2}")
            rows = []
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # Lọc theo khoảng thời gian (nếu có)
    def _to_date(val):
        if val is None:
            return None
        if isinstance(val, datetime.datetime):
            return val.date()
        if isinstance(val, datetime.date):
            return val
        if isinstance(val, str) and len(val) >= 10:
            try:
                return datetime.datetime.strptime(val[:10], "%Y-%m-%d").date()
            except Exception:
                return None
        return None

    # Lọc: giữ lại NV "đang tham gia BHXH trong kỳ" = đã bắt đầu BH từ kỳ này
    # trở về trước, VÀ (chưa kết thúc BH hoặc kết thúc BH từ kỳ này trở về sau).
    # Dùng CHUNG 2 mốc thang_bat_dau_bh / thang_ket_thuc_bh (fallback ngay_ket_thuc)
    # với báo cáo D02-LT để đảm bảo số NV tăng/giảm giữa 2 báo cáo khớp nhau.
    if tu_ngay or den_ngay:
        filtered = []
        for r in rows:
            ngay_bd = _to_date(r.get("thang_bat_dau_bh"))
            if den_ngay and ngay_bd and ngay_bd > den_ngay:
                continue  # BH bắt đầu sau kỳ báo cáo

            ngay_kt = _to_date(r.get("thang_ket_thuc_bh")) or _to_date(r.get("ngay_ket_thuc"))
            if tu_ngay and ngay_kt and ngay_kt < tu_ngay:
                continue  # Đã kết thúc BH trước kỳ báo cáo

            filtered.append(r)
        rows = filtered

    # ── Tạo Excel ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DS Nop BH"

    # Styles
    font_title = Font(name="Times New Roman", size=13, bold=True)
    font_header = Font(name="Times New Roman", size=10, bold=True)
    font_normal = Font(name="Times New Roman", size=10)
    font_header_white = Font(name="Times New Roman", size=10, bold=True, color="FFFFFF")
    font_note = Font(name="Times New Roman", size=9, italic=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_total = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Column widths - THÊM cột KPCĐ 2% và cột Trạng thái
    col_widths = {"A": 5, "B": 24, "C": 16, "D": 16, "E": 18, "F": 16, "G": 16,
                  "H": 18,  # KPCĐ 2%
                  "I": 16, "J": 14, "K": 16, "L": 16,  # CCCD
                  "M": 24, "N": 20, "O": 16, "P": 14}  # Chức vụ, Nơi ĐK KCB, Ghi chú, Trạng thái
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Title (mở rộng đến cột P)
    ws.merge_cells("A1:P1")
    ws["A1"] = "DANH SÁCH LAO ĐỘNG NỘP BẢO HIỂM"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    ky_label = ""
    if tu_ngay and den_ngay:
        if tu_ngay.year == den_ngay.year and tu_ngay.month == den_ngay.month:
            ky_label = f"Kỳ: Tháng {tu_ngay.month:02d}/{tu_ngay.year}"
        else:
            ky_label = f"Kỳ: từ {tu_ngay.strftime('%d/%m/%Y')} đến {den_ngay.strftime('%d/%m/%Y')}"
    else:
        ky_label = f"Kỳ: {datetime.date.today().strftime('%m/%Y')}"
    if ky_label:
        ws.merge_cells("A2:P2")
        ws["A2"] = ky_label
        ws["A2"].font = font_normal
        ws["A2"].alignment = align_center

    # Header - THÊM cột KPCĐ 2% và cột Trạng thái
    headers = [
        "STT", "HỌ VÀ TÊN", "THÁNG\nBẮT ĐẦU\nNỘP BH", "TIỀN\nLƯƠNG\nNỘP BH",
        "SỐ TIỀN\nBH PHẢI\nNỘP/THÁNG\n32%",
        "DN PHẢI\nNỘP\n21,5%", "NLĐ PHẢI\nNỘP\n10,5%",
        "KINH PHÍ\nCÔNG ĐOÀN\n2%",  # Cột mới
        "SỐ HĐLĐ", "MÃ SỐ BH", "CCCD", "NGÀY CẤP\nCCCD",
        "CHỨC VỤ", "NƠI ĐK\nKCB", "GHI CHÚ", "TRẠNG\nTHÁI"
    ]

    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = font_header_white
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = fill_header
    ws.row_dimensions[header_row].height = 50

    # Data rows
    data_row = header_row + 1
    tong_luong = 0
    tong_32 = 0
    tong_dn = 0
    tong_nld = 0
    tong_kpcd = 0
    tong_nv_hop_le = 0  # Đếm số NV có lương > 0 và đang làm

    # Lấy tháng/năm hiện tại của kỳ báo cáo
    ky_thang = tu_ngay.month if tu_ngay else datetime.date.today().month
    ky_nam = tu_ngay.year if tu_ngay else datetime.date.today().year

    for idx, r in enumerate(rows, 1):
        # Lấy lương, nếu NULL hoặc rỗng thì = 0
        luong_bh = _safe_float(r.get("luong_bao_hiem", 0))
        trang_thai = r.get("trang_thai", "")
        phuong_an = r.get("phuong_an_dieu_chinh", "") or ""
        thang_phuong_an = r.get("thang_phuong_an", "") or ""
        
        # Mặc định tính bình thường
        tinh_binh_thuong = True
        tinh_chi_phi_cong_doan = False
        tat_ca_bang_0 = False
        
        # LOGIC XỬ LÝ PHƯƠNG ÁN ĐIỀU CHỈNH
        # GH1/GH2/GH3/GH4 đều là "Giảm hẳn" -> so sánh KỲ ĐANG BÁO CÁO với THÁNG QĐNS
        # (thang_phuong_an): kỳ TRƯỚC tháng QĐNS -> NV vẫn đang đóng BH bình thường
        # trong kỳ đó -> tính bình thường; kỳ BẰNG/SAU tháng QĐNS -> đã giảm hẳn -> = 0.
        # ===== QUY TẮC TÍNH THEO PHƯƠNG ÁN ĐIỀU CHỈNH (đã chốt theo luật BHXH) =====
        # - GH1/GH2/GH3/GH4 và KL/OF (nghỉ không lương): 0 HOÀN TOÀN kể từ tháng phương
        #   án trở đi (tính bình thường ở các tháng TRƯỚC đó).
        # - TS (thai sản) và TNLD (tai nạn lao động): LUÔN tính bình thường, không bao
        #   giờ zero, bất kể thang_phuong_an là tháng nào.
        # - Phương án khác (TD/TM/TC/ON) hoặc để trống: mặc định tính bình thường.
        PHUONG_AN_GIAM_HAN = ("GH1", "GH2", "GH3", "GH4", "KL", "OF")
        PHUONG_AN_LUON_BINH_THUONG = ("TS", "TNLD")

        if phuong_an in PHUONG_AN_LUON_BINH_THUONG:
            tinh_binh_thuong = True
        elif phuong_an in PHUONG_AN_GIAM_HAN and thang_phuong_an:
            try:
                if "/" in thang_phuong_an:
                    pa_thang, pa_nam = map(int, thang_phuong_an.split("/"))
                else:
                    pa_date = datetime.datetime.strptime(thang_phuong_an[:10], "%Y-%m-%d").date()
                    pa_thang, pa_nam = pa_date.month, pa_date.year
                
                if (ky_nam, ky_thang) < (pa_nam, pa_thang):
                    tinh_binh_thuong = True
                else:
                    tat_ca_bang_0 = True
            except:
                tinh_binh_thuong = True
        # else: giữ mặc định tinh_binh_thuong = True (TD/TM/TC/ON hoặc phuong_an trống)
        
        # Lương = 0 thì luôn = 0, KHÔNG gộp chung điều kiện với trang_thai nữa — trạng
        # thái hiện tại không còn quyết định được có tính tiền hay không (đã chuyển hẳn
        # sang xét tat_ca_bang_0 ở trên, dựa theo KỲ so với tháng phương án).
        if luong_bh == 0:
            so_tien_32 = 0
            dn_215 = 0
            nld_105 = 0
            kpcd_2 = 0
        elif tat_ca_bang_0:
            so_tien_32 = 0
            dn_215 = 0
            nld_105 = 0
            kpcd_2 = 0
        elif tinh_chi_phi_cong_doan:
            so_tien_32 = 0
            dn_215 = 0
            nld_105 = 0
            kpcd_2 = round(luong_bh * 0.02)  # Chỉ tính công đoàn
        else:
            # Tính bình thường
            so_tien_32 = round(luong_bh * 0.32)
            dn_215 = round(luong_bh * 0.215)
            nld_105 = round(luong_bh * 0.105)
            kpcd_2 = round(luong_bh * 0.02)

        # Chỉ tính tổng cho NV thực sự đang đóng BH bình thường/công đoàn TRONG KỲ NÀY
        # (dùng tat_ca_bang_0 đã xét theo kỳ, KHÔNG dùng trạng thái hiện tại nữa)
        if luong_bh > 0 and not tat_ca_bang_0:
            tong_luong += luong_bh
            tong_32 += so_tien_32
            tong_dn += dn_215
            tong_nld += nld_105
            tong_kpcd += kpcd_2
            tong_nv_hop_le += 1

        # Format ngày tháng bắt đầu nộp BH: chỉ lấy tháng/năm
        ngay_bh_str = _safe_date_format(r.get("thang_bat_dau_bh"), "%m/%Y")
        # Format ngày cấp CCCD: đầy đủ ngày/tháng/năm
        ngay_cap_str = _safe_date_format(r.get("ngay_cap_cccd"), "%d/%m/%Y")
        
        # Hiển thị trạng thái tiếng Việt — PHẢI phản ánh đúng trạng thái TẠI KỲ báo cáo,
        # không phải trạng thái HIỆN TẠI trong hồ sơ. NV đã báo giảm GH1-4 nhưng kỳ báo
        # cáo còn TRƯỚC tháng phương án thì tại kỳ đó vẫn đang đóng BH bình thường ->
        # không được hiện "Nghỉ việc" (sẽ gây hiểu nhầm khi đối chiếu với cột tiền khác 0).
        if tat_ca_bang_0:
            if phuong_an == 'KL':
                trang_thai_display = 'Nghỉ không lương (KL)'
            elif phuong_an == 'OF':
                trang_thai_display = 'Nghỉ ốm dài ngày (OF)'
            else:
                trang_thai_display = 'Nghỉ việc'
        elif trang_thai == 'THU_VIEC':
            trang_thai_display = 'Thử việc'
        else:
            trang_thai_display = 'Đang làm'

        values = [
            idx,
            r.get("ho_ten", ""),
            ngay_bh_str,
            luong_bh,
            so_tien_32,
            dn_215,
            nld_105,
            kpcd_2,  # KPCĐ 2%
            r.get("so_hdld", "") or "",
            r.get("ma_so_bhxh", "") or "",
            r.get("so_cccd", "") or "",
            ngay_cap_str,
            r.get("chuc_danh_nghe", "") or "",
            r.get("noi_dang_ky_kcb", "") or "",
            r.get("ghi_chu", "") or "",
            trang_thai_display,
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=data_row, column=col_idx, value=val)
            cell.font = font_normal
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx in (4, 5, 6, 7, 8):  # Các cột số (thêm KPCĐ)
                cell.alignment = align_right
                cell.number_format = '#,##0'
            elif col_idx in (3, 9, 10, 11, 12, 16):  # Các cột center (thêm trạng thái)
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        data_row += 1

    # Dòng tổng (mở rộng đến cột C)
    ws.merge_cells(f"A{data_row}:C{data_row}")
    cell_tong = ws.cell(row=data_row, column=1, value="TỔNG CỘNG")
    cell_tong.font = font_header
    cell_tong.alignment = align_center
    cell_tong.border = thin_border
    cell_tong.fill = fill_total
    # Border cho B, C (đã merge)
    for mc in [2, 3]:
        ws.cell(row=data_row, column=mc).border = thin_border

    # Các cột tổng từ D đến H (bao gồm KPCĐ)
    for col_idx, val in [(4, tong_luong), (5, tong_32), (6, tong_dn), 
                         (7, tong_nld), (8, tong_kpcd)]:
        cell = ws.cell(row=data_row, column=col_idx, value=val)
        cell.font = font_header
        cell.alignment = align_right
        cell.border = thin_border
        cell.number_format = '#,##0'
        cell.fill = fill_total

    for col_idx in range(9, 17):
        ws.cell(row=data_row, column=col_idx).border = thin_border
        ws.cell(row=data_row, column=col_idx).fill = fill_total

    # Ghi chú
    data_row += 2
    ws.merge_cells(f"A{data_row}:P{data_row}")
    ws[f"A{data_row}"] = "Ghi chú: DN phải nộp 21,5% = BHXH 14% + BHYT 3% + BHTN 1% + BHTNLĐ-BNN 0,5% + Quỹ HT 3%"
    ws[f"A{data_row}"].font = font_note
    data_row += 1
    ws.merge_cells(f"A{data_row}:P{data_row}")
    ws[f"A{data_row}"] = "         NLĐ phải nộp 10,5% = BHXH 8% + BHYT 1,5% + BHTN 1%"
    ws[f"A{data_row}"].font = font_note
    data_row += 1
    ws.merge_cells(f"A{data_row}:P{data_row}")
    ws[f"A{data_row}"] = "         DN phải nộp Kinh phí công đoàn 2%"
    ws[f"A{data_row}"].font = font_note
    data_row += 1
    ws.merge_cells(f"A{data_row}:P{data_row}")
    ws[f"A{data_row}"] = f"         Tổng số nhân viên đang tham gia BH: {tong_nv_hop_le}"
    ws[f"A{data_row}"].font = font_note

    # Ký tên
    data_row += 2
    ws.merge_cells(f"I{data_row}:P{data_row}")
    ws[f"I{data_row}"] = f"Ngày ..... tháng ..... năm {datetime.date.today().year}"
    ws[f"I{data_row}"].font = font_normal
    ws[f"I{data_row}"].alignment = align_center
    data_row += 1
    ws.merge_cells(f"A{data_row}:D{data_row}")
    ws[f"A{data_row}"] = "KẾ TOÁN"
    ws[f"A{data_row}"].font = font_header
    ws[f"A{data_row}"].alignment = align_center
    ws.merge_cells(f"I{data_row}:P{data_row}")
    ws[f"I{data_row}"] = "GIÁM ĐỐC"
    ws[f"I{data_row}"].font = font_header
    ws[f"I{data_row}"].alignment = align_center

    # Xuất bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), len(rows)


def render_xuat_bao_cao_bhxh(db_engine):
    """UI xuất báo cáo trích nộp BHXH."""
    st.subheader("📥 Xuất báo cáo trích nộp BHXH")
    st.caption("Mẫu: Danh sách lao động nộp bảo hiểm (32% = DN 21,5% + NLĐ 10,5%) + Kinh phí công đoàn 2%")

    col1, col2 = st.columns(2)
    with col1:
        thang_chon = st.selectbox("📅 Tháng:", list(range(1, 13)),
                                   index=datetime.date.today().month - 1,
                                   key="bhxh_bc_thang")
    with col2:
        nam_chon = st.selectbox("📅 Năm:",
                                 list(range(datetime.date.today().year - 2, datetime.date.today().year + 2)),
                                 index=2, key="bhxh_bc_nam")

    import calendar
    tu_ngay = datetime.date(nam_chon, thang_chon, 1)
    den_ngay = datetime.date(nam_chon, thang_chon, calendar.monthrange(nam_chon, thang_chon)[1])

    if st.button("📥 Xuất báo cáo", type="primary", key="btn_xuat_bc_bhxh"):
        try:
            with st.spinner("Đang tạo báo cáo..."):
                excel_bytes, so_nv = xuat_bao_cao_trich_nop_bhxh(db_engine, tu_ngay, den_ngay)

                if so_nv == 0:
                    st.warning("⚠️ Không có NV nào thỏa điều kiện.")
                else:
                    ten_file = f"DS_Nop_BH_{nam_chon}{thang_chon:02d}.xlsx"
                    st.success(f"✅ Đã tạo báo cáo cho {so_nv} NV! Đang tự động tải file...")
                    _auto_download_excel(excel_bytes, ten_file)
        except Exception as e:
            st.error(f"❌ Lỗi: {e}")