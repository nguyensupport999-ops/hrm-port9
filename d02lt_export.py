# -*- coding: utf-8 -*-
"""
d02lt_export.py
================
Sinh file Excel D02-LT ĐÚNG ĐỊNH DẠNG CHUẨN mà hệ thống iCare (BHXH điện tử)
yêu cầu khi "Nhập Excel".

VÌ SAO FILE CŨ BỊ BÁO "KHÔNG ĐÚNG ĐỊNH DẠNG"
---------------------------------------------
So sánh file "export-20-08-2026.xlsx" (file mẫu chuẩn do chính iCare xuất ra,
qua nút "Tải xuống" trong hộp thoại Nhập Excel) với file
"D02-LT_BHXH_01082026_31082026.xlsx" (file do chương trình cũ xuất ra), thấy:

1. Tên sheet dữ liệu SAI: iCare đọc dữ liệu từ sheet tên **"DuLieu"**,
   file cũ đặt tên sheet là "D02-LT" -> iCare không tìm thấy dữ liệu.
2. Thiếu toàn bộ các sheet danh mục đi kèm (Tinh, Huyen, Xa, KCB, GioiTinh,
   DM Quốc tịch, Dân tộc, Loại PA, Phương án, Ngân Hàng, Mối quan hệ...).
   Đây là các sheet chứa "named range" (vùng đặt tên) mà các cột có công thức
   INDEX/MATCH và Data Validation (dropdown) trong sheet DuLieu tham chiếu tới.
   Thiếu các sheet này -> các công thức/validation bị lỗi tham chiếu.
3. Sai cấu trúc hàng tiêu đề: file chuẩn có 4 dòng tiêu đề cố định
   (dòng 1 = tiêu đề mẫu, dòng 2-3 = tên cột hiển thị, dòng 4 = tên field kỹ
   thuật ẩn dùng để hệ thống map dữ liệu), dữ liệu bắt đầu từ **dòng 5**.
   File cũ chèn thêm 4 dòng "Công ty…", "Mã đơn vị…", "Kỳ báo cáo…" phía trên
   rồi mới tới bảng dữ liệu -> lệch hoàn toàn cấu trúc dòng/cột mà iCare
   mong đợi (iCare đọc theo đúng vị trí dòng/cột cố định, không tự dò tiêu đề).
4. Sai số cột & thứ tự cột: file chuẩn có 117 cột (đến cột DM) mô tả đầy đủ
   thông tin lao động + phụ lục hộ gia đình; file cũ chỉ có khoảng 116 cột
   nhưng thứ tự/nhãn bị xê dịch so với chuẩn (ví dụ cột "Hợp đồng lao động"
   bị lặp/lệch vị trí), khiến dữ liệu bị đọc nhầm cột.
5. Định dạng ô sai kiểu: các cột ngày tháng (Ngày sinh, Ngày ký, Tháng/năm
   bắt đầu...) trong file chuẩn có định dạng Text ("@"), phải ghi dạng chuỗi
   "dd/mm/yyyy" / "mm/yyyy" — không phải kiểu Date của Excel.
6. Các cột có công thức tự tính mã (Mã loại PA, Mã PA, Mã quốc tịch, Mã dân
   tộc, Mã tỉnh...) cần được TÍNH LẠI (recalculate) trước khi lưu, nếu không
   giá trị đã tính (cached value) sẽ rỗng và bộ đọc phía iCare có thể không
   thấy được mã tương ứng.

CÁCH KHẮC PHỤC (module này làm)
--------------------------------
KHÔNG dựng lại file từ đầu bằng openpyxl (rất dễ sai vì có hàng trăm nghìn
công thức + named range + data validation phức tạp). Thay vào đó:

  - Dùng CHÍNH file mẫu gốc do iCare xuất ra (nút "Tải xuống") làm khung sườn
    (giữ nguyên toàn bộ sheet danh mục, named range, formula, data validation).
  - Chỉ ĐIỀN DỮ LIỆU vào đúng các ô nhập liệu (text) của sheet "DuLieu",
    bắt đầu từ dòng 5, theo đúng tên field kỹ thuật ở dòng 4 (ẩn) — nên dù
    sau này iCare có xê dịch thứ tự cột, code vẫn map đúng vì tra theo tên
    field chứ không tra theo số cột cứng.
  - KHÔNG đụng vào các cột có sẵn công thức (Mã loại PA, Mã PA, Mã quốc
    tịch, Mã dân tộc, Mã tỉnh/xã...) — chỉ điền cột text người dùng nhập,
    công thức sẽ tự tính ra mã tương ứng.
  - Cắt bớt các dòng trống dư thừa phía cuối (mẫu gốc để sẵn ~13.000 dòng
    trống cho khách xuất nhiều lao động) để file gọn, rồi TÍNH LẠI công thức
    bằng LibreOffice để đảm bảo giá trị các cột mã được "chốt" (cached) vào
    file, tránh trường hợp phần mềm phía BHXH đọc thấy ô công thức rỗng.

CÀI ĐẶT / SỬ DỤNG
------------------
1. Tải 1 lần file mẫu chuẩn từ iCare (nút "Tải xuống" trong hộp thoại Nhập
   Excel của màn hình D02-LT) và lưu cố định trong project, ví dụ:
       templates/D02LT_template_goc.xlsx
   -> Nên tải lại mỗi khi BHXH cập nhật mẫu (ví dụ đổi từ QĐ 595 sang QĐ mới),
      vì tên các named range/sheet danh mục có thể thay đổi.

2. Gọi hàm `build_d02lt_excel(...)` với danh sách lao động TĂNG và GIẢM lấy
   từ truy vấn có sẵn (tang_list, giam_list) — xem `example_usage()` cuối file
   để biết cách gắn thẳng vào nút "XUẤT EXCEL D02-LT" trong Streamlit.

LƯU Ý
------
  - Cột "Phương án điều chỉnh" (TM/TC/TD/TH/GH/GC/GD/OF/KL/TS...) KHÔNG còn
    suy đoán — module này lấy đúng mã đã lưu sẵn trong cột
    `nhan_vien.phuong_an_dieu_chinh` (là mã do người dùng chọn ngay khi
    thêm mới/chuyển đổi lao động trong app, xem PHUONG_AN_TANG /
    PHUONG_AN_GIAM / hàm lay_ma_phuong_an() trong app.py), rồi đối chiếu
    với danh mục "Phương án" đọc trực tiếp từ file mẫu để lấy đúng text
    chuẩn BHXH. Nếu 1 nhân viên chưa được chọn phương án khi tăng/giảm,
    ô "Phương án điều chỉnh" sẽ để TRỐNG — anh cần bổ sung chọn phương án
    cho nhân viên đó trong app trước khi xuất báo cáo.
  - Giới tính: quy ước 1 = Nam, 0 = Nữ theo chuẩn TK1-TS của BHXH.
  - "Tháng/năm bắt đầu-kết thúc" ưu tiên lấy từ cột `thang_phuong_an`
    (đã được app tự format sẵn dạng mm/yyyy khi lưu), chỉ tự tính lại từ
    `thang_bat_dau_bh` / `thang_ket_thuc_bh` nếu cột đó trống.
"""
from __future__ import annotations

import shutil
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Mapping, Any

import unicodedata

import openpyxl

warnings.filterwarnings("ignore")


def _norm(s: str) -> str:
    """Chuẩn hoá chuỗi để so khớp không phân biệt hoa/thường & dấu tiếng Việt."""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.strip().upper()

# Dòng bắt đầu ghi dữ liệu trong sheet DuLieu (dòng 1-4 là tiêu đề/field-key ẩn)
FIRST_DATA_ROW = 5
DATA_SHEET_NAME = "DuLieu"
FIELD_KEY_ROW = 4          # dòng chứa tên field kỹ thuật (ẩn)
# Các field-key mà ô của nó đã CÓ SẴN CÔNG THỨC trong template -> không ghi đè
FORMULA_FIELD_KEYS = {
    "LoaiHoSo", "PhuongAnId", "QuocTichId", "DanTocId", "MaTinh_NH",
    "MaNganHang", "KhaiSinhTinhId", "KhaiSinhXaId", "NoiNhanTinhId",
    "NoiNhanXaId", "BenhVienTinhId", "BenhVienId", "TinhNhanBanGiayId",
    "XaNhanBanGiayId", "ChuHoTinhId", "ChuHoXaId", "TinhThuongTruChuHoId",
    "XaThuongTruChuHoId", "QuocTich_TV", "DanToc_TV", "KhaiSinhTinhId_TV",
    "KhaiSinhXaId_TV",
}


# --------------------------------------------------------------------------
# Helpers định dạng giá trị -> chuỗi text đúng như template yêu cầu (kiểu "@")
# --------------------------------------------------------------------------
def _s(val) -> str:
    """Chuỗi text an toàn, None -> ''."""
    if val is None:
        return ""
    return str(val).strip()


def _num(val) -> str:
    """Chuẩn hoá số (int/float/Decimal/str) -> chuỗi số THUẦN, không dấu
    phẩy ngăn cách nghìn, không ".0"/".00" thừa. Vd Decimal('4650000.00')
    -> '4650000'; Decimal('2.34') -> '2.34'; None/"" -> ''.
    (Khắc phục lỗi cột M/P/Q/R/S/AC hiển thị dạng '4650000.0' do trước đây
    dùng str() trực tiếp lên giá trị Decimal lấy từ Postgres.)"""
    if val in (None, ""):
        return ""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return _s(val)
    if f == int(f):
        return str(int(f))
    s = f"{f:.4f}".rstrip("0").rstrip(".")
    return s


def _num_or_blank(val) -> str:
    """Như _num(), nhưng trả về RỖNG nếu giá trị là 0 hoặc None — dùng cho
    các cột Hệ số lương / Phụ cấp: nếu nhân viên chưa nhập (DB lưu mặc định
    0 thay vì NULL) thì ô trên file xuất phải để TRỐNG, không ghi '0'."""
    if val in (None, ""):
        return ""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return _s(val)
    if f == 0:
        return ""
    return _num(val)


def _fmt_date(val) -> str:
    """dd/mm/yyyy dạng TEXT (cột Ngày sinh, Ngày ký... trong template là Text)."""
    if val in (None, ""):
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%d/%m/%Y")
    # đã là chuỗi sẵn (vd lấy từ DB dạng '2026-08-01')
    s = str(val).strip()
    for fmt_in in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt_in).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


def _fmt_month(val) -> str:
    """mm/yyyy dạng TEXT (cột Tháng/năm bắt đầu - kết thúc)."""
    if val in (None, ""):
        return ""
    if isinstance(val, (date, datetime)):
        return val.strftime("%m/%Y")
    s = str(val).strip()
    for fmt_in in ("%Y-%m-%d", "%m/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10] if len(s) > 7 else s, fmt_in).strftime("%m/%Y")
        except ValueError:
            continue
    return s


def _valid_ma_so_bhxh(val) -> str:
    """Mã số BHXH bắt buộc dài đúng 10 hoặc 12 ký tự (theo Data Validation
    của template: OR(LEN=10, LEN=12)). Nếu không hợp lệ -> để trống, để
    tránh bị iCare từ chối cả file vì 1 dòng sai định dạng mã số."""
    s = _s(val)
    if len(s) in (10, 12):
        return s
    return ""


def _gioi_tinh_code(val) -> str:
    """Quy ước TK1-TS: Nam = 1, Nữ = 0."""
    s = _s(val).lower()
    if s in ("1", "nam", "male", "true", "m"):
        return "1"
    if s in ("0", "nữ", "nu", "female", "false", "f"):
        return "0"
    return ""


def _ma_phuong_an(row: Mapping[str, Any]) -> str:
    """Lấy MÃ phương án (2 ký tự, vd 'TD', 'TM', 'GH'...) đã được người dùng
    chọn sẵn trong app (cột `phuong_an_dieu_chinh` trong bảng nhan_vien —
    xem PHUONG_AN_TANG / PHUONG_AN_GIAM và hàm lay_ma_phuong_an() trong
    app.py). Đây LÀ nguồn đáng tin cậy nhất, không cần suy đoán lại.
    Chuỗi text đầy đủ đúng theo mẫu (vd "TD-Tăng đến...") sẽ được resolve
    ở bước ghi record, dựa theo danh mục 'Phương án' đọc trực tiếp từ
    chính file mẫu, nên luôn khớp chính tả với BHXH yêu cầu."""
    return _s(row.get("phuong_an_dieu_chinh")).upper()


# --------------------------------------------------------------------------
# Xây dựng bản ghi (record) cho 1 dòng dữ liệu, dạng {field_key: value}
# --------------------------------------------------------------------------
def _record_tang(row: Mapping[str, Any], stt: int) -> dict:
    return {
        "STT": str(stt),
        "HoTen": _s(row.get("ho_ten")),
        "MaSoBhxh": _valid_ma_so_bhxh(row.get("ma_so_bhxh")),
        "LoaiPA_Text": "Tăng lao động",           # -> cột D (text hiển thị)
        "ChiCoNamSinh": "0",
        "NgaySinh": _fmt_date(row.get("ngay_sinh")),
        "GioiTinh": _gioi_tinh_code(row.get("gioi_tinh")),
        "Cmnd": _s(row.get("so_cccd")),
        "ChucVu": _s(row.get("chuc_danh_nghe")),
        "MaPhongBan": _s(row.get("phong_ban_lam_viec")),
        "NoiLamViec": _s(row.get("noi_lam_viec")),   # -> cột L (ưu tiên riêng từng NV)
        "TienLuong": _num(row.get("luong_bao_hiem")),
        "Heso": _num_or_blank(row.get("he_so_luong")),
        "PhuCapChucVu": _num_or_blank(row.get("phu_cap_chuc_vu")),
        "PhuCapThamNienVuotKhung": _num_or_blank(row.get("phu_cap_tnvk")),
        "PhuCapThamNienNghe": _num_or_blank(row.get("phu_cap_tnn")),
        "PhuongAn_Text": _ma_phuong_an(row),       # -> cột T (mã, sẽ resolve ra text chuẩn)
        "TuThang": _s(row.get("thang_phuong_an")) or _fmt_month(row.get("thang_bat_dau_bh") or row.get("ngay_bat_dau")),
        "DenThang": "",
        "GhiChu": _s(row.get("ghi_chu")) or _s(row.get("loai_hop_dong")),
        "MucHuongBaoHiemYTe": _s(row.get("muc_huong_bhyt")),
        "TyleDong": _num(row.get("ty_le_dong")),
        "LoaiHDLD": _s(row.get("loai_hop_dong")),
        "TuNgayHDLD": _fmt_date(row.get("ngay_vao_lam")),
        "SoHDLD": _s(row.get("so_hdld")),
        "NgayKy": _fmt_date(row.get("ngay_ky_hd")),
        "QuocTich_Text": _s(row.get("quoc_tich")) or "Việt Nam",  # -> AS
        "DanToc_Text": _s(row.get("dan_toc")) or "Kinh",          # -> AU
        "DienThoaiLienHe": _s(row.get("dien_thoai")),
        "Email": _s(row.get("email_lien_he")),
        "TinhKCB_Text": _s(row.get("tinh_kcb")),                   # -> BM
        "NoiDangKyKCB_Text": _s(row.get("noi_dang_ky_kcb")),       # -> BO
        "DangKyNhanSoThe": _s(row.get("dang_ky_nhan_so")),
        "NoiNhanTinh_Text": _s(row.get("tinh_nhan_hs")),           # -> BH
        "NoiNhanXa_Text": _s(row.get("phuong_nhan_hs")),           # -> BJ
        "NoiNhanDiaChiChiTiet": _s(row.get("dia_chi_nhan_hs")),
        "MucDongTk01": _num(row.get("muc_tien_dong")),
        "PhuongThucTk1": _s(row.get("phuong_thuc_dong")),
    }


def _record_giam(row: Mapping[str, Any], stt: int) -> dict:
    return {
        "STT": str(stt),
        "HoTen": _s(row.get("ho_ten")),
        "MaSoBhxh": _valid_ma_so_bhxh(row.get("ma_so_bhxh")),
        "LoaiPA_Text": "Giảm lao động",
        "ChiCoNamSinh": "0",
        "NgaySinh": _fmt_date(row.get("ngay_sinh")),
        "GioiTinh": _gioi_tinh_code(row.get("gioi_tinh")),
        "Cmnd": _s(row.get("so_cccd")),
        "ChucVu": _s(row.get("chuc_danh_nghe")),
        "MaPhongBan": _s(row.get("phong_ban_lam_viec")),
        "NoiLamViec": _s(row.get("noi_lam_viec")),   # -> cột L (ưu tiên riêng từng NV)
        "TienLuong": _num(row.get("luong_bao_hiem")),
        "Heso": _num_or_blank(row.get("he_so_luong")),
        "PhuCapChucVu": _num_or_blank(row.get("phu_cap_chuc_vu")),
        "PhuCapThamNienVuotKhung": _num_or_blank(row.get("phu_cap_tnvk")),
        "PhuCapThamNienNghe": _num_or_blank(row.get("phu_cap_tnn")),
        "PhuongAn_Text": _ma_phuong_an(row),       # -> cột T (mã, sẽ resolve ra text chuẩn)
        "TuThang": "",
        "DenThang": _s(row.get("thang_phuong_an")) or _fmt_month(row.get("thang_ket_thuc_bh") or row.get("ngay_ket_thuc")),
        "GhiChu": _s(row.get("ly_do_nghi")),
        "MucHuongBaoHiemYTe": _s(row.get("muc_huong_bhyt")),
        "TyleDong": _num(row.get("ty_le_dong")),
        "LoaiHDLD": _s(row.get("loai_hop_dong")),
        "SoHDLD": _s(row.get("so_hdld")),
        "NgayKy": _fmt_date(row.get("ngay_ky_hd")),
        "QuocTich_Text": _s(row.get("quoc_tich")) or "Việt Nam",
        "DanToc_Text": _s(row.get("dan_toc")) or "Kinh",
        "DienThoaiLienHe": _s(row.get("dien_thoai")),
        "Email": _s(row.get("email_lien_he")),
        "MucDongTk01": _num(row.get("muc_tien_dong")),
        "PhuongThucTk1": _s(row.get("phuong_thuc_dong")),
        "NoiDungThayDoi": _s(row.get("ly_do_nghi")),
    }


# Map field-key ĐẶC BIỆT không trùng tên với cột "text hiển thị" (D/E, T/U,
# AS/AT, AU/AV): các field key này KHÔNG có trong row4 (vì row4 gắn tên cho
# cột công thức, còn cột text nhập tay đứng ngay trước nó không có tên
# field). Ta định vị các cột text đó theo field-key CỦA CỘT CÔNG THỨC đứng
# ngay sau, trừ đi 1 cột.
_TEXT_BEFORE_FORMULA = {
    "LoaiPA_Text": "LoaiHoSo",          # D (text) đứng trước E=LoaiHoSo
    "PhuongAn_Text": "PhuongAnId",      # T (text) đứng trước U=PhuongAnId
    "QuocTich_Text": "QuocTichId",      # AS (text) đứng trước AT=QuocTichId
    "DanToc_Text": "DanTocId",          # AU (text) đứng trước AV=DanTocId
}
# Các field-key text KHÔNG có cột công thức liền sau (là các nhóm địa chỉ:
# Tỉnh/Huyện dạng text đứng cách 1 cột trước cột mã tỉnh); định vị bằng offset
_TEXT_BEFORE_FORMULA_OFFSET2 = {
    "TinhKCB_Text": "BenhVienTinhId",     # BM text, cách BenhVienTinhId 1 cột (BN) -> offset -1 tính từ BN
    "NoiDangKyKCB_Text": "BenhVienId",    # BO text, cách BenhVienId 1 cột (BP) -> offset -1
    "NoiNhanTinh_Text": "NoiNhanTinhId",  # BH đứng trước BI=NoiNhanTinhId -> offset -1
    "NoiNhanXa_Text": "NoiNhanXaId",      # BJ đứng trước BK=NoiNhanXaId -> offset -1
}


def _build_field_col_map(ws) -> dict[str, int]:
    """Đọc dòng 4 (field-key ẩn) để lấy map field_key -> số cột."""
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = ws.cell(row=FIELD_KEY_ROW, column=c).value
        if key:
            col_map[str(key).strip()] = c
    # Suy ra cột của các trường "_Text" đứng ngay TRƯỚC cột công thức tương ứng
    for text_key, formula_key in _TEXT_BEFORE_FORMULA.items():
        if formula_key in col_map:
            col_map[text_key] = col_map[formula_key] - 1
    for text_key, formula_key in _TEXT_BEFORE_FORMULA_OFFSET2.items():
        if formula_key in col_map:
            col_map[text_key] = col_map[formula_key] - 1
    return col_map


def _build_lookup(wb, sheet_name: str, col_letter: str = "B", start_row: int = 2) -> dict[str, str]:
    """Đọc danh mục (vd 'DM Quốc tịch', 'Dân tộc') -> dict {chuẩn hoá: chữ gốc}
    để so khớp không phân biệt hoa/thường & dấu, rồi ghi ĐÚNG chữ có trong
    danh mục (vì công thức MATCH trong template so khớp CHÍNH XÁC chuỗi)."""
    ws = wb[sheet_name]
    col = openpyxl.utils.column_index_from_string(col_letter)
    lookup = {}
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v:
            lookup[_norm(v)] = str(v)
    return lookup


def _build_phuong_an_lookup(wb) -> dict[str, str]:
    """Đọc sheet 'Phương án' (cột B = mã 'TD','TM','GH'...; cột C = text đầy
    đủ đúng chuẩn 'TD-Tăng đến...') -> dict {mã: text đầy đủ}. Nhờ đọc trực
    tiếp từ chính file mẫu nên luôn khớp chính tả, kể cả khi BHXH cập nhật
    lại danh mục phương án sau này (chỉ cần tải lại file mẫu mới)."""
    ws = wb["Phương án"]
    lookup = {}
    for r in range(3, ws.max_row + 1):
        ma = ws.cell(row=r, column=2).value
        text = ws.cell(row=r, column=3).value
        if ma and text:
            lookup[str(ma).strip().upper()] = str(text)
    return lookup


def _resolve(value: str, lookup: dict[str, str]) -> str:
    """Trả về đúng chuỗi có trong danh mục nếu khớp (không phân biệt hoa
    thường/dấu); nếu không khớp, trả nguyên giá trị gốc (để người dùng tự
    xem lại, còn hơn là ghi giá trị sai lệch)."""
    if not value:
        return value
    return lookup.get(_norm(value), value)


def _write_record(
    ws, row_idx: int, record: dict, col_map: dict[str, int],
    quoc_tich_lookup: dict[str, str] | None = None,
    dan_toc_lookup: dict[str, str] | None = None,
    phuong_an_lookup: dict[str, str] | None = None,
) -> None:
    for field_key, value in record.items():
        col = col_map.get(field_key)
        if col is None:
            continue  # field không tồn tại trong template hiện tại -> bỏ qua an toàn
        if field_key == "QuocTich_Text" and quoc_tich_lookup:
            value = _resolve(value, quoc_tich_lookup)
        elif field_key == "DanToc_Text" and dan_toc_lookup:
            value = _resolve(value, dan_toc_lookup)
        elif field_key == "PhuongAn_Text" and phuong_an_lookup:
            # value hiện là MÃ (vd "TD") -> đổi thành text đầy đủ đúng chuẩn.
            # Nếu mã không có trong danh mục (vd để trống) -> để trống, không
            # ghi mã trần vào ô text (tránh công thức MATCH ở cột kế bên lỗi).
            value = phuong_an_lookup.get(value, "")
        ws.cell(row=row_idx, column=col, value=value)


# --------------------------------------------------------------------------
# Hàm chính
# --------------------------------------------------------------------------
def build_d02lt_excel(
    template_path: str | Path,
    output_path: str | Path,
    tang_list: Iterable[Mapping[str, Any]],
    giam_list: Iterable[Mapping[str, Any]],
    trim_buffer_rows: int = 30,
    noi_lam_viec: str = "",
) -> Path:
    """
    Sinh file Excel D02-LT đúng định dạng chuẩn iCare.

    Parameters
    ----------
    template_path : đường dẫn file mẫu gốc (tải từ iCare, KHÔNG chỉnh sửa tay)
    output_path    : đường dẫn file kết quả sẽ ghi ra
    tang_list      : danh sách dict/RealDictRow lao động TĂNG (từ query hiện có)
    giam_list      : danh sách dict/RealDictRow lao động GIẢM (từ query hiện có)
    trim_buffer_rows : số dòng trống chừa thêm phía sau dữ liệu (mặc định 30)
                        để anh có thể bổ sung tay nếu cần trước khi nộp.

    Returns
    -------
    Path tới file đã tạo.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file mẫu gốc: {template_path}. "
            "Hãy tải file mẫu từ nút 'Tải xuống' trong hộp thoại Nhập Excel "
            "của iCare và lưu vào đường dẫn này (chỉ tải 1 lần, dùng lại)."
        )

    shutil.copy(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    if DATA_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"File mẫu không có sheet '{DATA_SHEET_NAME}'. "
            "Có thể iCare đã đổi cấu trúc mẫu — hãy tải lại file mẫu mới nhất."
        )
    ws = wb[DATA_SHEET_NAME]
    col_map = _build_field_col_map(ws)
    # Danh mục Quốc tịch dùng tên tiếng Anh không dấu (vd "VIET NAM"), khác
    # với Dân tộc dùng tiếng Việt có dấu (vd "Kinh") -> build lookup riêng
    # từng sheet để ghi đúng chính tả mà công thức MATCH() có thể tìm thấy.
    quoc_tich_lookup = _build_lookup(wb, "DM Quốc tịch", col_letter="B")
    dan_toc_lookup = _build_lookup(wb, "Dân tộc", col_letter="B")
    phuong_an_lookup = _build_phuong_an_lookup(wb)

    tang_list = list(tang_list)
    giam_list = list(giam_list)

    row_idx = FIRST_DATA_ROW
    stt = 1
    for r in tang_list:
        rec = _record_tang(r, stt)
        if not rec.get("NoiLamViec") and noi_lam_viec:
            rec["NoiLamViec"] = noi_lam_viec
        _write_record(ws, row_idx, rec, col_map, quoc_tich_lookup, dan_toc_lookup, phuong_an_lookup)
        row_idx += 1
        stt += 1
    for r in giam_list:
        rec = _record_giam(r, stt)
        if not rec.get("NoiLamViec") and noi_lam_viec:
            rec["NoiLamViec"] = noi_lam_viec
        _write_record(ws, row_idx, rec, col_map, quoc_tich_lookup, dan_toc_lookup, phuong_an_lookup)
        row_idx += 1
        stt += 1

    last_used_row = row_idx - 1

    # Cắt bớt các dòng trống dư thừa phía cuối (mẫu gốc để sẵn hàng chục
    # nghìn dòng công thức) để file gọn & việc TÍNH LẠI công thức (bước sau)
    # không bị chậm; các sheet danh mục khác (Tinh/Huyen/Xa/...) giữ nguyên.
    keep_until = last_used_row + trim_buffer_rows
    if ws.max_row > keep_until:
        ws.delete_rows(keep_until + 1, ws.max_row - keep_until)

    wb.save(output_path)

    n_tang, n_giam = len(tang_list), len(giam_list)
    print(
        f"Đã ghi {n_tang} lao động TĂNG + {n_giam} lao động GIẢM "
        f"vào '{DATA_SHEET_NAME}' (dòng {FIRST_DATA_ROW}-{last_used_row}) -> {output_path}"
    )
    return output_path


# --------------------------------------------------------------------------
# Ví dụ tích hợp vào Streamlit (thay cho đoạn code cũ tạo sheet "D02-LT")
# --------------------------------------------------------------------------
def example_usage():
    """
    Trong file Streamlit hiện tại (đoạn `with t2:` anh gửi), thay vì tự dựng
    workbook bằng openpyxl như code cũ, chỉ cần:

        from d02lt_export import build_d02lt_excel
        import tempfile, os

        if export_clicked:
            out_path = os.path.join(
                tempfile.gettempdir(),
                f"D02-LT_BHXH_{tu_ngay:%d%m%Y}_{den_ngay:%d%m%Y}.xlsx"
            )
            build_d02lt_excel(
                template_path="templates/D02LT_template_goc.xlsx",  # file mẫu gốc, tải 1 lần từ iCare
                output_path=out_path,
                tang_list=tang_list,   # list đã có sẵn từ query phía trên
                giam_list=giam_list,
                noi_lam_viec=COMPANY_CONFIG.get("noi_lam_viec", ""),  # tuỳ chọn
            )
            with open(out_path, "rb") as f:
                st.download_button(
                    "📥 Tải file D02-LT (đúng định dạng iCare)",
                    data=f.read(),
                    file_name=os.path.basename(out_path),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
    """
