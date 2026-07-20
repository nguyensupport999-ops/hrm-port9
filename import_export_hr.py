"""
import_export_hr.py
====================
Chức năng: xuất file Excel mẫu (nhiều sheet, đúng thứ tự phụ thuộc) để admin
điền dữ liệu khi chuyển đổi từ hệ thống/file cũ sang HRM Master, và nhập
ngược dữ liệu từ file Excel đó vào database của tenant.

Cách dùng trong app.py:
    from import_export_hr import render_import_export_ui
    render_import_export_ui(lambda: st.session_state.db_engine.get_connection())

Nguyên tắc quan trọng (đã thống nhất với khách hàng):
- Khi nhập liệu, nếu có BẤT KỲ lỗi nào (thiếu bắt buộc, sai định dạng, trùng
  dữ liệu, không tìm thấy dữ liệu tham chiếu...) -> DỪNG LẠI TOÀN BỘ, KHÔNG
  ghi bất kỳ dòng nào vào database, hiển thị đầy đủ danh sách lỗi (kèm tên
  sheet + số dòng Excel) để admin tự sửa file rồi tải lên lại.
- Vì vậy toàn bộ quá trình nhập chạy trong 1 transaction DB duy nhất:
  có lỗi -> rollback toàn bộ; không lỗi -> commit toàn bộ 1 lần ở cuối.
"""

import io
import datetime
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

try:
    import streamlit as st
except ImportError:
    st = None  # cho phép import module này để unit-test ngoài Streamlit


# ============================================================
# 1. ĐỊNH NGHĨA CẤU TRÚC CÁC SHEET (THEO ĐÚNG THỨ TỰ PHỤ THUỘC)
# ============================================================
# Mỗi cột: (db_col, tieu_de_excel, bat_buoc, loai, enum_values_hoac_None, vi_du)
# loai in {'text','int','decimal','date','enum'}

TABLES = [
    {
        "key": "phong_ban",
        "table": "danh_muc_phong_ban",
        "sheet_title": "1. Phòng ban",
        "unique_cols": ["ten_phong_ban"],
        "columns": [
            ("ten_phong_ban", "Tên phòng ban *", True, "text", None, "Kinh doanh"),
            ("thu_tu", "Thứ tự hiển thị", False, "int", None, "1"),
            ("trang_thai", "Trạng thái", False, "enum", ["active", "inactive"], "active"),
        ],
    },
    {
        "key": "chuc_vu",
        "table": "chuc_vu_danh_muc",
        "sheet_title": "2. Chức vụ",
        "unique_cols": ["ten_chuc_vu"],
        "columns": [
            ("ten_chuc_vu", "Tên chức vụ *", True, "text", None, "Trưởng phòng"),
            ("thu_tu", "Thứ tự hiển thị", False, "int", None, "1"),
            ("trang_thai", "Trạng thái", False, "enum", ["Hoạt động", "Ngừng"], "Hoạt động"),
        ],
    },
    {
        "key": "vi_tri",
        "table": "vi_tri_cong_tac",
        "sheet_title": "3. Vị trí công tác",
        "unique_cols": ["ten_vi_tri"],
        "columns": [
            ("ten_vi_tri", "Tên vị trí công tác *", True, "text", None, "Nhân viên kinh doanh"),
            ("so_luong_can_tuyen", "Số lượng cần tuyển", False, "int", None, "2"),
            ("phu_cap_tang_ca", "Phụ cấp tăng ca (đ)", False, "int", None, "0"),
            ("phu_cap_ca_dem", "Phụ cấp ca đêm (đ)", False, "int", None, "0"),
            ("ghi_chu", "Ghi chú", False, "text", None, ""),
        ],
    },
    {
        "key": "trinh_do",
        "table": "danh_muc_trinh_do_hoc_van",
        "sheet_title": "4. Trình độ học vấn",
        "unique_cols": ["ten_trinh_do"],
        "columns": [
            ("ten_trinh_do", "Tên trình độ *", True, "text", None, "Cử nhân"),
            ("thu_tu", "Thứ tự hiển thị", False, "int", None, "1"),
            ("trang_thai", "Trạng thái", False, "enum", ["active", "inactive"], "active"),
        ],
    },
    {
        "key": "nhan_vien",
        "table": "nhan_vien",
        "sheet_title": "5. Nhân viên",
        "unique_cols": ["dien_thoai"],
        "lookup": {
            # excel_col        -> (bảng tham chiếu,      cột hiển thị,  cột db cần điền)
            "ten_vi_tri_ref": ("vi_tri_cong_tac", "ten_vi_tri", "vi_tri_id"),
        },
        "columns": [
            ("ma_nv", "Mã nhân viên *", True, "text", None, "NV001"),
            ("ho_ten", "Họ tên *", True, "text", None, "Nguyễn Văn A"),
            ("dien_thoai", "Số điện thoại *", True, "text", None, "0901234567"),
            ("gioi_tinh", "Giới tính", False, "enum", ["Nam", "Nữ", "Khác"], "Nam"),
            ("ngay_sinh", "Ngày sinh", False, "date", None, "1990-01-15"),
            ("tinh_trang_hon_nhan", "Tình trạng hôn nhân", False, "text", None, "Độc thân"),
            ("so_cccd", "Số CCCD", False, "text", None, "038090012345"),
            ("ngay_cap_cccd", "Ngày cấp CCCD", False, "date", None, "2021-05-01"),
            ("noi_cap_cccd", "Nơi cấp CCCD", False, "text", None, "Cục CS QLHC về TTXH"),
            ("nguyen_quan", "Nguyên quán", False, "text", None, ""),
            ("thuong_tru", "Địa chỉ thường trú", False, "text", None, ""),
            ("email", "Email cá nhân", False, "text", None, "a@gmail.com"),
            ("email_lien_he", "Email công việc", False, "text", None, "a@congty.com"),
            ("quoc_tich", "Quốc tịch", False, "text", None, "Việt Nam"),
            ("dan_toc", "Dân tộc", False, "text", None, "Kinh"),
            ("phong_ban_lam_viec", "Phòng ban làm việc", False, "text", None, "Kinh doanh"),
            ("chuc_vu", "Chức vụ", False, "text", None, "Trưởng phòng"),
            ("ten_vi_tri_ref", "Vị trí công tác (theo sheet 3)", False, "text", None, "Nhân viên kinh doanh"),
            ("trinh_do", "Trình độ học vấn", False, "text", None, "Cử nhân"),
            ("ngay_vao_lam", "Ngày vào làm", False, "date", None, "2023-03-01"),
            ("so_hdld", "Số hợp đồng lao động", False, "text", None, "HD-001/2023"),
            ("ngay_ky_hd", "Ngày ký hợp đồng", False, "date", None, "2023-03-01"),
            ("loai_hop_dong", "Loại hợp đồng", False, "enum",
             ["Thử việc", "Xác định thời hạn", "Không xác định thời hạn"], "Xác định thời hạn"),
            ("thoi_han_hd", "Thời hạn hợp đồng", False, "text", None, "12 tháng"),
            ("he_so_luong", "Hệ số lương", False, "decimal", None, "2.34"),
            ("trang_thai", "Trạng thái làm việc", False, "enum",
             ["DANG_LAM", "THU_VIEC", "NGHI_VIEC"], "DANG_LAM"),
            ("ghi_chu", "Ghi chú", False, "text", None, ""),
        ],
    },
    {
        "key": "lich_su_cong_tac",
        "table": "lich_su_cong_tac",
        "sheet_title": "6. Lịch sử công tác",
        "unique_cols": [],
        "lookup": {
            "dien_thoai_nv": ("nhan_vien", "dien_thoai", "nhan_vien_id"),
        },
        "columns": [
            ("dien_thoai_nv", "SĐT nhân viên (theo sheet 5) *", True, "text", None, "0901234567"),
            ("tu_ngay", "Từ ngày *", True, "date", None, "2023-03-01"),
            ("den_ngay", "Đến ngày", False, "date", None, ""),
            ("chuc_danh", "Chức danh", False, "text", None, "Nhân viên kinh doanh"),
            ("phong_ban", "Phòng ban", False, "text", None, "Kinh doanh"),
            ("noi_lam_viec", "Nơi làm việc", False, "text", None, "Trụ sở chính"),
            ("loai_hop_dong", "Loại hợp đồng", False, "enum",
             ["Thử việc", "Xác định thời hạn", "Không xác định thời hạn"], "Xác định thời hạn"),
            ("he_so_luong", "Hệ số lương", False, "decimal", None, "2.34"),
            ("so_hop_dong", "Số hợp đồng", False, "text", None, "HD-001/2023"),
            ("ghi_chu", "Ghi chú", False, "text", None, ""),
        ],
    },
    {
        "key": "quyet_dinh_nhan_su",
        "table": "quyet_dinh_nhan_su",
        "sheet_title": "7. Quyết định nhân sự",
        "unique_cols": [],
        "lookup": {
            "dien_thoai_nv": ("nhan_vien", "dien_thoai", "nhan_vien_id"),
        },
        "columns": [
            ("dien_thoai_nv", "SĐT nhân viên (theo sheet 5) *", True, "text", None, "0901234567"),
            ("loai_quyet_dinh", "Loại quyết định *", True, "enum",
             ["BO_NHIEM", "MIEN_NHIEM", "DOI_CHUC_DANH", "DIEU_CHUYEN", "CHAM_DUT_HD"], "BO_NHIEM"),
            ("so_quyet_dinh", "Số quyết định", False, "text", None, "QD-001/2023"),
            ("ngay_quyet_dinh", "Ngày quyết định *", True, "date", None, "2023-03-01"),
            ("ngay_hieu_luc", "Ngày hiệu lực *", True, "date", None, "2023-03-01"),
            ("nguoi_ky", "Người ký", False, "text", None, "Giám đốc"),
            ("chuc_danh_cu", "Chức danh cũ", False, "text", None, ""),
            ("chuc_danh_moi", "Chức danh mới", False, "text", None, ""),
            ("phong_ban_cu", "Phòng ban cũ", False, "text", None, ""),
            ("phong_ban_moi", "Phòng ban mới", False, "text", None, ""),
            ("so_hd_cu", "Số HĐ cũ", False, "text", None, ""),
            ("so_hd_moi", "Số HĐ mới", False, "text", None, ""),
            ("noi_dung", "Nội dung quyết định", False, "text", None, ""),
            ("ghi_chu", "Ghi chú", False, "text", None, ""),
            ("trang_thai", "Trạng thái", False, "enum",
             ["CO_HIEU_LUC", "HET_HIEU_LUC"], "CO_HIEU_LUC"),
        ],
    },
]

# Các cột được ghi vào CẢ 2 nơi (vì app hiện đang đọc lẫn lộn 2 tên cột khác
# nhau cho cùng 1 dữ liệu — xem lịch sử trao đổi). Không tự ý bỏ cột nào.
QUYET_DINH_ALIAS_COLS = {
    "loai_quyet_dinh": "loai_qd",
    "so_quyet_dinh": "so_qd",
    "ngay_quyet_dinh": "ngay_qd",
}

EXAMPLE_MARK = "VD:"  # đánh dấu dòng ví dụ ở hàng 3, sẽ bị bỏ qua khi import


# ============================================================
# 2. SINH FILE EXCEL MẪU
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXAMPLE_FONT = Font(italic=True, color="808080")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _write_instruction_sheet(wb):
    ws = wb.active
    ws.title = "HƯỚNG DẪN"
    lines = [
        ("HƯỚNG DẪN NHẬP DỮ LIỆU TỪ FILE EXCEL NÀY", True),
        ("", False),
        ("1. File gồm nhiều sheet, ĐÁNH SỐ THEO ĐÚNG THỨ TỰ phải điền "
         "(sheet 1 → sheet 7). Các sheet sau tham chiếu tới sheet trước "
         "(ví dụ sheet Nhân viên cần Phòng ban/Vị trí đã có ở sheet 1, 3).", False),
        ("2. Dòng có chữ 'VD:' ở đầu là DÒNG VÍ DỤ MẪU — hãy XOÁ dòng này "
         "trước khi nhập dữ liệu thật (hoặc để nguyên, hệ thống sẽ tự bỏ "
         "qua dòng này khi nhập).", False),
        ("3. Cột có dấu (*) ở tiêu đề là BẮT BUỘC phải điền.", False),
        ("4. Các cột có sẵn danh sách chọn (dropdown) — vui lòng chọn đúng "
         "1 trong các giá trị được liệt kê, không tự gõ giá trị khác.", False),
        ("5. Cột ngày tháng nhập theo định dạng YYYY-MM-DD (ví dụ 2023-03-01) "
         "hoặc chọn ngày trực tiếp trong Excel.", False),
        ("6. QUAN TRỌNG: nếu có BẤT KỲ dòng nào bị lỗi (thiếu bắt buộc, sai "
         "định dạng, trùng số điện thoại/tên đã có...), hệ thống sẽ DỪNG LẠI "
         "TOÀN BỘ và KHÔNG nhập bất kỳ dữ liệu nào — sẽ hiện danh sách lỗi "
         "chi tiết theo từng sheet/dòng để bạn sửa lại rồi tải lên lần nữa.", False),
        ("7. Cột 'SĐT nhân viên' ở các sheet Lịch sử công tác / Quyết định "
         "nhân sự phải khớp CHÍNH XÁC với số điện thoại đã điền ở sheet "
         "Nhân viên (hoặc đã có sẵn trong hệ thống).", False),
    ]
    ws.column_dimensions["A"].width = 100
    for i, (text, bold) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            c.font = Font(bold=True, size=14)
    ws.sheet_view.showGridLines = False


def generate_template_workbook() -> io.BytesIO:
    wb = Workbook()
    _write_instruction_sheet(wb)

    for tbl in TABLES:
        ws = wb.create_sheet(tbl["sheet_title"])
        cols = tbl["columns"]

        # Header (dòng 1)
        for ci, (db_col, header, required, kind, enum_vals, example) in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = max(18, len(header) + 2)

            # Dòng 2: ví dụ mẫu
            ex_val = f"{EXAMPLE_MARK} {example}" if example else EXAMPLE_MARK
            ex_cell = ws.cell(row=2, column=ci, value=ex_val)
            ex_cell.font = EXAMPLE_FONT

            # Dropdown cho cột enum (áp dụng từ dòng 3 tới dòng 500)
            if kind == "enum" and enum_vals:
                dv = DataValidation(
                    type="list",
                    formula1='"' + ",".join(enum_vals) + '"',
                    allow_blank=not required,
                )
                dv.error = "Vui lòng chọn 1 giá trị trong danh sách."
                dv.prompt = "Chọn 1 giá trị"
                ws.add_data_validation(dv)
                col_letter = get_column_letter(ci)
                dv.add(f"{col_letter}3:{col_letter}500")

        ws.freeze_panes = "A3"
        ws.sheet_view.showGridLines = True

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ============================================================
# 3. ĐỌC FILE EXCEL NGƯỜI DÙNG TẢI LÊN
# ============================================================
def _read_sheet_rows(ws, columns):
    """Đọc các dòng dữ liệu thật (bỏ header dòng 1, bỏ dòng ví dụ dòng 2,
    bỏ các dòng trống hoàn toàn). Trả về list[(excel_row_number, {db_col: value})]."""
    rows = []
    for r in range(3, ws.max_row + 1):
        values = [ws.cell(row=r, column=ci).value for ci in range(1, len(columns) + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        row_dict = {}
        for (db_col, header, required, kind, enum_vals, example), v in zip(columns, values):
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    v = None
            row_dict[db_col] = v
        rows.append((r, row_dict))
    return rows


def _convert_value(v, kind, errors, sheet_title, excel_row, header):
    if v is None or v == "":
        return None
    try:
        if kind == "int":
            return int(float(v))
        if kind == "decimal":
            return float(v)
        if kind == "date":
            if isinstance(v, (datetime.date, datetime.datetime)):
                return v.date() if isinstance(v, datetime.datetime) else v
            return datetime.datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        return str(v)
    except Exception:
        errors.append(f"[{sheet_title} - dòng {excel_row}] Cột '{header}': giá trị "
                       f"'{v}' không đúng định dạng ({kind}).")
        return None


def import_excel(file_bytes, get_connection):
    """Đọc & nhập dữ liệu từ file Excel. Trả về (ok: bool, errors: list[str], summary: dict).

    get_connection: hàm không tham số, trả về 1 psycopg2 connection MỚI (giống
    st.session_state.db_engine.get_connection()).
    """
    from openpyxl import load_workbook

    errors = []
    summary = {}

    try:
        wb = load_workbook(file_bytes, data_only=True)
    except Exception as e:
        return False, [f"Không đọc được file Excel: {e}"], {}

    # Đọc toàn bộ dữ liệu thô trước (chưa đụng DB)
    parsed = {}
    for tbl in TABLES:
        if tbl["sheet_title"] not in wb.sheetnames:
            errors.append(f"Thiếu sheet bắt buộc: '{tbl['sheet_title']}'. "
                           f"Vui lòng dùng đúng file mẫu đã tải, không đổi tên sheet.")
            continue
        ws = wb[tbl["sheet_title"]]
        col_defs = {c[0]: c for c in tbl["columns"]}
        raw_rows = _read_sheet_rows(ws, tbl["columns"])
        converted_rows = []
        for excel_row, row_dict in raw_rows:
            conv = {}
            for db_col, v in row_dict.items():
                db_col_o, header, required, kind, enum_vals, example = col_defs[db_col]
                if required and (v is None or str(v).strip() == ""):
                    errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] "
                                   f"Cột '{header}' là bắt buộc, đang để trống.")
                cv = _convert_value(v, kind, errors, tbl["sheet_title"], excel_row, header)
                if kind == "enum" and cv is not None and enum_vals and cv not in enum_vals:
                    errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] Cột '{header}': "
                                   f"giá trị '{cv}' không hợp lệ. Chỉ chấp nhận: {', '.join(enum_vals)}.")
                conv[db_col] = cv
            converted_rows.append((excel_row, conv))
        parsed[tbl["key"]] = converted_rows

    if errors:
        return False, errors, {}

    # ---- Bắt đầu transaction: kiểm tra trùng + tham chiếu + insert ----
    conn = get_connection()
    try:
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        name_to_id_cache = {}  # (table, col) -> {lower(value): id}

        def existing_map(table, col):
            key = (table, col)
            if key not in name_to_id_cache:
                cur.execute(f"SELECT id, {col} FROM {table}")
                name_to_id_cache[key] = {
                    (row[col] or "").strip().lower(): row["id"] for row in cur.fetchall()
                }
            return name_to_id_cache[key]

        for tbl in TABLES:
            if tbl["key"] not in parsed:
                continue
            rows = parsed[tbl["key"]]
            if not rows:
                continue

            col_defs = {c[0]: c for c in tbl["columns"]}
            unique_cols = tbl.get("unique_cols", [])
            lookup = tbl.get("lookup", {})
            table = tbl["table"]

            seen_in_file = {c: set() for c in unique_cols}
            existing_unique = {c: existing_map(table, c) for c in unique_cols}

            inserted = 0
            for excel_row, row in rows:
                # 1) kiểm tra trùng (trong file + đã có trong DB)
                dup = False
                for c in unique_cols:
                    v = (row.get(c) or "").strip().lower() if row.get(c) else None
                    if not v:
                        continue
                    if v in seen_in_file[c]:
                        errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] "
                                       f"Giá trị '{row.get(c)}' ở cột trùng với 1 dòng khác "
                                       f"TRONG CHÍNH FILE này.")
                        dup = True
                    elif v in existing_unique[c]:
                        errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] "
                                       f"Giá trị '{row.get(c)}' đã TỒN TẠI SẴN trong hệ thống.")
                        dup = True
                    else:
                        seen_in_file[c].add(v)
                if dup:
                    continue

                # 2) resolve các cột lookup (tham chiếu sheet khác)
                insert_row = {k: v for k, v in row.items() if k not in lookup}
                skip_row = False
                for excel_col, (ref_table, ref_col, target_col) in lookup.items():
                    v = row.get(excel_col)
                    if v is None or str(v).strip() == "":
                        continue
                    ref_map = existing_map(ref_table, ref_col)
                    key = str(v).strip().lower()
                    if key not in ref_map:
                        errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] "
                                       f"Không tìm thấy '{v}' trong sheet/dữ liệu tương ứng "
                                       f"({ref_table}.{ref_col}). Hãy thêm trước ở sheet đó, "
                                       f"hoặc kiểm tra lại chính tả.")
                        skip_row = True
                    else:
                        insert_row[target_col] = ref_map[key]
                if skip_row:
                    continue

                # 3) ghi thêm cột "alias" (áp dụng riêng cho quyet_dinh_nhan_su)
                if table == "quyet_dinh_nhan_su":
                    for src, dst in QUYET_DINH_ALIAS_COLS.items():
                        if src in insert_row:
                            insert_row[dst] = insert_row[src]

                # 4) INSERT
                cols_list = list(insert_row.keys())
                placeholders = ", ".join(["%s"] * len(cols_list))
                col_names = ", ".join(cols_list)
                try:
                    cur.execute(
                        f"INSERT INTO {table} ({col_names}) VALUES ({placeholders}) RETURNING id",
                        [insert_row[c] for c in cols_list],
                    )
                    new_id = cur.fetchone()["id"]
                    inserted += 1
                    # cập nhật cache để các sheet sau (vd nhân viên) tham
                    # chiếu được ngay tới dòng vừa thêm trong CÙNG lần import
                    for c in unique_cols:
                        v = row.get(c)
                        if v:
                            name_to_id_cache.setdefault((table, c), {})[str(v).strip().lower()] = new_id
                except Exception as e:
                    errors.append(f"[{tbl['sheet_title']} - dòng {excel_row}] "
                                   f"Lỗi khi ghi vào database: {e}")

            summary[tbl["sheet_title"]] = inserted

        if errors:
            conn.rollback()
            return False, errors, {}

        conn.commit()
        return True, [], summary

    except Exception as e:
        conn.rollback()
        return False, [f"Lỗi hệ thống khi nhập dữ liệu: {e}"], {}
    finally:
        conn.close()


# ============================================================
# 4. GIAO DIỆN STREAMLIT (dùng chung cho admin công ty & Super Admin)
# ============================================================
def render_import_export_ui(get_connection, extra_caption=""):
    """get_connection: hàm không tham số trả về 1 connection MỚI tới đúng
    database của tenant cần thao tác (dùng lại được nhiều lần trong 1 lượt gọi)."""
    st.subheader("📥 Nhập / 📤 Xuất dữ liệu Excel")
    if extra_caption:
        st.caption(extra_caption)

    tab1, tab2 = st.tabs(["📤 Tải mẫu Excel", "📥 Nhập dữ liệu từ Excel"])

    with tab1:
        st.write("Tải file Excel mẫu (đã có sẵn hướng dẫn, danh sách chọn, "
                 "và đúng thứ tự các sheet cần điền) để chuyển đổi dữ liệu "
                 "từ hệ thống/file cũ sang HRM Master.")
        buf = generate_template_workbook()
        st.download_button(
            "⬇️ Tải file mẫu Excel (mau_nhap_lieu_hrm.xlsx)",
            data=buf,
            file_name="mau_nhap_lieu_hrm.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with tab2:
        st.write("Sau khi điền xong file mẫu, tải lên đây để nhập vào hệ thống.")
        st.warning("⚠️ Nếu có BẤT KỲ dòng nào bị lỗi, hệ thống sẽ **không nhập bất kỳ "
                   "dữ liệu nào** — vui lòng sửa hết lỗi rồi tải lên lại.")
        up = st.file_uploader("Chọn file Excel đã điền", type=["xlsx"])
        if up is not None and st.button("🔍 Kiểm tra & Nhập dữ liệu", type="primary"):
            with st.spinner("Đang kiểm tra và nhập dữ liệu..."):
                ok, errors, summary = import_excel(up, get_connection)
            if ok:
                st.success("✅ Nhập dữ liệu thành công!")
                for sheet, count in summary.items():
                    st.write(f"- {sheet}: đã thêm **{count}** dòng")
                st.cache_data.clear()
            else:
                st.error(f"❌ Phát hiện {len(errors)} lỗi — CHƯA nhập bất kỳ dữ liệu nào. "
                         f"Vui lòng sửa file rồi tải lên lại:")
                for e in errors[:200]:
                    st.write("- " + e)
                if len(errors) > 200:
                    st.caption(f"... và {len(errors) - 200} lỗi khác.")
