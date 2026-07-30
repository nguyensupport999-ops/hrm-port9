# -*- coding: utf-8 -*-
"""
thue_hkd.py
============
Module quản lý Thuế cho Hộ kinh doanh (HKD).
Gồm 3 tab chính:
  - ⚙️ Cấu hình HKD: thông tin hộ, ngành nghề, tỷ lệ thuế, BHXH chủ hộ
  - 📊 Theo dõi Doanh thu & Thuế: nhập DT theo kỳ, tự động tính thuế
  - 📋 Tổng hợp & Cảnh báo: báo cáo lũy kế, nhắc hạn kê khai

Cách dùng trong app.py:
    from thue_hkd import render_thue_hkd
    if menu == "🧾 Thuế HKD":
        render_thue_hkd(db_engine)
"""

import datetime
import streamlit as st
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

# ═══════════════════════════════════════════════════════════════════
#  HẰNG SỐ
# ═══════════════════════════════════════════════════════════════════

NGUONG_MIEN_THUE = 500_000_000       # 500 triệu
NGUONG_NHOM_2 = 3_000_000_000        # 3 tỷ
NGUONG_NHOM_3 = 50_000_000_000_000   # 50 tỷ — dùng cho phân nhóm, thực tế HKD hiếm khi đạt

# Tỷ lệ thuế theo ngành nghề (tham chiếu Thông tư 40/2021/TT-BTC)
TY_LE_THUE_THEO_NGANH = {
    "THUONG_MAI":   {"ten": "Phân phối, cung cấp hàng hóa",                       "gtgt": 1.0, "tncn": 0.5},
    "DICH_VU":      {"ten": "Dịch vụ, xây dựng không bao thầu NVL",               "gtgt": 5.0, "tncn": 2.0},
    "SAN_XUAT":     {"ten": "Sản xuất, vận tải, dịch vụ gắn hàng hóa, XD có NVL", "gtgt": 3.0, "tncn": 1.5},
    "CHO_THUE":     {"ten": "Cho thuê tài sản, BĐS; sản phẩm/dịch vụ nội dung số","gtgt": 5.0, "tncn": 5.0},
    "DAI_LY":       {"ten": "Đại lý bảo hiểm, xổ số, bán hàng đa cấp",           "gtgt": 0.0, "tncn": 5.0},
    "KHAC":         {"ten": "Hoạt động kinh doanh khác",                           "gtgt": 2.0, "tncn": 1.0},
}

THUE_SUAT_LOI_NHUAN = {
    "500TR_3TY":  15.0,
    "3TY_50TY":   17.0,
    "TREN_50TY":  20.0,
}

MUC_THAM_CHIEU_BHXH = 2_340_000  # Mức tham chiếu BHXH 2025-2026 (= lương cơ sở)


# ═══════════════════════════════════════════════════════════════════
#  HÀM TIỆN ÍCH
# ═══════════════════════════════════════════════════════════════════

def _can_edit():
    """Kiểm tra quyền chỉnh sửa — tái sử dụng logic can_edit() từ app.py."""
    role = st.session_state.get("role", "")
    if role in ("xem_toan_bo", "demo_readonly", "viewer"):
        return False
    return True


def _fmt_tien(so):
    """Định dạng số tiền VND: 1,234,567"""
    if so is None:
        return "0"
    return f"{int(so):,}".replace(",", ".")


def _fmt_tien_trieu(so):
    """Hiển thị số tiền dạng triệu: 500.000.000 → 500 triệu"""
    if so is None or so == 0:
        return "0"
    if so >= 1_000_000_000:
        return f"{so / 1_000_000_000:.1f} tỷ"
    if so >= 1_000_000:
        return f"{so / 1_000_000:.0f} triệu"
    return _fmt_tien(so)


def _phan_nhom_doanh_thu(doanh_thu_nam):
    """Phân nhóm HKD theo doanh thu lũy kế năm."""
    if doanh_thu_nam <= NGUONG_MIEN_THUE:
        return "MIEN_THUE"
    elif doanh_thu_nam <= NGUONG_NHOM_2:
        return "500TR_3TY"
    elif doanh_thu_nam <= NGUONG_NHOM_3:
        return "3TY_50TY"
    else:
        return "TREN_50TY"


def _ten_nhom_doanh_thu(nhom):
    MAP = {
        "MIEN_THUE": "≤ 500 triệu — Miễn thuế",
        "500TR_3TY": "500 triệu – 3 tỷ",
        "3TY_50TY":  "3 tỷ – 50 tỷ",
        "TREN_50TY": "> 50 tỷ",
    }
    return MAP.get(nhom, nhom)


def _tinh_thue_ty_le_doanh_thu(doanh_thu, ty_le_gtgt, ty_le_tncn):
    """Tính thuế theo tỷ lệ % trên doanh thu (sau khi trừ 500 triệu)."""
    phan_chiu_thue = max(0, doanh_thu - NGUONG_MIEN_THUE)
    thue_gtgt = round(phan_chiu_thue * ty_le_gtgt / 100)
    thue_tncn = round(phan_chiu_thue * ty_le_tncn / 100)
    return thue_gtgt, thue_tncn


def _tinh_thue_loi_nhuan(doanh_thu, chi_phi, nhom):
    """Tính thuế theo lợi nhuận (DT - CP) × thuế suất."""
    loi_nhuan = max(0, doanh_thu - chi_phi)
    thue_suat = THUE_SUAT_LOI_NHUAN.get(nhom, 15.0)
    # Thuế TNCN = lợi nhuận × thuế suất
    thue_tncn = round(loi_nhuan * thue_suat / 100)
    # Thuế GTGT: vẫn tính theo tỷ lệ trên doanh thu (không đổi)
    return thue_tncn


def _ky_ke_khai_options(nam=None):
    """Tạo danh sách kỳ kê khai cho 1 năm."""
    if nam is None:
        nam = datetime.date.today().year
    return [f"Q1/{nam}", f"Q2/{nam}", f"Q3/{nam}", f"Q4/{nam}", str(nam)]


def _han_ke_khai(ky_thue):
    """Xác định hạn kê khai dựa trên kỳ thuế."""
    if "/" in ky_thue:
        # Kỳ quý: Q1/2026 → hạn 30/4/2026
        quy, nam = ky_thue.split("/")
        nam = int(nam)
        quy_num = int(quy.replace("Q", ""))
        thang_han = quy_num * 3 + 1
        if thang_han > 12:
            thang_han = 1
            nam += 1
        return datetime.date(nam, thang_han, 30 if thang_han in (4, 6, 9, 11) else 31 if thang_han in (1, 3, 5, 7, 8, 10, 12) else 28)
    else:
        # Kỳ năm: 2026 → hạn 31/01/2027
        nam = int(ky_thue)
        return datetime.date(nam + 1, 1, 31)

def _lay_doanh_thu_theo_nganh(conn, ky_thue):
    """Lấy doanh thu từ bảng doanh_thu_hkd, nhóm theo ngành nghề.
    ky_thue = 'Q1/2026' → lấy tháng 01,02,03/2026
    ky_thue = '2026' → lấy cả năm 2026
    Trả về dict: {nhom_nganh: tong_doanh_thu}
    """
    c = conn.cursor()
    if "/" in ky_thue:
        # Quý: Q1/2026 → tháng 01,02,03
        quy_str, nam_str = ky_thue.split("/")
        quy_num = int(quy_str.replace("Q", ""))
        thang_bat_dau = (quy_num - 1) * 3 + 1
        ds_thang = [f"{m:02d}/{nam_str}" for m in range(thang_bat_dau, thang_bat_dau + 3)]
        placeholders = ",".join(["%s"] * len(ds_thang))
        c.execute(f"""
            SELECT nhom_nganh, COALESCE(SUM(doanh_thu), 0) as tong
            FROM doanh_thu_hkd
            WHERE thang IN ({placeholders})
            GROUP BY nhom_nganh
        """, ds_thang)
    else:
        # Cả năm
        c.execute("""
            SELECT nhom_nganh, COALESCE(SUM(doanh_thu), 0) as tong
            FROM doanh_thu_hkd
            WHERE thang LIKE %s
            GROUP BY nhom_nganh
        """, (f"%/{ky_thue}",))

    result = {}
    for row in c.fetchall():
        result[row[0]] = float(row[1])
    return result


def _lay_chi_phi_ky(conn, ky_thue):
    """Lấy tổng chi phí được trừ thuế từ bảng chi_phi_hkd theo kỳ.
    Trả về (tong_chi_phi, tong_duoc_tru)
    """
    c = conn.cursor()
    if "/" in ky_thue:
        quy_str, nam_str = ky_thue.split("/")
        quy_num = int(quy_str.replace("Q", ""))
        thang_bat_dau = (quy_num - 1) * 3 + 1
        ds_thang = [f"{m:02d}/{nam_str}" for m in range(thang_bat_dau, thang_bat_dau + 3)]
        placeholders = ",".join(["%s"] * len(ds_thang))
        c.execute(f"""
            SELECT COALESCE(SUM(so_tien), 0),
                   COALESCE(SUM(CASE WHEN duoc_tru_thue THEN so_tien ELSE 0 END), 0)
            FROM chi_phi_hkd
            WHERE thang IN ({placeholders})
        """, ds_thang)
    else:
        c.execute("""
            SELECT COALESCE(SUM(so_tien), 0),
                   COALESCE(SUM(CASE WHEN duoc_tru_thue THEN so_tien ELSE 0 END), 0)
            FROM chi_phi_hkd
            WHERE thang LIKE %s
        """, (f"%/{ky_thue}",))

    row = c.fetchone()
    return (float(row[0]), float(row[1])) if row else (0, 0)


def _tinh_thue_chi_tiet_theo_nganh(dt_theo_nganh):
    """Tính thuế GTGT + TNCN chi tiết theo từng ngành nghề.
    Trả về dict: {nhom_nganh: {doanh_thu, thue_gtgt, thue_tncn, ty_le_gtgt, ty_le_tncn}}
    """
    result = {}
    tong_gtgt = 0
    tong_tncn = 0
    tong_dt = 0
    for nganh, dt in dt_theo_nganh.items():
        if dt <= 0:
            continue
        info = TY_LE_THUE_THEO_NGANH.get(nganh, TY_LE_THUE_THEO_NGANH["KHAC"])
        gtgt = round(dt * info["gtgt"] / 100)
        tncn = round(dt * info["tncn"] / 100)
        result[nganh] = {
            "ten": info["ten"],
            "doanh_thu": dt,
            "ty_le_gtgt": info["gtgt"],
            "ty_le_tncn": info["tncn"],
            "thue_gtgt": gtgt,
            "thue_tncn": tncn,
        }
        tong_gtgt += gtgt
        tong_tncn += tncn
        tong_dt += dt
    return result, tong_dt, tong_gtgt, tong_tncn

# ═══════════════════════════════════════════════════════════════════
#  DB HELPERS
# ═══════════════════════════════════════════════════════════════════

def _get_cau_hinh(conn):
    """Lấy cấu hình HKD (1 dòng duy nhất). Trả None nếu chưa có."""
    try:
        c = conn.cursor()
        c.execute("SELECT * FROM cau_hinh_thue_hkd ORDER BY id LIMIT 1")
        cols = [desc[0] for desc in c.description]
        row = c.fetchone()
        if row:
            return dict(zip(cols, row))
    except Exception:
        pass
    return None


def _luu_cau_hinh(conn, data):
    """Insert hoặc Update cấu hình HKD."""
    c = conn.cursor()
    # Kiểm tra đã có chưa
    c.execute("SELECT id FROM cau_hinh_thue_hkd LIMIT 1")
    existing = c.fetchone()

    if existing:
        c.execute("""
            UPDATE cau_hinh_thue_hkd SET
                ten_hkd = %s, dia_chi = %s, so_dkkd = %s, ngay_cap_dkkd = %s,
                co_quan_cap = %s, chu_ho_ten = %s, chu_ho_cccd = %s,
                nganh_nghe = %s, ty_le_thue_gtgt = %s, ty_le_thue_tncn = %s,
                phuong_phap_tinh_thue = %s, ky_ke_khai = %s,
                chu_ho_dong_bhxh = %s, muc_luong_dong_bhxh_chu_ho = %s,
                phuong_thuc_dong_bhxh = %s,
                updated_at = now()
            WHERE id = %s
        """, (
            data["ten_hkd"], data["dia_chi"], data["so_dkkd"], data["ngay_cap_dkkd"],
            data["co_quan_cap"], data["chu_ho_ten"], data["chu_ho_cccd"],
            data["nganh_nghe"], data["ty_le_thue_gtgt"], data["ty_le_thue_tncn"],
            data["phuong_phap_tinh_thue"], data["ky_ke_khai"],
            data["chu_ho_dong_bhxh"], data["muc_luong_dong_bhxh_chu_ho"],
            data["phuong_thuc_dong_bhxh"],
            existing[0],
        ))
    else:
        c.execute("""
            INSERT INTO cau_hinh_thue_hkd (
                ten_hkd, dia_chi, so_dkkd, ngay_cap_dkkd, co_quan_cap,
                chu_ho_ten, chu_ho_cccd,
                nganh_nghe, ty_le_thue_gtgt, ty_le_thue_tncn,
                phuong_phap_tinh_thue, ky_ke_khai,
                chu_ho_dong_bhxh, muc_luong_dong_bhxh_chu_ho, phuong_thuc_dong_bhxh
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["ten_hkd"], data["dia_chi"], data["so_dkkd"], data["ngay_cap_dkkd"],
            data["co_quan_cap"], data["chu_ho_ten"], data["chu_ho_cccd"],
            data["nganh_nghe"], data["ty_le_thue_gtgt"], data["ty_le_thue_tncn"],
            data["phuong_phap_tinh_thue"], data["ky_ke_khai"],
            data["chu_ho_dong_bhxh"], data["muc_luong_dong_bhxh_chu_ho"],
            data["phuong_thuc_dong_bhxh"],
        ))
    conn.commit()


def _get_ds_thue(conn, nam=None):
    """Lấy danh sách theo dõi thuế của 1 năm."""
    c = conn.cursor()
    if nam:
        c.execute("""SELECT * FROM theo_doi_thue_hkd
                     WHERE ky_thue LIKE %s OR ky_thue = %s
                     ORDER BY id""", (f"%/{nam}", str(nam)))
    else:
        c.execute("SELECT * FROM theo_doi_thue_hkd ORDER BY id DESC LIMIT 20")
    cols = [desc[0] for desc in c.description]
    rows = c.fetchall()
    return [dict(zip(cols, r)) for r in rows]


def _luu_ky_thue(conn, data):
    """Insert hoặc Update 1 kỳ thuế."""
    c = conn.cursor()
    # Kiểm tra kỳ đã tồn tại chưa
    c.execute("SELECT id FROM theo_doi_thue_hkd WHERE ky_thue = %s", (data["ky_thue"],))
    existing = c.fetchone()

    if existing:
        c.execute("""
            UPDATE theo_doi_thue_hkd SET
                doanh_thu_ky = %s, doanh_thu_luy_ke_nam = %s, chi_phi_ky = %s,
                nhom_doanh_thu = %s, thue_gtgt_phai_nop = %s, thue_tncn_phai_nop = %s,
                trang_thai = %s, han_ke_khai = %s, ngay_ke_khai = %s, ngay_nop = %s,
                ghi_chu = %s, updated_at = now()
            WHERE id = %s
        """, (
            data["doanh_thu_ky"], data["doanh_thu_luy_ke_nam"], data["chi_phi_ky"],
            data["nhom_doanh_thu"], data["thue_gtgt_phai_nop"], data["thue_tncn_phai_nop"],
            data["trang_thai"], data["han_ke_khai"], data.get("ngay_ke_khai"),
            data.get("ngay_nop"), data.get("ghi_chu", ""),
            existing[0],
        ))
    else:
        c.execute("""
            INSERT INTO theo_doi_thue_hkd (
                ky_thue, doanh_thu_ky, doanh_thu_luy_ke_nam, chi_phi_ky,
                nhom_doanh_thu, thue_gtgt_phai_nop, thue_tncn_phai_nop,
                trang_thai, han_ke_khai, ngay_ke_khai, ngay_nop, ghi_chu
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            data["ky_thue"], data["doanh_thu_ky"], data["doanh_thu_luy_ke_nam"],
            data["chi_phi_ky"], data["nhom_doanh_thu"],
            data["thue_gtgt_phai_nop"], data["thue_tncn_phai_nop"],
            data["trang_thai"], data["han_ke_khai"], data.get("ngay_ke_khai"),
            data.get("ngay_nop"), data.get("ghi_chu", ""),
        ))
    conn.commit()


def _xoa_ky_thue(conn, ky_thue):
    """Xóa 1 kỳ thuế."""
    c = conn.cursor()
    c.execute("DELETE FROM theo_doi_thue_hkd WHERE ky_thue = %s", (ky_thue,))
    conn.commit()


# ═══════════════════════════════════════════════════════════════════
#  TAB 1: CẤU HÌNH HKD
# ═══════════════════════════════════════════════════════════════════

def _render_tab_cau_hinh(db_engine):
    """Tab cấu hình thông tin Hộ kinh doanh + ngành nghề + BHXH chủ hộ."""
    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if cfg is None:
        cfg = {}

    st.subheader("📝 Thông tin Hộ kinh doanh")

    col1, col2 = st.columns(2)
    with col1:
        ten_hkd = st.text_input("Tên HKD", value=cfg.get("ten_hkd", ""), key="hkd_ten")
        so_dkkd = st.text_input("Số ĐKKD", value=cfg.get("so_dkkd", ""), key="hkd_dkkd")
        chu_ho_ten = st.text_input("Họ tên chủ hộ", value=cfg.get("chu_ho_ten", ""), key="hkd_chuten")
    with col2:
        dia_chi = st.text_input("Địa chỉ", value=cfg.get("dia_chi", ""), key="hkd_diachi")
        co_quan_cap = st.text_input("Cơ quan cấp ĐKKD", value=cfg.get("co_quan_cap", ""), key="hkd_coquan")
        chu_ho_cccd = st.text_input("Số CCCD chủ hộ", value=cfg.get("chu_ho_cccd", ""), key="hkd_cccd")

    ngay_cap_val = cfg.get("ngay_cap_dkkd")
    if ngay_cap_val and isinstance(ngay_cap_val, str):
        try:
            ngay_cap_val = datetime.datetime.strptime(ngay_cap_val, "%Y-%m-%d").date()
        except Exception:
            ngay_cap_val = None
    ngay_cap_dkkd = st.date_input("Ngày cấp ĐKKD", value=ngay_cap_val, key="hkd_ngaycap")

    st.divider()
    st.subheader("🏷️ Ngành nghề & Tỷ lệ thuế")

    nganh_nghe_list = list(TY_LE_THUE_THEO_NGANH.keys())
    nganh_hien_tai = cfg.get("nganh_nghe", "THUONG_MAI")
    idx_nganh = nganh_nghe_list.index(nganh_hien_tai) if nganh_hien_tai in nganh_nghe_list else 0

    nganh_nghe = st.selectbox(
        "Ngành nghề chính",
        nganh_nghe_list,
        index=idx_nganh,
        format_func=lambda k: TY_LE_THUE_THEO_NGANH[k]["ten"],
        key="hkd_nganh",
    )

    # Tỷ lệ thuế auto theo ngành nghề (luật định — không cho sửa tay)
    ty_le_mac_dinh = TY_LE_THUE_THEO_NGANH[nganh_nghe]
    ty_le_gtgt = ty_le_mac_dinh["gtgt"]
    ty_le_tncn = ty_le_mac_dinh["tncn"]
    col_a, col_b = st.columns(2)
    with col_a:
        st.metric("Tỷ lệ thuế GTGT", f"{ty_le_gtgt}%")
    with col_b:
        st.metric("Tỷ lệ thuế TNCN", f"{ty_le_tncn}%")
    st.caption("📌 Tỷ lệ thuế GTGT: Điều 12(2b) Luật Thuế GTGT 48/2024/QH15. Tỷ lệ thuế TNCN: Điều 7(3) Luật Thuế TNCN 109/2025/QH15.")

    # Phương pháp tính thuế: HKD 500tr–3tỷ được chọn 1 trong 2; nhóm khác fix theo luật
    phuong_phap_options = ["TY_LE_DOANH_THU", "LOI_NHUAN"]
    phuong_phap_labels = {
        "TY_LE_DOANH_THU": "Theo tỷ lệ % trên doanh thu (không cần chứng từ chi phí)",
        "LOI_NHUAN": "Theo lợi nhuận: (DT − Chi phí) × thuế suất (cần hóa đơn đầu vào)",
    }
    pp_hien_tai = cfg.get("phuong_phap_tinh_thue", "TY_LE_DOANH_THU")
    idx_pp = phuong_phap_options.index(pp_hien_tai) if pp_hien_tai in phuong_phap_options else 0
    phuong_phap = st.selectbox(
        "Phương pháp tính thuế (HKD DT 500tr–3tỷ được chọn)",
        phuong_phap_options,
        index=idx_pp,
        format_func=lambda k: phuong_phap_labels[k],
        key="hkd_phuongphap",
    )

    # Kỳ kê khai: auto theo luật — DT ≤ 500tr → năm, DT > 500tr → quý
    # Lần đầu admin tự chọn dự kiến; sau đó Tab "Theo dõi" sẽ tự điều chỉnh khi có DT thực tế
    ky_options = ["NAM", "QUY"]
    ky_labels = {"NAM": "Kê khai theo năm (DT ≤ 500 triệu)", "QUY": "Kê khai theo quý (DT > 500 triệu)"}
    ky_hien_tai = cfg.get("ky_ke_khai", "NAM")
    idx_ky = ky_options.index(ky_hien_tai) if ky_hien_tai in ky_options else 0
    ky_ke_khai = st.selectbox(
        "Kỳ kê khai (dự kiến — sẽ tự điều chỉnh theo DT thực tế)",
        ky_options,
        index=idx_ky,
        format_func=lambda k: ky_labels[k],
        key="hkd_kykekhai",
    )

    st.divider()
    st.subheader("🏥 BHXH Chủ hộ kinh doanh")
    st.caption("Chủ HKD đóng BHXH (25%) + BHYT (4,5%) = 29,5%. Không đóng BHTN. Mức lương tự chọn.")

    chu_ho_dong_bhxh = st.checkbox(
        "Chủ hộ tham gia BHXH bắt buộc",
        value=cfg.get("chu_ho_dong_bhxh", True),
        key="hkd_bhxh",
    )

    if chu_ho_dong_bhxh:
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            _ml_hien_tai = int(cfg.get("muc_luong_dong_bhxh_chu_ho", MUC_THAM_CHIEU_BHXH))
            _ml_text = st.text_input(
                f"Mức lương đóng BHXH (tối thiểu {_fmt_tien(MUC_THAM_CHIEU_BHXH)}đ)",
                value=_fmt_tien(_ml_hien_tai),
                key="hkd_mucluong",
            )
            # Parse: bỏ dấu chấm phân hàng → số nguyên
            try:
                muc_luong = int(_ml_text.replace(".", "").replace(",", "").strip())
            except ValueError:
                muc_luong = _ml_hien_tai
                st.caption("⚠️ Nhập số, VD: 2.340.000")
            # Validate
            if muc_luong < MUC_THAM_CHIEU_BHXH:
                st.caption(f"⚠️ Tối thiểu {_fmt_tien(MUC_THAM_CHIEU_BHXH)}đ")
                muc_luong = MUC_THAM_CHIEU_BHXH
            elif muc_luong > MUC_THAM_CHIEU_BHXH * 20:
                st.caption(f"⚠️ Tối đa {_fmt_tien(MUC_THAM_CHIEU_BHXH * 20)}đ")
                muc_luong = MUC_THAM_CHIEU_BHXH * 20
        with col_m2:
            pt_options = ["HANG_THANG", "3_THANG", "6_THANG"]
            pt_labels = {"HANG_THANG": "Hàng tháng", "3_THANG": "3 tháng/lần", "6_THANG": "6 tháng/lần"}
            pt_hien_tai = cfg.get("phuong_thuc_dong_bhxh", "HANG_THANG")
            idx_pt = pt_options.index(pt_hien_tai) if pt_hien_tai in pt_options else 0
            phuong_thuc_dong = st.selectbox(
                "Phương thức đóng",
                pt_options,
                index=idx_pt,
                format_func=lambda k: pt_labels[k],
                key="hkd_ptdong",
            )

        # Hiển thị ước tính
        tien_bhxh = round(muc_luong * 0.25)
        tien_bhyt = round(muc_luong * 0.045)
        tong = tien_bhxh + tien_bhyt
        st.info(f"💰 Ước tính: BHXH {_fmt_tien(tien_bhxh)}đ + BHYT {_fmt_tien(tien_bhyt)}đ = **{_fmt_tien(tong)}đ/tháng**")
    else:
        muc_luong = MUC_THAM_CHIEU_BHXH
        phuong_thuc_dong = "HANG_THANG"

    st.divider()

    # Nút Lưu
    if st.button("💾 Lưu cấu hình", disabled=not _can_edit(), type="primary", key="hkd_luu"):
        data = {
            "ten_hkd": ten_hkd,
            "dia_chi": dia_chi,
            "so_dkkd": so_dkkd,
            "ngay_cap_dkkd": ngay_cap_dkkd,
            "co_quan_cap": co_quan_cap,
            "chu_ho_ten": chu_ho_ten,
            "chu_ho_cccd": chu_ho_cccd,
            "nganh_nghe": nganh_nghe,
            "ty_le_thue_gtgt": ty_le_gtgt,
            "ty_le_thue_tncn": ty_le_tncn,
            "phuong_phap_tinh_thue": phuong_phap,
            "ky_ke_khai": ky_ke_khai,
            "chu_ho_dong_bhxh": chu_ho_dong_bhxh,
            "muc_luong_dong_bhxh_chu_ho": muc_luong,
            "phuong_thuc_dong_bhxh": phuong_thuc_dong,
        }
        try:
            conn = db_engine.get_connection()
            _luu_cau_hinh(conn, data)
            conn.close()
            st.success("✅ Đã lưu cấu hình HKD!")
            st.rerun()
        except Exception as e:
            st.error(f"❌ Lỗi: {e}")


# ═══════════════════════════════════════════════════════════════════
#  TAB 2: THEO DÕI DOANH THU & THUẾ
# ═══════════════════════════════════════════════════════════════════

def _render_tab_theo_doi(db_engine):
    """Tab kê khai thuế theo kỳ — tích hợp DT chi tiết theo ngành từ bảng doanh_thu_hkd."""
    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if not cfg:
        st.warning("⚠️ Chưa có cấu hình HKD. Vui lòng vào **⚙️ Cấu hình HKD** để thiết lập trước.")
        return

    nam_hien_tai = datetime.date.today().year
    nam = st.selectbox("Năm", [nam_hien_tai - 1, nam_hien_tai, nam_hien_tai + 1],
                       index=1, key="hkd_nam_theodoi")

    conn = db_engine.get_connection()
    try:
        ds_thue = _get_ds_thue(conn, nam)
    finally:
        conn.close()

    ky_da_nhap = {d["ky_thue"] for d in ds_thue}

    # ══════════════════════════════════════════════
    # PHẦN 1: Tạo kê khai kỳ mới
    # ══════════════════════════════════════════════
    st.divider()
    st.subheader("➕ Kê khai kỳ thuế mới")

    if cfg.get("ky_ke_khai") == "QUY":
        ky_chon_list = [f"Q{i}/{nam}" for i in range(1, 5)]
    else:
        ky_chon_list = [str(nam)]

    ky_chua_nhap = [k for k in ky_chon_list if k not in ky_da_nhap]

    if not ky_chua_nhap:
        st.info("✅ Đã kê khai đủ các kỳ thuế trong năm này.")
    else:
        ky_thue = st.selectbox("Chọn kỳ thuế", ky_chua_nhap, key="hkd_ky_moi")

        # ── Tự động lấy DT từ bảng doanh_thu_hkd theo kỳ ──
        conn = db_engine.get_connection()
        try:
            dt_theo_nganh = _lay_doanh_thu_theo_nganh(conn, ky_thue)
            chi_phi_ky_tong, chi_phi_duoc_tru = _lay_chi_phi_ky(conn, ky_thue)
        finally:
            conn.close()

        if not dt_theo_nganh:
            st.warning(f"⚠️ Chưa có doanh thu nào trong kỳ **{ky_thue}**. "
                       f"Vui lòng vào **💰 Doanh thu & Chi phí** nhập DT trước.")
            # Cho phép nhập tay fallback
            st.caption("Hoặc nhập tay doanh thu tổng:")
            col_fb1, col_fb2 = st.columns(2)
            with col_fb1:
                _dt_text = st.text_input("Doanh thu kỳ (VNĐ)", value="0", key="hkd_dt_fb")
                try:
                    dt_nhap_tay = int(_dt_text.replace(".", "").replace(",", "").strip())
                except ValueError:
                    dt_nhap_tay = 0
            with col_fb2:
                nganh_fb = st.selectbox("Ngành nghề", list(TY_LE_THUE_THEO_NGANH.keys()),
                                        format_func=lambda k: TY_LE_THUE_THEO_NGANH[k]["ten"],
                                        key="hkd_nganh_fb")
            if dt_nhap_tay > 0:
                dt_theo_nganh = {nganh_fb: dt_nhap_tay}

        # ── Hiển thị bảng DT chi tiết theo ngành ──
        if dt_theo_nganh:
            chi_tiet, tong_dt, tong_gtgt, tong_tncn = _tinh_thue_chi_tiet_theo_nganh(dt_theo_nganh)

            st.markdown("#### 📊 Doanh thu & Thuế theo ngành nghề")

            # Bảng chi tiết
            header = "| Ngành nghề | Doanh thu | GTGT (%) | Thuế GTGT | TNCN (%) | Thuế TNCN |"
            sep = "|---|---:|:---:|---:|:---:|---:|"
            rows = []
            for nganh, info in chi_tiet.items():
                rows.append(
                    f"| {info['ten']} | {_fmt_tien(info['doanh_thu'])}đ | "
                    f"{info['ty_le_gtgt']}% | {_fmt_tien(info['thue_gtgt'])}đ | "
                    f"{info['ty_le_tncn']}% | {_fmt_tien(info['thue_tncn'])}đ |"
                )
            rows.append(f"| **TỔNG** | **{_fmt_tien(tong_dt)}đ** | | **{_fmt_tien(tong_gtgt)}đ** | | **{_fmt_tien(tong_tncn)}đ** |")
            st.markdown("\n".join([header, sep] + rows))

            # ── Xử lý phương pháp thuế TNCN theo lợi nhuận ──
            phuong_phap = cfg.get("phuong_phap_tinh_thue", "TY_LE_DOANH_THU")
            thue_tncn_final = tong_tncn

            if phuong_phap == "LOI_NHUAN":
                st.markdown("#### 💼 Thuế TNCN theo lợi nhuận")
                st.markdown(f"""
| Chỉ tiêu | Giá trị |
|---|---:|
| Doanh thu kỳ | {_fmt_tien(tong_dt)}đ |
| Chi phí được trừ (từ module Chi phí) | {_fmt_tien(chi_phi_duoc_tru)}đ |
| **Lợi nhuận tính thuế** | **{_fmt_tien(max(0, tong_dt - chi_phi_duoc_tru))}đ** |
                """)
                # Tính lại TNCN theo lợi nhuận
                dt_luy_ke_nam = sum(d.get("doanh_thu_ky", 0) or 0 for d in ds_thue) + tong_dt
                nhom_nam = _phan_nhom_doanh_thu(dt_luy_ke_nam)
                thue_tncn_final = _tinh_thue_loi_nhuan(tong_dt, chi_phi_duoc_tru, nhom_nam)
                st.info(f"Thuế TNCN theo lợi nhuận: **{_fmt_tien(thue_tncn_final)}đ** "
                        f"(thay vì {_fmt_tien(tong_tncn)}đ theo tỷ lệ)")

            # ── DT lũy kế + Nhóm ──
            dt_luy_ke = sum(d.get("doanh_thu_ky", 0) or 0 for d in ds_thue) + tong_dt
            nhom = _phan_nhom_doanh_thu(dt_luy_ke)
            han = _han_ke_khai(ky_thue)

            st.divider()
            st.markdown(f"""
#### 📋 Tổng hợp kỳ {ky_thue}
| Chỉ tiêu | Giá trị |
|---|---:|
| Doanh thu kỳ (tất cả ngành) | **{_fmt_tien(tong_dt)}đ** |
| DT lũy kế năm {nam} | {_fmt_tien(dt_luy_ke)}đ |
| Nhóm doanh thu | {_ten_nhom_doanh_thu(nhom)} |
| Thuế GTGT | **{_fmt_tien(tong_gtgt)}đ** |
| Thuế TNCN | **{_fmt_tien(thue_tncn_final)}đ** |
| **Tổng thuế phải nộp** | **{_fmt_tien(tong_gtgt + thue_tncn_final)}đ** |
| Hạn kê khai | {han.strftime('%d/%m/%Y')} |
            """)

            # Cảnh báo vượt ngưỡng
            if nhom == "MIEN_THUE":
                st.success("✅ DT lũy kế ≤ 500 triệu — miễn thuế GTGT & TNCN.")
            if nhom != "MIEN_THUE" and dt_luy_ke > 1_000_000_000:
                st.warning("⚠️ DT > 1 tỷ — bắt buộc sử dụng hóa đơn điện tử (NĐ 68/2026)")

            # ── Nút Lưu ──
            if st.button("💾 Lưu kỳ thuế", disabled=not _can_edit(), type="primary", key="hkd_luu_ky"):
                data = {
                    "ky_thue": ky_thue,
                    "doanh_thu_ky": tong_dt,
                    "doanh_thu_luy_ke_nam": dt_luy_ke,
                    "chi_phi_ky": chi_phi_duoc_tru,
                    "nhom_doanh_thu": nhom,
                    "thue_gtgt_phai_nop": tong_gtgt,
                    "thue_tncn_phai_nop": thue_tncn_final,
                    "trang_thai": "CHUA_KE_KHAI",
                    "han_ke_khai": han,
                    "ngay_ke_khai": None,
                    "ngay_nop": None,
                    "ghi_chu": "",
                }
                try:
                    conn = db_engine.get_connection()
                    _luu_ky_thue(conn, data)
                    conn.close()
                    st.success(f"✅ Đã lưu kỳ thuế {ky_thue}!")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Lỗi: {e}")

    # ══════════════════════════════════════════════
    # PHẦN 2: Danh sách kỳ đã kê khai (giữ nguyên logic cũ)
    # ══════════════════════════════════════════════
    if ds_thue:
        st.divider()
        st.subheader(f"📋 Các kỳ thuế năm {nam}")

        for d in ds_thue:
            trang_thai_icon = {"CHUA_KE_KHAI": "⏳", "DA_KE_KHAI": "📝", "DA_NOP": "✅"}.get(
                d.get("trang_thai", ""), "❓"
            )
            trang_thai_ten = {"CHUA_KE_KHAI": "Chưa kê khai", "DA_KE_KHAI": "Đã kê khai", "DA_NOP": "Đã nộp"}.get(
                d.get("trang_thai", ""), d.get("trang_thai", "")
            )

            with st.expander(
                f"{trang_thai_icon} **{d['ky_thue']}** — DT: {_fmt_tien(d.get('doanh_thu_ky', 0))}đ | "
                f"Thuế: {_fmt_tien((d.get('thue_gtgt_phai_nop', 0) or 0) + (d.get('thue_tncn_phai_nop', 0) or 0))}đ | "
                f"{trang_thai_ten}"
            ):
                st.markdown(f"""
- **Doanh thu kỳ:** {_fmt_tien(d.get('doanh_thu_ky', 0))}đ
- **DT lũy kế năm:** {_fmt_tien(d.get('doanh_thu_luy_ke_nam', 0))}đ
- **Nhóm:** {_ten_nhom_doanh_thu(d.get('nhom_doanh_thu', ''))}
- **Thuế GTGT:** {_fmt_tien(d.get('thue_gtgt_phai_nop', 0))}đ
- **Thuế TNCN:** {_fmt_tien(d.get('thue_tncn_phai_nop', 0))}đ
- **Hạn kê khai:** {d.get('han_ke_khai', '')}
                """)

                col_s1, col_s2, col_s3 = st.columns(3)
                with col_s1:
                    if d.get("trang_thai") == "CHUA_KE_KHAI":
                        if st.button("📝 Đánh dấu Đã kê khai", key=f"kk_{d['ky_thue']}",
                                     disabled=not _can_edit()):
                            conn = db_engine.get_connection()
                            d_copy = dict(d)
                            d_copy["trang_thai"] = "DA_KE_KHAI"
                            d_copy["ngay_ke_khai"] = datetime.date.today()
                            _luu_ky_thue(conn, d_copy)
                            conn.close()
                            st.rerun()
                with col_s2:
                    if d.get("trang_thai") in ("CHUA_KE_KHAI", "DA_KE_KHAI"):
                        if st.button("✅ Đánh dấu Đã nộp", key=f"nop_{d['ky_thue']}",
                                     disabled=not _can_edit()):
                            conn = db_engine.get_connection()
                            d_copy = dict(d)
                            d_copy["trang_thai"] = "DA_NOP"
                            d_copy["ngay_nop"] = datetime.date.today()
                            if not d_copy.get("ngay_ke_khai"):
                                d_copy["ngay_ke_khai"] = datetime.date.today()
                            _luu_ky_thue(conn, d_copy)
                            conn.close()
                            st.rerun()
                with col_s3:
                    if st.button("🗑️ Xóa", key=f"xoa_{d['ky_thue']}",
                                 disabled=not _can_edit()):
                        conn = db_engine.get_connection()
                        _xoa_ky_thue(conn, d["ky_thue"])
                        conn.close()
                        st.rerun()


# ═══════════════════════════════════════════════════════════════════
#  TAB 3: TỔNG HỢP & CẢNH BÁO
# ═══════════════════════════════════════════════════════════════════

def _render_tab_tong_hop(db_engine):
    """Tab báo cáo tổng hợp và cảnh báo hạn kê khai."""
    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if not cfg:
        st.warning("⚠️ Chưa có cấu hình HKD.")
        return

    nam = datetime.date.today().year
    conn = db_engine.get_connection()
    try:
        ds_thue = _get_ds_thue(conn, nam)
    finally:
        conn.close()

    hom_nay = datetime.date.today()

    # ── Tổng hợp năm ──
    st.subheader(f"📊 Tổng hợp năm {nam}")

    tong_dt = sum(d.get("doanh_thu_ky", 0) or 0 for d in ds_thue)
    tong_gtgt = sum(d.get("thue_gtgt_phai_nop", 0) or 0 for d in ds_thue)
    tong_tncn = sum(d.get("thue_tncn_phai_nop", 0) or 0 for d in ds_thue)
    tong_thue = tong_gtgt + tong_tncn

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Doanh thu lũy kế", _fmt_tien_trieu(tong_dt))
    col2.metric("Thuế GTGT", _fmt_tien(tong_gtgt) + "đ")
    col3.metric("Thuế TNCN", _fmt_tien(tong_tncn) + "đ")
    col4.metric("Tổng thuế", _fmt_tien(tong_thue) + "đ")

    nhom = _phan_nhom_doanh_thu(tong_dt)
    st.info(f"📌 Nhóm doanh thu hiện tại: **{_ten_nhom_doanh_thu(nhom)}** | "
            f"Phương pháp: **{cfg.get('phuong_phap_tinh_thue', '')}** | "
            f"Kỳ kê khai: **{cfg.get('ky_ke_khai', '')}**")

    # ── BHXH chủ hộ ──
    if cfg.get("chu_ho_dong_bhxh"):
        st.divider()
        st.subheader("🏥 BHXH Chủ hộ")
        muc_luong = int(cfg.get("muc_luong_dong_bhxh_chu_ho", MUC_THAM_CHIEU_BHXH))
        tien_thang = round(muc_luong * 0.295)
        pt = cfg.get("phuong_thuc_dong_bhxh", "HANG_THANG")
        pt_label = {"HANG_THANG": "hàng tháng", "3_THANG": "3 tháng/lần", "6_THANG": "6 tháng/lần"}.get(pt, pt)
        he_so = {"HANG_THANG": 1, "3_THANG": 3, "6_THANG": 6}.get(pt, 1)
        tien_moi_lan = tien_thang * he_so

        col_b1, col_b2, col_b3 = st.columns(3)
        col_b1.metric("Mức lương đóng", _fmt_tien(muc_luong) + "đ")
        col_b2.metric(f"Số tiền đóng ({pt_label})", _fmt_tien(tien_moi_lan) + "đ")
        col_b3.metric("Đóng cả năm", _fmt_tien(tien_thang * 12) + "đ")

    # ── Cảnh báo hạn kê khai ──
    st.divider()
    st.subheader("⏰ Cảnh báo hạn kê khai")

    canh_bao = []
    for d in ds_thue:
        if d.get("trang_thai") in ("CHUA_KE_KHAI", "DA_KE_KHAI"):
            han = d.get("han_ke_khai")
            if han:
                if isinstance(han, str):
                    try:
                        han = datetime.datetime.strptime(han, "%Y-%m-%d").date()
                    except Exception:
                        continue
                so_ngay_con = (han - hom_nay).days
                canh_bao.append({
                    "ky": d["ky_thue"],
                    "han": han,
                    "so_ngay": so_ngay_con,
                    "trang_thai": d["trang_thai"],
                })

    if not canh_bao:
        st.success("✅ Không có kỳ nào đang chờ kê khai/nộp.")
    else:
        for cb in sorted(canh_bao, key=lambda x: x["so_ngay"]):
            if cb["so_ngay"] < 0:
                st.error(f"🔴 **{cb['ky']}** — QUÁ HẠN {abs(cb['so_ngay'])} ngày! (hạn {cb['han'].strftime('%d/%m/%Y')})")
            elif cb["so_ngay"] <= 7:
                st.warning(f"🟡 **{cb['ky']}** — còn **{cb['so_ngay']} ngày** (hạn {cb['han'].strftime('%d/%m/%Y')})")
            elif cb["so_ngay"] <= 30:
                st.info(f"🔵 **{cb['ky']}** — còn {cb['so_ngay']} ngày (hạn {cb['han'].strftime('%d/%m/%Y')})")
            else:
                st.caption(f"⚪ **{cb['ky']}** — còn {cb['so_ngay']} ngày (hạn {cb['han'].strftime('%d/%m/%Y')})")

# ═══════════════════════════════════════════════════════════════════
#  AUTO MIGRATION — đảm bảo các bảng HKD tồn tại
# ═══════════════════════════════════════════════════════════════════

def ensure_hkd_tables(db_engine):
    """Tự động tạo các bảng HKD nếu chưa tồn tại. Gọi 1 lần khi vào module."""
    if st.session_state.get("_hkd_tables_ensured"):
        return
    try:
        conn = db_engine.get_connection()
        c = conn.cursor()

        # Bảng doanh thu chi tiết
        c.execute("""
            CREATE TABLE IF NOT EXISTS doanh_thu_hkd (
                id SERIAL PRIMARY KEY,
                thang TEXT NOT NULL,
                nhom_nganh TEXT NOT NULL DEFAULT 'THUONG_MAI',
                mo_ta TEXT,
                doanh_thu NUMERIC NOT NULL DEFAULT 0,
                so_hoa_don TEXT,
                ghi_chu TEXT,
                created_by TEXT,
                created_at TIMESTAMPTZ DEFAULT now(),
                updated_at TIMESTAMPTZ DEFAULT now()
            )
        """)

        # Bảng chi phí
        c.execute("""
            CREATE TABLE IF NOT EXISTS chi_phi_hkd (
                id SERIAL PRIMARY KEY,
                thang TEXT NOT NULL,
                loai_chi_phi TEXT NOT NULL DEFAULT 'KHAC',
                mo_ta TEXT NOT NULL,
                so_tien NUMERIC NOT NULL DEFAULT 0,
                chung_tu TEXT,
                duoc_tru_thue BOOLEAN DEFAULT true,
                ghi_chu TEXT,
                created_by TEXT,
                created_at TIMESTAMPTZ DEFAULT now(),
                updated_at TIMESTAMPTZ DEFAULT now()
            )
        """)

        # Bảng BHXH chủ hộ
        c.execute("""
            CREATE TABLE IF NOT EXISTS bhxh_chu_ho (
                id SERIAL PRIMARY KEY,
                thang TEXT NOT NULL UNIQUE,
                muc_luong_dong NUMERIC NOT NULL,
                ty_le_bhxh NUMERIC DEFAULT 25,
                ty_le_bhyt NUMERIC DEFAULT 4.5,
                so_tien_bhxh NUMERIC DEFAULT 0,
                so_tien_bhyt NUMERIC DEFAULT 0,
                tong_phai_dong NUMERIC DEFAULT 0,
                da_dong BOOLEAN DEFAULT false,
                ngay_dong DATE,
                bien_lai TEXT,
                ghi_chu TEXT,
                created_at TIMESTAMPTZ DEFAULT now(),
                updated_at TIMESTAMPTZ DEFAULT now()
            )
        """)

        # Bảng tỷ lệ thuế theo ngành
        c.execute("""
            CREATE TABLE IF NOT EXISTS ty_le_thue_nganh (
                id SERIAL PRIMARY KEY,
                nhom_nganh TEXT UNIQUE NOT NULL,
                ten_nhom TEXT NOT NULL,
                ty_le_gtgt NUMERIC NOT NULL,
                ty_le_tncn NUMERIC NOT NULL,
                ghi_chu TEXT,
                created_at TIMESTAMPTZ DEFAULT now()
            )
        """)
        c.execute("""
            INSERT INTO ty_le_thue_nganh (nhom_nganh, ten_nhom, ty_le_gtgt, ty_le_tncn) VALUES
            ('THUONG_MAI', 'Phân phối, cung cấp hàng hóa', 1, 0.5),
            ('DICH_VU', 'Dịch vụ, xây dựng không bao thầu NVL', 5, 2),
            ('SAN_XUAT', 'Sản xuất, vận tải, xây dựng có NVL', 3, 1.5),
            ('CHO_THUE', 'Cho thuê tài sản, BĐS, nội dung số', 5, 5),
            ('DAI_LY', 'Đại lý bảo hiểm, xổ số, bán hàng đa cấp', 0, 5),
            ('KHAC', 'Hoạt động kinh doanh khác', 2, 1)
            ON CONFLICT (nhom_nganh) DO NOTHING
        """)

        # Thêm cột mới vào bảng cau_hinh_thue_hkd nếu thiếu
        for col, typ in [
            ("ma_so_thue", "TEXT"),
            ("email", "TEXT"),
            ("dien_thoai", "TEXT"),
            ("giay_phep_kd", "TEXT"),
            ("nhom_doanh_thu", "TEXT DEFAULT 'NHOM_1'"),
            ("su_dung_hd_dien_tu", "BOOLEAN DEFAULT false"),
            ("so_tai_khoan_ngan_hang", "TEXT"),
            ("vi_dien_tu", "TEXT"),
            ("da_nganh", "BOOLEAN DEFAULT false"),
        ]:
            try:
                c.execute(f"ALTER TABLE cau_hinh_thue_hkd ADD COLUMN IF NOT EXISTS {col} {typ}")
            except Exception:
                pass

        conn.commit()
        conn.close()
        st.session_state["_hkd_tables_ensured"] = True
    except Exception as e:
        st.warning(f"⚠️ Không thể kiểm tra bảng HKD: {e}")

# ═══════════════════════════════════════════════════════════════════
#  DASHBOARD HKD — Tổng quan
# ═══════════════════════════════════════════════════════════════════

def render_dashboard_hkd(db_engine):
    """Dashboard tổng hợp cho Hộ kinh doanh."""
    ensure_hkd_tables(db_engine)
    st.title("📊 Tổng quan Hộ kinh doanh")

    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if not cfg:
        st.warning("⚠️ Chưa cấu hình thông tin HKD. Vui lòng vào **⚙️ Cấu hình HKD** để thiết lập.")
        return

    hom_nay = datetime.date.today()
    thang_hien_tai = hom_nay.strftime("%m/%Y")
    nam = hom_nay.year

    # ── Lấy dữ liệu ──
    conn = db_engine.get_connection()
    try:
        c = conn.cursor()

        # Doanh thu tháng này
        c.execute("SELECT COALESCE(SUM(doanh_thu), 0) FROM doanh_thu_hkd WHERE thang = %s",
                  (thang_hien_tai,))
        dt_thang = c.fetchone()[0]

        # Doanh thu lũy kế năm
        c.execute("SELECT COALESCE(SUM(doanh_thu), 0) FROM doanh_thu_hkd WHERE thang LIKE %s",
                  (f"%/{nam}",))
        dt_luy_ke = c.fetchone()[0]

        # Chi phí tháng này
        c.execute("SELECT COALESCE(SUM(so_tien), 0) FROM chi_phi_hkd WHERE thang = %s",
                  (thang_hien_tai,))
        cp_thang = c.fetchone()[0]

        # Thuế đã kê khai trong năm
        ds_thue = _get_ds_thue(conn, nam)
        tong_thue = sum(
            (d.get("thue_gtgt_phai_nop", 0) or 0) + (d.get("thue_tncn_phai_nop", 0) or 0)
            for d in ds_thue
        )

        # BHXH chủ hộ
        c.execute("""SELECT COUNT(*) FILTER (WHERE da_dong = false),
                            COALESCE(SUM(tong_phai_dong) FILTER (WHERE da_dong = true), 0)
                     FROM bhxh_chu_ho WHERE thang LIKE %s""", (f"%/{nam}",))
        bhxh_row = c.fetchone()
        bhxh_chua_dong = bhxh_row[0] if bhxh_row else 0
        bhxh_da_dong = bhxh_row[1] if bhxh_row else 0

        # Số NLĐ (nếu có bảng nhan_vien)
        so_nld = 0
        try:
            c.execute("SELECT COUNT(*) FROM nhan_vien WHERE trang_thai IN ('DANG_LAM', 'THU_VIEC')")
            so_nld = c.fetchone()[0]
        except Exception:
            pass

    finally:
        conn.close()

    # ── Hiển thị KPI ──
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 Doanh thu tháng", _fmt_tien_trieu(dt_thang))
    col2.metric("📉 Chi phí tháng", _fmt_tien_trieu(cp_thang))
    col3.metric("🧾 Thuế năm nay", _fmt_tien_trieu(tong_thue))

    if cfg.get("chu_ho_dong_bhxh"):
        muc_luong = int(cfg.get("muc_luong_dong_bhxh_chu_ho", MUC_THAM_CHIEU_BHXH))
        bhxh_thang = round(muc_luong * 0.295)
        col4.metric("📋 BHXH/tháng", _fmt_tien(bhxh_thang) + "đ")
    else:
        col4.metric("👥 Nhân sự", str(so_nld) + " người")

    st.divider()

    # ── Nhóm doanh thu + Cảnh báo ──
    nhom = _phan_nhom_doanh_thu(dt_luy_ke)
    ten_nhom = _ten_nhom_doanh_thu(nhom)

    col_info, col_canh_bao = st.columns([1, 1])

    with col_info:
        st.subheader("📌 Thông tin thuế")
        st.markdown(f"""
- **Tên HKD:** {cfg.get('ten_hkd', '')}
- **Chủ hộ:** {cfg.get('chu_ho_ten', '')}
- **DT lũy kế {nam}:** {_fmt_tien_trieu(dt_luy_ke)}
- **Nhóm DT:** {ten_nhom}
- **Phương pháp thuế:** {cfg.get('phuong_phap_tinh_thue', '')}
- **Kỳ kê khai:** {cfg.get('ky_ke_khai', '')}
        """)

    with col_canh_bao:
        st.subheader("⚠️ Cảnh báo")

        canh_bao_list = []

        # Cảnh báo vượt ngưỡng nhóm
        if nhom == "MIEN_THUE" and dt_luy_ke > 400_000_000:
            canh_bao_list.append(("🟡", f"DT lũy kế {_fmt_tien_trieu(dt_luy_ke)} — sắp vượt ngưỡng 500 triệu (phải nộp thuế)"))
        if nhom == "500TR_3TY" and dt_luy_ke > 2_500_000_000:
            canh_bao_list.append(("🟡", f"DT lũy kế {_fmt_tien_trieu(dt_luy_ke)} — sắp vượt 3 tỷ (phải đổi phương pháp thuế TNCN)"))
        if dt_luy_ke > 1_000_000_000 and not cfg.get("su_dung_hd_dien_tu"):
            canh_bao_list.append(("🔴", "DT > 1 tỷ — BẮT BUỘC sử dụng hóa đơn điện tử (NĐ 68/2026)"))

        # Cảnh báo hạn kê khai
        for d in ds_thue:
            if d.get("trang_thai") in ("CHUA_KE_KHAI", "DA_KE_KHAI"):
                han = d.get("han_ke_khai")
                if han:
                    if isinstance(han, str):
                        try:
                            han = datetime.datetime.strptime(han, "%Y-%m-%d").date()
                        except Exception:
                            continue
                    so_ngay = (han - hom_nay).days
                    if so_ngay < 0:
                        canh_bao_list.append(("🔴", f"**{d['ky_thue']}** — QUÁ HẠN {abs(so_ngay)} ngày!"))
                    elif so_ngay <= 7:
                        canh_bao_list.append(("🟡", f"**{d['ky_thue']}** — còn {so_ngay} ngày"))

        # Cảnh báo BHXH
        if bhxh_chua_dong > 0:
            canh_bao_list.append(("🟡", f"BHXH chủ hộ: {bhxh_chua_dong} tháng chưa đóng"))

        if not canh_bao_list:
            st.success("✅ Không có cảnh báo nào.")
        else:
            for icon, msg in canh_bao_list:
                if icon == "🔴":
                    st.error(f"{icon} {msg}")
                else:
                    st.warning(f"{icon} {msg}")

    # ── Biểu đồ DT theo tháng ──
    st.divider()
    st.subheader(f"📈 Doanh thu theo tháng — Năm {nam}")

    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT thang, SUM(doanh_thu) as tong_dt
            FROM doanh_thu_hkd
            WHERE thang LIKE %s
            GROUP BY thang
            ORDER BY thang
        """, (f"%/{nam}",))
        dt_theo_thang = c.fetchall()
    finally:
        conn.close()

    if dt_theo_thang:
        import pandas as pd
        df = pd.DataFrame(dt_theo_thang, columns=["Tháng", "Doanh thu"])
        st.bar_chart(df.set_index("Tháng"))
    else:
        st.info("Chưa có dữ liệu doanh thu. Vào **💰 Doanh thu & Chi phí** để nhập.")
        
# ═══════════════════════════════════════════════════════════════════
#  DOANH THU & CHI PHÍ — Nhập chi tiết theo tháng, theo ngành
# ═══════════════════════════════════════════════════════════════════

LOAI_CHI_PHI = {
    "NVL": "Nguyên vật liệu",
    "NHAN_CONG": "Nhân công / Lương NLĐ",
    "THUE_PHI": "Thuế, phí, lệ phí",
    "DIEN_NUOC": "Điện, nước, viễn thông",
    "KHAU_HAO": "Khấu hao tài sản",
    "VAN_PHONG": "Văn phòng phẩm, thuê mặt bằng",
    "KHAC": "Chi phí khác",
}


def render_doanh_thu_chi_phi(db_engine):
    """Module nhập Doanh thu & Chi phí chi tiết."""
    ensure_hkd_tables(db_engine)
    st.title("💰 Doanh thu & Chi phí")

    tab_dt, tab_cp = st.tabs(["📈 Doanh thu", "📉 Chi phí"])

    with tab_dt:
        _render_doanh_thu(db_engine)
    with tab_cp:
        _render_chi_phi(db_engine)


def _render_doanh_thu(db_engine):
    """Tab nhập doanh thu theo tháng & ngành nghề."""
    hom_nay = datetime.date.today()
    thang_mac_dinh = hom_nay.strftime("%m/%Y")

    st.subheader("➕ Thêm doanh thu mới")

    col1, col2, col3 = st.columns(3)
    with col1:
        thang_nhap = st.text_input("Tháng (mm/yyyy)", value=thang_mac_dinh, key="dt_thang")
    with col2:
        nhom_nganh = st.selectbox(
            "Nhóm ngành",
            list(TY_LE_THUE_THEO_NGANH.keys()),
            format_func=lambda k: TY_LE_THUE_THEO_NGANH[k]["ten"],
            key="dt_nganh"
        )
    with col3:
        _dt_text = st.text_input("Doanh thu (VNĐ)", value="0", key="dt_sotien")
        try:
            doanh_thu_nhap = int(_dt_text.replace(".", "").replace(",", "").strip())
            if doanh_thu_nhap > 0:
                st.caption(f"= {_fmt_tien(doanh_thu_nhap)}đ")
        except ValueError:
            doanh_thu_nhap = 0

    mo_ta = st.text_input("Mô tả (tùy chọn)", key="dt_mota")
    so_hd = st.text_input("Số hóa đơn (tùy chọn)", key="dt_sohd")

    if st.button("💾 Lưu doanh thu", disabled=not _can_edit(), type="primary", key="dt_luu"):
        if doanh_thu_nhap <= 0:
            st.error("❌ Doanh thu phải > 0")
        else:
            try:
                conn = db_engine.get_connection()
                c = conn.cursor()
                c.execute("""
                    INSERT INTO doanh_thu_hkd (thang, nhom_nganh, mo_ta, doanh_thu, so_hoa_don, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (thang_nhap, nhom_nganh, mo_ta, doanh_thu_nhap, so_hd,
                      st.session_state.get("user_display_name", "")))
                conn.commit()
                conn.close()
                st.success(f"✅ Đã lưu doanh thu {_fmt_tien(doanh_thu_nhap)}đ — {thang_nhap}")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Lỗi: {e}")

    # ── Danh sách doanh thu đã nhập ──
    st.divider()
    nam = hom_nay.year
    st.subheader(f"📋 Doanh thu đã nhập — Năm {nam}")

    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT id, thang, nhom_nganh, mo_ta, doanh_thu, so_hoa_don, created_at
            FROM doanh_thu_hkd
            WHERE thang LIKE %s
            ORDER BY thang DESC, created_at DESC
        """, (f"%/{nam}",))
        cols_dt = [desc[0] for desc in c.description]
        rows_dt = [dict(zip(cols_dt, r)) for r in c.fetchall()]
    finally:
        conn.close()

    if not rows_dt:
        st.info("Chưa có dữ liệu doanh thu năm nay.")
    else:
        tong = sum(r["doanh_thu"] for r in rows_dt)
        st.metric("Tổng DT năm", _fmt_tien(tong) + "đ")

        for r in rows_dt:
            ten_nganh = TY_LE_THUE_THEO_NGANH.get(r["nhom_nganh"], {}).get("ten", r["nhom_nganh"])
            with st.expander(f"📌 {r['thang']} — {ten_nganh} — {_fmt_tien(r['doanh_thu'])}đ"):
                st.markdown(f"- Mô tả: {r.get('mo_ta', '')}")
                st.markdown(f"- Hóa đơn: {r.get('so_hoa_don', '')}")
                if st.button("🗑️ Xóa", key=f"xoa_dt_{r['id']}", disabled=not _can_edit()):
                    conn = db_engine.get_connection()
                    c = conn.cursor()
                    c.execute("DELETE FROM doanh_thu_hkd WHERE id = %s", (r["id"],))
                    conn.commit()
                    conn.close()
                    st.rerun()


def _render_chi_phi(db_engine):
    """Tab nhập chi phí theo tháng."""
    hom_nay = datetime.date.today()
    thang_mac_dinh = hom_nay.strftime("%m/%Y")

    st.subheader("➕ Thêm chi phí mới")

    col1, col2, col3 = st.columns(3)
    with col1:
        thang_cp = st.text_input("Tháng (mm/yyyy)", value=thang_mac_dinh, key="cp_thang")
    with col2:
        loai_cp = st.selectbox(
            "Loại chi phí",
            list(LOAI_CHI_PHI.keys()),
            format_func=lambda k: LOAI_CHI_PHI[k],
            key="cp_loai"
        )
    with col3:
        _cp_text = st.text_input("Số tiền (VNĐ)", value="0", key="cp_sotien")
        try:
            so_tien_cp = int(_cp_text.replace(".", "").replace(",", "").strip())
            if so_tien_cp > 0:
                st.caption(f"= {_fmt_tien(so_tien_cp)}đ")
        except ValueError:
            so_tien_cp = 0

    mo_ta_cp = st.text_input("Mô tả chi phí", key="cp_mota")
    chung_tu = st.text_input("Số chứng từ/hóa đơn", key="cp_chungtu")
    duoc_tru = st.checkbox("Được trừ khi tính thuế TNCN", value=True, key="cp_duoctru")

    if st.button("💾 Lưu chi phí", disabled=not _can_edit(), type="primary", key="cp_luu"):
        if so_tien_cp <= 0 or not mo_ta_cp:
            st.error("❌ Cần nhập mô tả và số tiền > 0")
        else:
            try:
                conn = db_engine.get_connection()
                c = conn.cursor()
                c.execute("""
                    INSERT INTO chi_phi_hkd (thang, loai_chi_phi, mo_ta, so_tien, chung_tu,
                                             duoc_tru_thue, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (thang_cp, loai_cp, mo_ta_cp, so_tien_cp, chung_tu, duoc_tru,
                      st.session_state.get("user_display_name", "")))
                conn.commit()
                conn.close()
                st.success(f"✅ Đã lưu chi phí {_fmt_tien(so_tien_cp)}đ")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Lỗi: {e}")

    # ── Danh sách chi phí ──
    st.divider()
    nam = hom_nay.year
    st.subheader(f"📋 Chi phí đã nhập — Năm {nam}")

    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT id, thang, loai_chi_phi, mo_ta, so_tien, chung_tu, duoc_tru_thue
            FROM chi_phi_hkd WHERE thang LIKE %s
            ORDER BY thang DESC, created_at DESC
        """, (f"%/{nam}",))
        cols_cp = [desc[0] for desc in c.description]
        rows_cp = [dict(zip(cols_cp, r)) for r in c.fetchall()]
    finally:
        conn.close()

    if not rows_cp:
        st.info("Chưa có dữ liệu chi phí năm nay.")
    else:
        tong_cp = sum(r["so_tien"] for r in rows_cp)
        tong_duoc_tru = sum(r["so_tien"] for r in rows_cp if r.get("duoc_tru_thue"))
        c1, c2 = st.columns(2)
        c1.metric("Tổng chi phí", _fmt_tien(tong_cp) + "đ")
        c2.metric("Chi phí được trừ thuế", _fmt_tien(tong_duoc_tru) + "đ")

        for r in rows_cp:
            ten_loai = LOAI_CHI_PHI.get(r["loai_chi_phi"], r["loai_chi_phi"])
            tru_icon = "✅" if r.get("duoc_tru_thue") else "❌"
            with st.expander(f"{tru_icon} {r['thang']} — {ten_loai} — {_fmt_tien(r['so_tien'])}đ"):
                st.markdown(f"- Mô tả: {r['mo_ta']}")
                st.markdown(f"- Chứng từ: {r.get('chung_tu', '')}")
                if st.button("🗑️ Xóa", key=f"xoa_cp_{r['id']}", disabled=not _can_edit()):
                    conn = db_engine.get_connection()
                    c = conn.cursor()
                    c.execute("DELETE FROM chi_phi_hkd WHERE id = %s", (r["id"],))
                    conn.commit()
                    conn.close()
                    st.rerun()
                    
# ═══════════════════════════════════════════════════════════════════
#  WRAPPER FUNCTIONS — gọi từ app.py menu HKD
# ═══════════════════════════════════════════════════════════════════

def render_cau_hinh_hkd(db_engine):
    """Entry point cho menu ⚙️ Cấu hình HKD — hiển thị tab Cấu hình riêng."""
    ensure_hkd_tables(db_engine)
    st.title("⚙️ Cấu hình Hộ kinh doanh")
    _render_tab_cau_hinh(db_engine)


def render_chu_ho_nhan_su(db_engine):
    """Entry point cho menu 👤 Chủ hộ & Nhân sự."""
    ensure_hkd_tables(db_engine)
    st.title("👤 Chủ hộ & Nhân sự")

    tab_chu_ho, tab_nhan_su = st.tabs(["🏪 Thông tin Chủ hộ", "👥 Người lao động"])

    with tab_chu_ho:
        conn = db_engine.get_connection()
        try:
            cfg = _get_cau_hinh(conn)
        finally:
            conn.close()

        if not cfg:
            st.warning("⚠️ Chưa cấu hình HKD. Vào **⚙️ Cấu hình HKD** trước.")
            return

        st.markdown(f"""
| Thông tin | Giá trị |
|---|---|
| Tên HKD | **{cfg.get('ten_hkd', '')}** |
| Chủ hộ | **{cfg.get('chu_ho_ten', '')}** |
| CCCD | {cfg.get('chu_ho_cccd', '')} |
| Địa chỉ | {cfg.get('dia_chi', '')} |
| Số ĐKKD | {cfg.get('so_dkkd', '')} |
| Ngành nghề | {TY_LE_THUE_THEO_NGANH.get(cfg.get('nganh_nghe', ''), {}).get('ten', cfg.get('nganh_nghe', ''))} |
| BHXH | {'Có tham gia' if cfg.get('chu_ho_dong_bhxh') else 'Chưa tham gia'} |
        """)

        if cfg.get("chu_ho_dong_bhxh"):
            muc_luong = int(cfg.get("muc_luong_dong_bhxh_chu_ho", MUC_THAM_CHIEU_BHXH))
            tien_thang = round(muc_luong * 0.295)
            st.info(f"💰 BHXH: đóng {_fmt_tien(tien_thang)}đ/tháng (mức lương {_fmt_tien(muc_luong)}đ × 29,5%)")

    with tab_nhan_su:
        st.info("👥 Danh sách người lao động — dùng chung module **✅ Nhân viên** của HRM Master.")
        st.caption("Nếu HKD có NLĐ, thêm/quản lý nhân viên tại menu tương ứng trong app.")
        # TODO Phase 3: tích hợp trực tiếp danh sách NV vào đây  

def render_bhxh_chu_ho(db_engine):
    """Module BHXH cho chủ hộ kinh doanh — theo dõi đóng BHXH hàng tháng."""
    ensure_hkd_tables(db_engine)
    st.title("📋 BHXH Hộ kinh doanh")

    tab_chu_ho, tab_nld = st.tabs(["🏪 BHXH Chủ hộ", "👥 BHXH Người lao động"])

    with tab_chu_ho:
        _render_bhxh_chu_ho_tab(db_engine)

    with tab_nld:
        st.info("👥 BHXH cho người lao động của HKD — dùng chung module BHXH của HRM Master.")
        st.caption("Nếu HKD có NLĐ, quản lý BHXH tại module BHXH chính của app.")


def _render_bhxh_chu_ho_tab(db_engine):
    """Tab BHXH chủ hộ: tạo bản ghi hàng tháng, theo dõi đóng/chưa đóng."""

    # Đọc cấu hình
    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if not cfg:
        st.warning("⚠️ Chưa cấu hình HKD. Vào **⚙️ Cấu hình HKD** trước.")
        return

    if not cfg.get("chu_ho_dong_bhxh"):
        st.info("ℹ️ Chủ hộ chưa đăng ký tham gia BHXH bắt buộc. "
                "Vào **⚙️ Cấu hình HKD** để bật.")
        return

    muc_luong = int(cfg.get("muc_luong_dong_bhxh_chu_ho", MUC_THAM_CHIEU_BHXH))
    phuong_thuc = cfg.get("phuong_thuc_dong_bhxh", "HANG_THANG")
    pt_label = {"HANG_THANG": "hàng tháng", "3_THANG": "3 tháng/lần", "6_THANG": "6 tháng/lần"}.get(phuong_thuc, phuong_thuc)

    # Thông tin tổng quan
    tien_bhxh = round(muc_luong * 0.25)
    tien_bhyt = round(muc_luong * 0.045)
    tong_thang = tien_bhxh + tien_bhyt

    col_i1, col_i2, col_i3, col_i4 = st.columns(4)
    col_i1.metric("Mức lương đóng", _fmt_tien(muc_luong) + "đ")
    col_i2.metric("BHXH (25%)", _fmt_tien(tien_bhxh) + "đ/tháng")
    col_i3.metric("BHYT (4,5%)", _fmt_tien(tien_bhyt) + "đ/tháng")
    col_i4.metric("Tổng đóng", _fmt_tien(tong_thang) + "đ/tháng")
    st.caption(f"Phương thức đóng: **{pt_label}** | Đóng cả năm: **{_fmt_tien(tong_thang * 12)}đ**")

    st.divider()

    # ── Tạo bản ghi BHXH cho các tháng trong năm ──
    nam = datetime.date.today().year
    st.subheader(f"📅 Theo dõi đóng BHXH năm {nam}")

    # Đọc dữ liệu đã có
    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""SELECT * FROM bhxh_chu_ho WHERE thang LIKE %s ORDER BY thang""",
                  (f"%/{nam}",))
        cols = [desc[0] for desc in c.description]
        ds_bhxh = [dict(zip(cols, r)) for r in c.fetchall()]
    finally:
        conn.close()

    thang_da_co = {d["thang"] for d in ds_bhxh}

    # Nút tạo bản ghi cho cả năm (12 tháng)
    if len(ds_bhxh) < 12:
        if st.button(f"📋 Tạo bản ghi BHXH cho {12 - len(ds_bhxh)} tháng còn thiếu trong năm {nam}",
                     disabled=not _can_edit(), key="bhxh_tao_nam"):
            try:
                conn = db_engine.get_connection()
                c = conn.cursor()
                for m in range(1, 13):
                    thang_str = f"{m:02d}/{nam}"
                    if thang_str not in thang_da_co:
                        so_bhxh = round(muc_luong * 0.25)
                        so_bhyt = round(muc_luong * 0.045)
                        c.execute("""
                            INSERT INTO bhxh_chu_ho (thang, muc_luong_dong, ty_le_bhxh, ty_le_bhyt,
                                                     so_tien_bhxh, so_tien_bhyt, tong_phai_dong)
                            VALUES (%s, %s, 25, 4.5, %s, %s, %s)
                            ON CONFLICT (thang) DO NOTHING
                        """, (thang_str, muc_luong, so_bhxh, so_bhyt, so_bhxh + so_bhyt))
                conn.commit()
                conn.close()
                st.success(f"✅ Đã tạo bản ghi BHXH cho năm {nam}")
                st.rerun()
            except Exception as e:
                st.error(f"❌ Lỗi: {e}")

    # ── Hiển thị bảng BHXH ──
    if ds_bhxh:
        # Tổng hợp
        da_dong_count = sum(1 for d in ds_bhxh if d.get("da_dong"))
        chua_dong_count = sum(1 for d in ds_bhxh if not d.get("da_dong"))
        tong_da_dong = sum(d.get("tong_phai_dong", 0) or 0 for d in ds_bhxh if d.get("da_dong"))
        tong_chua_dong = sum(d.get("tong_phai_dong", 0) or 0 for d in ds_bhxh if not d.get("da_dong"))

        col_t1, col_t2, col_t3, col_t4 = st.columns(4)
        col_t1.metric("✅ Đã đóng", f"{da_dong_count} tháng")
        col_t2.metric("⏳ Chưa đóng", f"{chua_dong_count} tháng")
        col_t3.metric("Tổng đã đóng", _fmt_tien(tong_da_dong) + "đ")
        col_t4.metric("Còn phải đóng", _fmt_tien(tong_chua_dong) + "đ")

        st.divider()

        for d in ds_bhxh:
            thang_label = d["thang"]
            da_dong = d.get("da_dong", False)
            icon = "✅" if da_dong else "⏳"
            tong = d.get("tong_phai_dong", 0) or 0

            col_a, col_b, col_c = st.columns([2, 1, 1])
            with col_a:
                st.markdown(f"{icon} **Tháng {thang_label}** — {_fmt_tien(tong)}đ"
                            + (f" — đóng ngày {d['ngay_dong']}" if da_dong and d.get('ngay_dong') else ""))
            with col_b:
                if not da_dong:
                    if st.button("✅ Đã đóng", key=f"bhxh_dong_{thang_label}",
                                 disabled=not _can_edit()):
                        conn = db_engine.get_connection()
                        c = conn.cursor()
                        c.execute("""UPDATE bhxh_chu_ho SET da_dong = true, ngay_dong = %s,
                                     updated_at = now() WHERE thang = %s""",
                                  (datetime.date.today(), thang_label))
                        conn.commit()
                        conn.close()
                        st.rerun()
            with col_c:
                if da_dong:
                    if st.button("↩️ Hủy đóng", key=f"bhxh_huy_{thang_label}",
                                 disabled=not _can_edit()):
                        conn = db_engine.get_connection()
                        c = conn.cursor()
                        c.execute("""UPDATE bhxh_chu_ho SET da_dong = false, ngay_dong = NULL,
                                     updated_at = now() WHERE thang = %s""",
                                  (thang_label,))
                        conn.commit()
                        conn.close()
                        st.rerun()
    else:
        st.info("Chưa có bản ghi BHXH. Nhấn nút phía trên để tạo.")

def _xuat_to_khai_01cnkd(cfg, ky_thue, dt_theo_nganh, chi_tiet_thue, tong_dt, tong_gtgt, tong_tncn, chi_phi_duoc_tru=0):
    """Tạo file Excel tờ khai thuế mẫu 01/CNKD (TT 18/2026).
    Trả về bytes của file Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "01-CNKD"

    # ── Styles ──
    font_title = Font(name="Times New Roman", size=13, bold=True)
    font_header = Font(name="Times New Roman", size=11, bold=True)
    font_normal = Font(name="Times New Roman", size=11)
    font_small = Font(name="Times New Roman", size=9, italic=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # ── Cột width ──
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # ── Header ──
    row = 1
    ws.merge_cells("A1:E1")
    ws["A1"] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws["A1"].font = font_header
    ws["A1"].alignment = align_center

    row = 2
    ws.merge_cells("A2:E2")
    ws["A2"] = "Độc lập - Tự do - Hạnh phúc"
    ws["A2"].font = font_normal
    ws["A2"].alignment = align_center

    row = 4
    ws.merge_cells("A4:E4")
    ws["A4"] = "TỜ KHAI THUẾ"
    ws["A4"].font = Font(name="Times New Roman", size=14, bold=True)
    ws["A4"].alignment = align_center

    row = 5
    ws.merge_cells("A5:E5")
    ws["A5"] = "Đối với hộ kinh doanh, cá nhân kinh doanh"
    ws["A5"].font = font_normal
    ws["A5"].alignment = align_center

    row = 6
    ws.merge_cells("A6:E6")
    ws["A6"] = f"(Mẫu số 01/CNKD — TT 18/2026/TT-BTC)"
    ws["A6"].font = font_small
    ws["A6"].alignment = align_center

    # ── Thông tin HKD ──
    row = 8
    ws["A8"] = "[01]"
    ws["A8"].font = font_normal
    ws["B8"] = f"Kỳ tính thuế: {ky_thue}"
    ws["B8"].font = font_header

    row = 9
    ws["A9"] = "[02]"
    ws["A9"].font = font_normal
    ws["B9"] = f"Tên HKD: {cfg.get('ten_hkd', '')}"
    ws["B9"].font = font_normal

    row = 10
    ws["A10"] = "[03]"
    ws["B10"] = f"Mã số thuế: {cfg.get('ma_so_thue', '')}"
    ws["B10"].font = font_normal

    row = 11
    ws["A11"] = "[04]"
    ws["B11"] = f"Địa chỉ kinh doanh: {cfg.get('dia_chi', '')}"
    ws["B11"].font = font_normal

    row = 12
    ws["A12"] = "[05]"
    ws["B12"] = f"Chủ hộ: {cfg.get('chu_ho_ten', '')} — CCCD: {cfg.get('chu_ho_cccd', '')}"
    ws["B12"].font = font_normal

    phuong_phap = cfg.get("phuong_phap_tinh_thue", "TY_LE_DOANH_THU")
    pp_label = "Tỷ lệ % trên doanh thu" if phuong_phap == "TY_LE_DOANH_THU" else "Thu nhập tính thuế (DT - CP)"
    row = 13
    ws["A13"] = "[06]"
    ws["B13"] = f"Phương pháp tính thuế TNCN: {pp_label}"
    ws["B13"].font = font_normal

    # ── Phần I: Hoạt động SXKD ──
    row = 15
    ws.merge_cells("A15:E15")
    ws["A15"] = "I. HOẠT ĐỘNG SẢN XUẤT, KINH DOANH HÀNG HÓA, CUNG CẤP DỊCH VỤ"
    ws["A15"].font = font_header

    # Header bảng
    row = 17
    headers = ["STT", "Ngành nghề kinh doanh", "Tỷ lệ\nGTGT (%)", "Doanh thu\n(VNĐ)", "Thuế GTGT\n(VNĐ)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = fill_header

    # Thêm cột TNCN
    headers2 = ["STT", "Ngành nghề kinh doanh", "Tỷ lệ\nTNCN (%)", "Doanh thu\n(VNĐ)", "Thuế TNCN\n(VNĐ)"]

    # Dữ liệu theo ngành
    row = 18
    stt = 1
    for nganh, info in chi_tiet_thue.items():
        ws.cell(row=row, column=1, value=stt).font = font_normal
        ws.cell(row=row, column=1).alignment = align_center
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=info["ten"]).font = font_normal
        ws.cell(row=row, column=2).alignment = align_left
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=f"{info['ty_le_gtgt']}%").font = font_normal
        ws.cell(row=row, column=3).alignment = align_center
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=info["doanh_thu"]).font = font_normal
        ws.cell(row=row, column=4).alignment = align_right
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0'

        ws.cell(row=row, column=5, value=info["thue_gtgt"]).font = font_normal
        ws.cell(row=row, column=5).alignment = align_right
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '#,##0'

        stt += 1
        row += 1

    # Dòng tổng GTGT
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value="TỔNG CỘNG THUẾ GTGT").font = font_header
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=4, value=tong_dt).font = font_header
    ws.cell(row=row, column=4).alignment = align_right
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).number_format = '#,##0'
    ws.cell(row=row, column=5, value=tong_gtgt).font = font_header
    ws.cell(row=row, column=5).alignment = align_right
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).number_format = '#,##0'
    ws.cell(row=row, column=5).fill = fill_yellow
    row += 2

    # ── Phần thuế TNCN ──
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "II. THUẾ THU NHẬP CÁ NHÂN"
    ws[f"A{row}"].font = font_header
    row += 1

    if phuong_phap == "TY_LE_DOANH_THU":
        # Tỷ lệ % trên DT — hiển thị theo ngành
        headers_tncn = ["STT", "Ngành nghề kinh doanh", "Tỷ lệ\nTNCN (%)", "Doanh thu\n(VNĐ)", "Thuế TNCN\n(VNĐ)"]
        row += 1
        for col_idx, h in enumerate(headers_tncn, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border
            cell.fill = fill_header
        row += 1

        stt = 1
        for nganh, info in chi_tiet_thue.items():
            ws.cell(row=row, column=1, value=stt).font = font_normal
            ws.cell(row=row, column=1).alignment = align_center
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=info["ten"]).font = font_normal
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value=f"{info['ty_le_tncn']}%").font = font_normal
            ws.cell(row=row, column=3).alignment = align_center
            ws.cell(row=row, column=3).border = thin_border
            ws.cell(row=row, column=4, value=info["doanh_thu"]).font = font_normal
            ws.cell(row=row, column=4).alignment = align_right
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5, value=info["thue_tncn"]).font = font_normal
            ws.cell(row=row, column=5).alignment = align_right
            ws.cell(row=row, column=5).border = thin_border
            ws.cell(row=row, column=5).number_format = '#,##0'
            stt += 1
            row += 1

        # Tổng TNCN
        ws.cell(row=row, column=2, value="TỔNG CỘNG THUẾ TNCN").font = font_header
        ws.cell(row=row, column=2).border = thin_border
        ws.cell(row=row, column=5, value=tong_tncn).font = font_header
        ws.cell(row=row, column=5).alignment = align_right
        ws.cell(row=row, column=5).border = thin_border
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).fill = fill_yellow
        for c in [1, 3, 4]:
            ws.cell(row=row, column=c).border = thin_border
    else:
        # Phương pháp lợi nhuận
        row += 1
        labels = [
            ("[07]", "Tổng doanh thu", tong_dt),
            ("[08]", "Tổng chi phí được trừ", chi_phi_duoc_tru),
            ("[09]", "Thu nhập tính thuế (= [07] - [08])", max(0, tong_dt - chi_phi_duoc_tru)),
            ("[10]", "Thuế TNCN phải nộp", tong_tncn),
        ]
        for code, label, val in labels:
            ws.cell(row=row, column=1, value=code).font = font_normal
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=label).font = font_normal
            ws.cell(row=row, column=2).border = thin_border
            ws.merge_cells(f"D{row}:E{row}")
            ws.cell(row=row, column=4, value=val).font = font_header
            ws.cell(row=row, column=4).alignment = align_right
            ws.cell(row=row, column=4).border = thin_border
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=3).border = thin_border
            if code == "[10]":
                ws.cell(row=row, column=4).fill = fill_yellow
            row += 1

    row += 2

    # ── Tổng hợp ──
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "III. TỔNG SỐ THUẾ PHẢI NỘP"
    ws[f"A{row}"].font = font_header
    row += 1

    summary = [
        ("Thuế GTGT", tong_gtgt),
        ("Thuế TNCN", tong_tncn),
        ("TỔNG CỘNG", tong_gtgt + tong_tncn),
    ]
    for label, val in summary:
        ws.cell(row=row, column=2, value=label).font = font_header if label == "TỔNG CỘNG" else font_normal
        ws.cell(row=row, column=2).border = thin_border
        ws.merge_cells(f"D{row}:E{row}")
        ws.cell(row=row, column=4, value=val).font = Font(name="Times New Roman", size=12, bold=True)
        ws.cell(row=row, column=4).alignment = align_right
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=3).border = thin_border
        if label == "TỔNG CỘNG":
            ws.cell(row=row, column=4).fill = fill_yellow
        row += 1

    # ── Ký tên ──
    row += 2
    ws.merge_cells(f"C{row}:E{row}")
    ws[f"C{row}"] = f"Ngày ..... tháng ..... năm {datetime.date.today().year}"
    ws[f"C{row}"].font = font_normal
    ws[f"C{row}"].alignment = align_center
    row += 1
    ws.merge_cells(f"C{row}:E{row}")
    ws[f"C{row}"] = "NGƯỜI NỘP THUẾ"
    ws[f"C{row}"].font = font_header
    ws[f"C{row}"].alignment = align_center
    row += 1
    ws.merge_cells(f"C{row}:E{row}")
    ws[f"C{row}"] = "(Ký, ghi rõ họ tên)"
    ws[f"C{row}"].font = font_small
    ws[f"C{row}"].alignment = align_center

    # ── Xuất bytes ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
    
def _xuat_bao_cao_tong_hop_nam(db_engine, cfg, nam):
    """Tạo file Excel báo cáo tổng hợp DT-CP-Thuế-BHXH theo năm.
    Trả về bytes của file Excel.
    """
    wb = openpyxl.Workbook()

    # ── Styles ──
    font_title = Font(name="Arial", size=13, bold=True)
    font_header = Font(name="Arial", size=10, bold=True)
    font_normal = Font(name="Arial", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_header_white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_total = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # ══════════════════════════════════════════
    # SHEET 1: DOANH THU THEO THÁNG & NGÀNH
    # ══════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Doanh thu"
    ws1.column_dimensions["A"].width = 12
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws1.column_dimensions[col_letter].width = 18

    ws1.merge_cells("A1:H1")
    ws1["A1"] = f"BÁO CÁO DOANH THU THEO THÁNG — NĂM {nam}"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_center

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"{cfg.get('ten_hkd', '')} — MST: {cfg.get('ma_so_thue', '')}"
    ws1["A2"].font = font_normal
    ws1["A2"].alignment = align_center

    # Header
    nganh_list = list(TY_LE_THUE_THEO_NGANH.keys())
    headers_dt = ["Tháng"] + [TY_LE_THUE_THEO_NGANH[n]["ten"][:20] for n in nganh_list] + ["TỔNG"]
    for col_idx, h in enumerate(headers_dt, 1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.font = font_header_white
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = fill_header

    # Lấy dữ liệu
    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT thang, nhom_nganh, SUM(doanh_thu) as tong
            FROM doanh_thu_hkd WHERE thang LIKE %s
            GROUP BY thang, nhom_nganh ORDER BY thang
        """, (f"%/{nam}",))
        dt_rows = c.fetchall()
    finally:
        conn.close()

    # Pivot: {thang: {nganh: dt}}
    dt_pivot = {}
    for thang, nganh, tong in dt_rows:
        if thang not in dt_pivot:
            dt_pivot[thang] = {}
        dt_pivot[thang][nganh] = float(tong)

    row = 5
    tong_col = {n: 0 for n in nganh_list}
    tong_tong = 0
    for m in range(1, 13):
        thang_str = f"{m:02d}/{nam}"
        ws1.cell(row=row, column=1, value=thang_str).font = font_normal
        ws1.cell(row=row, column=1).alignment = align_center
        ws1.cell(row=row, column=1).border = thin_border

        dt_thang = dt_pivot.get(thang_str, {})
        tong_thang_m = 0
        for col_idx, nganh in enumerate(nganh_list, 2):
            val = dt_thang.get(nganh, 0)
            ws1.cell(row=row, column=col_idx, value=val if val else "").font = font_normal
            ws1.cell(row=row, column=col_idx).alignment = align_right
            ws1.cell(row=row, column=col_idx).border = thin_border
            ws1.cell(row=row, column=col_idx).number_format = '#,##0'
            tong_col[nganh] += val
            tong_thang_m += val

        ws1.cell(row=row, column=len(nganh_list) + 2, value=tong_thang_m if tong_thang_m else "").font = font_normal
        ws1.cell(row=row, column=len(nganh_list) + 2).alignment = align_right
        ws1.cell(row=row, column=len(nganh_list) + 2).border = thin_border
        ws1.cell(row=row, column=len(nganh_list) + 2).number_format = '#,##0'
        tong_tong += tong_thang_m
        row += 1

    # Dòng tổng
    ws1.cell(row=row, column=1, value="TỔNG NĂM").font = font_header
    ws1.cell(row=row, column=1).border = thin_border
    ws1.cell(row=row, column=1).fill = fill_total
    for col_idx, nganh in enumerate(nganh_list, 2):
        ws1.cell(row=row, column=col_idx, value=tong_col[nganh]).font = font_header
        ws1.cell(row=row, column=col_idx).alignment = align_right
        ws1.cell(row=row, column=col_idx).border = thin_border
        ws1.cell(row=row, column=col_idx).number_format = '#,##0'
        ws1.cell(row=row, column=col_idx).fill = fill_total
    ws1.cell(row=row, column=len(nganh_list) + 2, value=tong_tong).font = font_header
    ws1.cell(row=row, column=len(nganh_list) + 2).alignment = align_right
    ws1.cell(row=row, column=len(nganh_list) + 2).border = thin_border
    ws1.cell(row=row, column=len(nganh_list) + 2).number_format = '#,##0'
    ws1.cell(row=row, column=len(nganh_list) + 2).fill = fill_total

    # ══════════════════════════════════════════
    # SHEET 2: CHI PHÍ THEO THÁNG
    # ══════════════════════════════════════════
    ws2 = wb.create_sheet("Chi phi")
    ws2.column_dimensions["A"].width = 12
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws2.column_dimensions[col_letter].width = 16

    ws2.merge_cells("A1:I1")
    ws2["A1"] = f"BÁO CÁO CHI PHÍ THEO THÁNG — NĂM {nam}"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = align_center

    loai_cp_list = ["NVL", "NHAN_CONG", "THUE_PHI", "DIEN_NUOC", "KHAU_HAO", "VAN_PHONG", "KHAC"]
    loai_cp_ten = {"NVL": "NVL", "NHAN_CONG": "Nhân công", "THUE_PHI": "Thuế/phí",
                   "DIEN_NUOC": "Điện nước", "KHAU_HAO": "Khấu hao",
                   "VAN_PHONG": "Văn phòng", "KHAC": "Khác"}

    headers_cp = ["Tháng"] + [loai_cp_ten.get(l, l) for l in loai_cp_list] + ["TỔNG"]
    for col_idx, h in enumerate(headers_cp, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_header_white
        cell.alignment = align_center
        cell.border = thin_border
        cell.fill = fill_header

    conn = db_engine.get_connection()
    try:
        c = conn.cursor()
        c.execute("""
            SELECT thang, loai_chi_phi, SUM(so_tien) as tong
            FROM chi_phi_hkd WHERE thang LIKE %s
            GROUP BY thang, loai_chi_phi ORDER BY thang
        """, (f"%/{nam}",))
        cp_rows = c.fetchall()
    finally:
        conn.close()

    cp_pivot = {}
    for thang, loai, tong in cp_rows:
        if thang not in cp_pivot:
            cp_pivot[thang] = {}
        cp_pivot[thang][loai] = float(tong)

    row2 = 4
    for m in range(1, 13):
        thang_str = f"{m:02d}/{nam}"
        ws2.cell(row=row2, column=1, value=thang_str).font = font_normal
        ws2.cell(row=row2, column=1).alignment = align_center
        ws2.cell(row=row2, column=1).border = thin_border
        cp_thang = cp_pivot.get(thang_str, {})
        tong_m = 0
        for col_idx, loai in enumerate(loai_cp_list, 2):
            val = cp_thang.get(loai, 0)
            ws2.cell(row=row2, column=col_idx, value=val if val else "").font = font_normal
            ws2.cell(row=row2, column=col_idx).alignment = align_right
            ws2.cell(row=row2, column=col_idx).border = thin_border
            ws2.cell(row=row2, column=col_idx).number_format = '#,##0'
            tong_m += val
        ws2.cell(row=row2, column=len(loai_cp_list) + 2, value=tong_m if tong_m else "").font = font_normal
        ws2.cell(row=row2, column=len(loai_cp_list) + 2).alignment = align_right
        ws2.cell(row=row2, column=len(loai_cp_list) + 2).border = thin_border
        ws2.cell(row=row2, column=len(loai_cp_list) + 2).number_format = '#,##0'
        row2 += 1

    # ══════════════════════════════════════════
    # SHEET 3: THUẾ & BHXH
    # ══════════════════════════════════════════
    ws3 = wb.create_sheet("Thue va BHXH")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 20

    ws3.merge_cells("A1:C1")
    ws3["A1"] = f"TỔNG HỢP THUẾ & BHXH — NĂM {nam}"
    ws3["A1"].font = font_title
    ws3["A1"].alignment = align_center

    # Lấy dữ liệu thuế
    conn = db_engine.get_connection()
    try:
        ds_thue = _get_ds_thue(conn, nam)
        c = conn.cursor()
        c.execute("""SELECT thang, tong_phai_dong, da_dong FROM bhxh_chu_ho
                     WHERE thang LIKE %s ORDER BY thang""", (f"%/{nam}",))
        bhxh_cols = [desc[0] for desc in c.description]
        ds_bhxh = [dict(zip(bhxh_cols, r)) for r in c.fetchall()]
    finally:
        conn.close()

    row3 = 3
    ws3.cell(row=row3, column=1, value="CHỈ TIÊU").font = font_header_white
    ws3.cell(row=row3, column=1).fill = fill_header
    ws3.cell(row=row3, column=1).border = thin_border
    ws3.cell(row=row3, column=2, value="SỐ TIỀN (VNĐ)").font = font_header_white
    ws3.cell(row=row3, column=2).fill = fill_header
    ws3.cell(row=row3, column=2).border = thin_border
    ws3.cell(row=row3, column=2).alignment = align_center
    ws3.cell(row=row3, column=3, value="GHI CHÚ").font = font_header_white
    ws3.cell(row=row3, column=3).fill = fill_header
    ws3.cell(row=row3, column=3).border = thin_border

    tong_thue_gtgt = sum(d.get("thue_gtgt_phai_nop", 0) or 0 for d in ds_thue)
    tong_thue_tncn = sum(d.get("thue_tncn_phai_nop", 0) or 0 for d in ds_thue)
    tong_dt_nam = sum(d.get("doanh_thu_ky", 0) or 0 for d in ds_thue)
    tong_cp_nam = sum(d.get("chi_phi_ky", 0) or 0 for d in ds_thue)
    bhxh_da_dong = sum(d.get("tong_phai_dong", 0) or 0 for d in ds_bhxh if d.get("da_dong"))
    bhxh_chua_dong = sum(d.get("tong_phai_dong", 0) or 0 for d in ds_bhxh if not d.get("da_dong"))

    data_rows = [
        ("A. DOANH THU & CHI PHÍ", "", ""),
        ("Tổng doanh thu năm", tong_dt_nam, ""),
        ("Tổng chi phí năm", tong_cp_nam, ""),
        ("Lợi nhuận (DT - CP)", max(0, tong_dt_nam - tong_cp_nam), ""),
        ("", "", ""),
        ("B. THUẾ", "", ""),
        ("Thuế GTGT đã kê khai", tong_thue_gtgt, f"{len(ds_thue)} kỳ"),
        ("Thuế TNCN đã kê khai", tong_thue_tncn, ""),
        ("Tổng thuế", tong_thue_gtgt + tong_thue_tncn, ""),
        ("", "", ""),
        ("C. BHXH CHỦ HỘ", "", ""),
        ("Đã đóng", bhxh_da_dong, f"{sum(1 for d in ds_bhxh if d.get('da_dong'))} tháng"),
        ("Chưa đóng", bhxh_chua_dong, f"{sum(1 for d in ds_bhxh if not d.get('da_dong'))} tháng"),
        ("Tổng BHXH năm", bhxh_da_dong + bhxh_chua_dong, ""),
        ("", "", ""),
        ("D. TỔNG CHI PHÍ PHÁP LÝ", "", ""),
        ("Thuế + BHXH", tong_thue_gtgt + tong_thue_tncn + bhxh_da_dong + bhxh_chua_dong, ""),
    ]

    row3 = 4
    for label, val, note in data_rows:
        is_header = label.startswith(("A.", "B.", "C.", "D."))
        cell_a = ws3.cell(row=row3, column=1, value=label)
        cell_a.font = font_header if is_header else font_normal
        cell_a.border = thin_border
        cell_b = ws3.cell(row=row3, column=2, value=val if val else "")
        cell_b.font = font_header if is_header or "Tổng" in label else font_normal
        cell_b.alignment = align_right
        cell_b.border = thin_border
        cell_b.number_format = '#,##0'
        if "Tổng" in label:
            cell_b.fill = fill_total
        cell_c = ws3.cell(row=row3, column=3, value=note)
        cell_c.font = font_normal
        cell_c.border = thin_border
        row3 += 1

    # ── Xuất bytes ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
    
def _render_tab_xuat_bao_cao(db_engine):
    """Tab xuất tờ khai thuế + báo cáo tổng hợp năm ra Excel."""
    conn = db_engine.get_connection()
    try:
        cfg = _get_cau_hinh(conn)
    finally:
        conn.close()

    if not cfg:
        st.warning("⚠️ Chưa cấu hình HKD.")
        return

    nam = datetime.date.today().year

    st.subheader("🧾 Xuất tờ khai thuế 01/CNKD")
    st.caption("Mẫu 01/CNKD theo TT 18/2026/TT-BTC — Tờ khai thuế quý/năm cho HKD có DT > 500 triệu")

    conn = db_engine.get_connection()
    try:
        ds_thue = _get_ds_thue(conn, nam)
    finally:
        conn.close()

    if not ds_thue:
        st.info("Chưa có kỳ thuế nào. Vào tab **📊 Kê khai theo kỳ** để tạo trước.")
    else:
        ky_list = [d["ky_thue"] for d in ds_thue]
        ky_chon = st.selectbox("Chọn kỳ thuế để xuất", ky_list, key="xuat_ky")

        if st.button("📥 Xuất tờ khai 01/CNKD", type="primary", key="btn_xuat_01cnkd"):
            # Lấy dữ liệu chi tiết theo ngành cho kỳ
            conn = db_engine.get_connection()
            try:
                dt_theo_nganh = _lay_doanh_thu_theo_nganh(conn, ky_chon)
                _, chi_phi_duoc_tru = _lay_chi_phi_ky(conn, ky_chon)
            finally:
                conn.close()

            if not dt_theo_nganh:
                # Fallback: dùng DT tổng từ bảng theo_doi_thue_hkd
                d_chon = next((d for d in ds_thue if d["ky_thue"] == ky_chon), None)
                if d_chon:
                    nganh = cfg.get("nganh_nghe", "THUONG_MAI")
                    dt_theo_nganh = {nganh: float(d_chon.get("doanh_thu_ky", 0) or 0)}
                    chi_phi_duoc_tru = float(d_chon.get("chi_phi_ky", 0) or 0)

            chi_tiet, tong_dt, tong_gtgt, tong_tncn = _tinh_thue_chi_tiet_theo_nganh(dt_theo_nganh)

            # Nếu phương pháp lợi nhuận, tính lại TNCN
            if cfg.get("phuong_phap_tinh_thue") == "LOI_NHUAN":
                nhom = _phan_nhom_doanh_thu(tong_dt)
                tong_tncn = _tinh_thue_loi_nhuan(tong_dt, chi_phi_duoc_tru, nhom)

            excel_bytes = _xuat_to_khai_01cnkd(
                cfg, ky_chon, dt_theo_nganh, chi_tiet,
                tong_dt, tong_gtgt, tong_tncn, chi_phi_duoc_tru
            )

            ten_file = f"01_CNKD_{ky_chon.replace('/', '_')}_{cfg.get('ma_so_thue', 'HKD')}.xlsx"
            st.download_button(
                label=f"⬇️ Tải {ten_file}",
                data=excel_bytes,
                file_name=ten_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.success(f"✅ Đã tạo tờ khai {ky_chon}!")

    st.divider()

    # ── Báo cáo tổng hợp năm ──
    st.subheader(f"📊 Báo cáo tổng hợp năm {nam}")
    st.caption("Gồm 3 sheet: Doanh thu theo tháng & ngành, Chi phí theo tháng, Tổng hợp Thuế & BHXH")

    nam_xuat = st.selectbox("Năm xuất báo cáo", [nam - 1, nam], index=1, key="xuat_nam")

    if st.button("📥 Xuất báo cáo tổng hợp", type="primary", key="btn_xuat_tonghop"):
        excel_bytes = _xuat_bao_cao_tong_hop_nam(db_engine, cfg, nam_xuat)
        ten_file = f"BaoCao_TongHop_{nam_xuat}_{cfg.get('ma_so_thue', 'HKD')}.xlsx"
        st.download_button(
            label=f"⬇️ Tải {ten_file}",
            data=excel_bytes,
            file_name=ten_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success(f"✅ Đã tạo báo cáo năm {nam_xuat}!")
        

# ═══════════════════════════════════════════════════════════════════
#  HÀM CHÍNH — gọi từ app.py
# ═══════════════════════════════════════════════════════════════════

def render_thue_hkd(db_engine):
    """Entry point — gọi từ app.py khi menu == '🧾 Kê khai Thuế'."""
    ensure_hkd_tables(db_engine)
    st.title("🧾 Kê khai Thuế")

    tab1, tab2, tab3 = st.tabs([
        "📊 Kê khai theo kỳ",
        "📋 Tổng hợp & Cảnh báo",
        "📥 Xuất báo cáo",
    ])

    with tab1:
        _render_tab_theo_doi(db_engine)
    with tab2:
        _render_tab_tong_hop(db_engine)
    with tab3:
        _render_tab_xuat_bao_cao(db_engine)
