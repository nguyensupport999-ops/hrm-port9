# -*- coding: utf-8 -*-
"""
tinh_thu_nhap.py
=================
Module "💰 Tính thu nhập" — chính sách lương 3P (Pay for Position - Person - Performance)
cho tenant CHL (và các tenant khác chọn phương pháp 3P).

CÁCH TÍCH HỢP VÀO app.py
------------------------
1) Thêm vào đầu app.py:
       import tinh_thu_nhap
2) Thay toàn bộ khối `elif menu == "💰 Tính thu nhập": ...` hiện có bằng:
       elif menu == "💰 Tính thu nhập":
           tinh_thu_nhap.show_tinh_thu_nhap()

Module này KHÔNG import app.py (tránh chạy lại toàn bộ ứng dụng Streamlit),
mà dùng lại đúng các quy ước sẵn có của app.py:
    - st.session_state.db_engine.get_connection()  (đa khách hàng, xem control_plane.py)
    - bảng `nhan_vien` sẵn có (id, ma_nv, ho_ten, phong_ban_lam_viec, chuc_vu,
      chuc_danh_nghe, he_so_luong, luong_bao_hiem, phu_cap_chuc_vu, phu_cap_tnvk,
      phu_cap_tnn, ngay_vao_lam, trang_thai, loai_hop_dong, ma_so_bhxh, so_tai_khoan_nh)
    - chat_noi_bo.create_private_room() / send_payslip_message() để gửi phiếu lương
    - st.session_state.nhan_vien_id: id nhân viên đang đăng nhập (người gửi phiếu lương)
    - st.session_state.role: 'admin' được setup chính sách 3P; 'admin'/'kt_luong' được tính lương

TRIẾT LÝ THIẾT KẾ 3P
--------------------
Thu nhập 3P của 1 nhân viên/tháng = P1 (lương vị trí) + P2 (lương năng lực) + P3 (lương hiệu quả)
    - P1 (Pay for Position): dựa trên GIÁ TRỊ CÔNG VIỆC (Job Evaluation) -> xếp Ngạch (Job Grade)
      -> tra Khung/Thang lương P1 (Ngạch x Bậc) đã được Admin thiết lập & gán cho từng nhân viên.
    - P2 (Pay for Person): quỹ P2 mục tiêu = % (cấu hình) x P1, điều chỉnh theo hệ số năng lực
      thực tế của nhân viên so với chuẩn (Khung năng lực P2).
    - P3 (Pay for Performance): quỹ P3 mục tiêu = % (cấu hình) x P1, điều chỉnh theo tỷ lệ hoàn
      thành KPI kỳ đó (Hệ thống KPI P3, liên kết với Bản đồ chiến lược BSC).
Phụ cấp (chức vụ, thâm niên, trách nhiệm, khác) được cộng THÊM ngoài 3P.
Từ đó tính Tổng thu nhập -> Khấu trừ BHXH/BHYT/BHTN/Đoàn phí -> Thu nhập chịu thuế -> Giảm trừ
gia cảnh -> Thuế TNCN luỹ tiến từng phần -> Thực nhận.

Toàn bộ hằng số nghiệp vụ (tỷ lệ trọng số đánh giá công việc, % quỹ P2/P3, biểu thuế, tỷ lệ BH...)
đều có giá trị MẶC ĐỊNH tham khảo và ĐỀU CÓ THỂ chỉnh sửa qua màn "⚙️ Cấu hình lương & thuế"
để cập nhật theo quy định hiện hành hoặc chính sách riêng của công ty.
"""

import io
import json
from datetime import datetime, date

import streamlit as st
import psycopg2
import psycopg2.extras

try:
    import chat_noi_bo
except Exception:
    chat_noi_bo = None


# ============================================================================
# 0) HẰNG SỐ MẶC ĐỊNH — GỢI Ý RIÊNG CHO CƠ CẤU TỔ CHỨC CỦA CHL
# ============================================================================

# Cơ cấu phòng ban tham khảo của CHL (dùng để tạo sẵn dữ liệu mẫu khi setup lần đầu;
# danh sách phòng ban THẬT vẫn luôn lấy động từ bảng nhan_vien để không lệch dữ liệu).
CHL_PHONG_BAN_GOI_Y = [
    "Hội Đồng Quản Trị", "Ban Tổng Giám Đốc", "Phòng Hành Chính Nhân Sự",
    "Phòng Kinh Doanh", "Phòng Tài Chính", "Phòng Điều Độ",
    "Tổ Cơ Giới", "Đội Bốc Xếp", "Phòng KT - Cơ Điện", "Đội Bảo Vệ",
]

BSC_KHIA_CANH = ["Tài chính", "Khách hàng", "Quy trình nội bộ", "Học hỏi & Phát triển"]

BSC_GOI_Y = [
    ("Tài chính", "Tăng trưởng doanh thu khai thác cảng", "Doanh thu dịch vụ cảng biển", "Tăng ≥10%/năm", "Phòng Kinh Doanh"),
    ("Tài chính", "Kiểm soát chi phí vận hành", "Tỷ lệ chi phí/doanh thu", "≤ 65%", "Phòng Tài Chính"),
    ("Khách hàng", "Nâng cao chất lượng dịch vụ xếp dỡ", "Điểm hài lòng khách hàng (CSAT)", "≥ 90%", "Phòng Kinh Doanh"),
    ("Khách hàng", "Rút ngắn thời gian giải phóng tàu", "Thời gian tàu nằm cảng bình quân", "Giảm ≥5%/năm", "Phòng Điều Độ"),
    ("Quy trình nội bộ", "Đảm bảo an toàn lao động & thiết bị", "Số vụ tai nạn lao động / sự cố thiết bị", "0 vụ nghiêm trọng", "Phòng KT - Cơ Điện"),
    ("Quy trình nội bộ", "Tối ưu hiệu suất điều độ tàu - bãi", "Năng suất xếp dỡ (tấn/giờ)", "Tăng ≥8%/năm", "Phòng Điều Độ"),
    ("Học hỏi & Phát triển", "Phát triển năng lực đội ngũ vận hành", "Tỷ lệ CBNV đạt chuẩn năng lực vị trí", "≥ 85%", "Phòng Hành Chính Nhân Sự"),
    ("Học hỏi & Phát triển", "Xây dựng văn hoá hiệu suất cao", "Tỷ lệ nghỉ việc tự nguyện", "≤ 8%/năm", "Phòng Hành Chính Nhân Sự"),
]

# Ma trận chức năng gợi ý: phòng ban x nhóm chức năng chính, mức độ tham gia
CAP_THAM_GIA = ["Chủ trì (A)", "Phối hợp/Hỗ trợ (S)", "Được tham vấn (C)", "Được thông báo (I)", "Không liên quan"]
FUNCTIONAL_MATRIX_GOI_Y = [
    ("Ban Tổng Giám Đốc", "Hoạch định chiến lược & kế hoạch SXKD", "Chủ trì (A)"),
    ("Phòng Kinh Doanh", "Phát triển khách hàng, ký hợp đồng dịch vụ cảng", "Chủ trì (A)"),
    ("Phòng Điều Độ", "Điều phối tàu - cầu bến - bãi - phương tiện", "Chủ trì (A)"),
    ("Tổ Cơ Giới", "Vận hành cẩu, xe nâng, phương tiện xếp dỡ", "Chủ trì (A)"),
    ("Đội Bốc Xếp", "Thực hiện xếp dỡ hàng hoá trực tiếp", "Chủ trì (A)"),
    ("Phòng KT - Cơ Điện", "Bảo trì, sửa chữa thiết bị - cơ điện", "Chủ trì (A)"),
    ("Đội Bảo Vệ", "An ninh trật tự, an toàn khu vực cảng", "Chủ trì (A)"),
    ("Phòng Tài Chính", "Quản lý dòng tiền, công nợ, báo cáo tài chính", "Chủ trì (A)"),
    ("Phòng Hành Chính Nhân Sự", "Tuyển dụng, đào tạo, chính sách lương thưởng", "Chủ trì (A)"),
]

# Bộ yếu tố đánh giá giá trị công việc (Job Evaluation) — phương pháp cho điểm theo yếu tố
# (điểm 1-5 mỗi yếu tố x trọng số %, tổng quy đổi thang 100)
YEU_TO_DANH_GIA_CONG_VIEC = [
    ("Kiến thức chuyên môn & kỹ năng yêu cầu", 0.25),
    ("Kinh nghiệm thực tế yêu cầu", 0.15),
    ("Mức độ ra quyết định & giải quyết vấn đề", 0.15),
    ("Trách nhiệm với kết quả tài chính / tài sản", 0.15),
    ("Trách nhiệm quản lý & phát triển con người", 0.10),
    ("Phạm vi ảnh hưởng & mối quan hệ công việc", 0.10),
    ("Điều kiện làm việc & mức độ áp lực/rủi ro", 0.10),
]

# Khung xếp Ngạch (Job Grade) gợi ý theo tổng điểm (thang 100) — 7 ngạch tương ứng cơ cấu CHL
NGACH_GOI_Y = [
    (0, 39, "G1", "Ngạch 1 — Lao động phổ thông / Bốc xếp"),
    (40, 49, "G2", "Ngạch 2 — Nhân viên nghiệp vụ / Công nhân kỹ thuật"),
    (50, 59, "G3", "Ngạch 3 — Chuyên viên / Kỹ thuật viên"),
    (60, 69, "G4", "Ngạch 4 — Chuyên viên chính / Phó phòng"),
    (70, 79, "G5", "Ngạch 5 — Trưởng phòng"),
    (80, 89, "G6", "Ngạch 6 — Phó Tổng Giám đốc"),
    (90, 100, "G7", "Ngạch 7 — Tổng Giám đốc / HĐQT"),
]

# Khung năng lực P2 gợi ý (nhóm cốt lõi áp dụng toàn công ty + nhóm quản lý)
P2_NHOM_GOI_Y = [
    ("Cốt lõi (toàn công ty)", "Chính trực - Trung thực", "Tuân thủ quy định, trung thực trong công việc và báo cáo"),
    ("Cốt lõi (toàn công ty)", "Tinh thần trách nhiệm & An toàn lao động", "Ý thức tuân thủ an toàn, chịu trách nhiệm với công việc được giao"),
    ("Cốt lõi (toàn công ty)", "Làm việc nhóm & Phối hợp", "Hợp tác hiệu quả với đồng nghiệp, các phòng ban liên quan"),
    ("Cốt lõi (toàn công ty)", "Thích ứng & Học hỏi", "Chủ động học hỏi, thích ứng với thay đổi quy trình/công nghệ"),
    ("Quản lý (áp dụng cấp quản lý)", "Hoạch định & Tổ chức công việc", "Lập kế hoạch, phân công, kiểm soát tiến độ đội/phòng ban"),
    ("Quản lý (áp dụng cấp quản lý)", "Lãnh đạo & Phát triển đội ngũ", "Dẫn dắt, đào tạo, tạo động lực cho nhân viên"),
    ("Quản lý (áp dụng cấp quản lý)", "Ra quyết định & Chịu trách nhiệm", "Ra quyết định kịp thời, chịu trách nhiệm về kết quả"),
    ("Chuyên môn — Khai thác cảng", "Điều phối tàu - bãi - phương tiện", "Am hiểu quy trình điều độ, tối ưu năng suất xếp dỡ"),
    ("Chuyên môn — Kỹ thuật/Cơ điện", "Vận hành & Bảo trì thiết bị cảng", "Vận hành an toàn, bảo trì phòng ngừa thiết bị/cơ điện"),
    ("Chuyên môn — Tài chính", "Phân tích & Kiểm soát tài chính", "Phân tích số liệu, kiểm soát ngân sách, tuân thủ kế toán"),
]
P2_MO_TA_CAP_DO = {
    1: "Mới bắt đầu — cần hướng dẫn sát sao",
    2: "Cơ bản — thực hiện được công việc thường xuyên, đôi khi cần hỗ trợ",
    3: "Thành thạo — đáp ứng đầy đủ yêu cầu vị trí (mức chuẩn)",
    4: "Nâng cao — xử lý tốt tình huống phức tạp, hỗ trợ người khác",
    5: "Chuyên gia — dẫn dắt, chuẩn hoá, đào tạo lại cho tổ chức",
}

# Hệ thống KPI P3 gợi ý theo phòng ban, liên kết BSC
P3_KPI_GOI_Y = [
    ("KPI-KD01", "Doanh thu dịch vụ cảng biển", "Phòng Kinh Doanh", "Tài chính", "Triệu đồng", 0.30, "Tháng"),
    ("KPI-KD02", "Số hợp đồng/khách hàng mới", "Phòng Kinh Doanh", "Khách hàng", "Hợp đồng", 0.20, "Tháng"),
    ("KPI-DD01", "Năng suất xếp dỡ bình quân", "Phòng Điều Độ", "Quy trình nội bộ", "Tấn/giờ", 0.30, "Tháng"),
    ("KPI-DD02", "Thời gian tàu nằm cảng bình quân", "Phòng Điều Độ", "Khách hàng", "Giờ", 0.20, "Tháng"),
    ("KPI-CG01", "Tỷ lệ thiết bị sẵn sàng vận hành", "Tổ Cơ Giới", "Quy trình nội bộ", "%", 0.30, "Tháng"),
    ("KPI-BX01", "Sản lượng xếp dỡ hoàn thành", "Đội Bốc Xếp", "Quy trình nội bộ", "Tấn", 0.35, "Tháng"),
    ("KPI-KT01", "Tỷ lệ sự cố thiết bị được xử lý đúng hạn", "Phòng KT - Cơ Điện", "Quy trình nội bộ", "%", 0.30, "Tháng"),
    ("KPI-BV01", "Số vụ mất an ninh trật tự", "Đội Bảo Vệ", "Quy trình nội bộ", "Vụ", 0.30, "Tháng"),
    ("KPI-TC01", "Tỷ lệ chi phí/doanh thu", "Phòng Tài Chính", "Tài chính", "%", 0.30, "Tháng"),
    ("KPI-NS01", "Tỷ lệ CBNV đạt chuẩn năng lực vị trí", "Phòng Hành Chính Nhân Sự", "Học hỏi & Phát triển", "%", 0.25, "Quý"),
]

# Biểu thuế TNCN luỹ tiến từng phần (Điều 22 Luật Thuế TNCN 2007, sửa đổi 2012)
BAC_THUE_TNCN_MAC_DINH = [
    {"tu": 0, "den": 5_000_000, "thue_suat": 0.05},
    {"tu": 5_000_000, "den": 10_000_000, "thue_suat": 0.10},
    {"tu": 10_000_000, "den": 18_000_000, "thue_suat": 0.15},
    {"tu": 18_000_000, "den": 32_000_000, "thue_suat": 0.20},
    {"tu": 32_000_000, "den": 52_000_000, "thue_suat": 0.25},
    {"tu": 52_000_000, "den": 80_000_000, "thue_suat": 0.30},
    {"tu": 80_000_000, "den": None, "thue_suat": 0.35},
]

CAU_HINH_MAC_DINH = {
    # Bảo hiểm bắt buộc — phần Người lao động đóng (khấu trừ vào lương)
    "bhxh_nld": 0.08,
    "bhyt_nld": 0.015,
    "bhtn_nld": 0.01,
    "doan_phi_nld": 0.01,          # Đoàn phí công đoàn (tuỳ chọn, mặc định 1%)
    "ap_dung_doan_phi": False,
    # Trần đóng bảo hiểm
    "luong_co_so": 2_340_000,       # dùng làm trần BHXH/BHYT = 20 x lương cơ sở
    "luong_toi_thieu_vung": 4_960_000,  # dùng làm trần BHTN = 20 x lương tối thiểu vùng
    # Giảm trừ gia cảnh thuế TNCN
    "giam_tru_ban_than": 11_000_000,
    "giam_tru_nguoi_phu_thuoc": 4_400_000,
    # Tỷ lệ quỹ P2 / P3 mục tiêu so với P1 (lương vị trí)
    "ty_le_quy_p2": 0.20,
    "ty_le_quy_p3": 0.15,
    "chuan_diem_nang_luc": 3.0,     # điểm P2 đạt "chuẩn" (thang 1-5) -> hệ số điều chỉnh = 1.0
    "he_so_p2_min": 0.70,
    "he_so_p2_max": 1.30,
    "ty_le_p3_min": 0.0,
    "ty_le_p3_max": 1.50,
    # Phụ cấp thâm niên công ty tự quy định (không phải TNVK/TNN nhà nước)
    "phu_cap_tham_nien_nam": 100_000,   # VNĐ / năm công tác
    "phu_cap_tham_nien_toi_da_nam": 10,  # tối đa tính đến 10 năm
    "bac_thue_tncn": BAC_THUE_TNCN_MAC_DINH,
}


# ============================================================================
# 1) DB HELPERS
# ============================================================================

def _conn(dict_cursor=True):
    db = st.session_state.db_engine.get_connection()
    c = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor) if dict_cursor else db.cursor()
    return db, c


def ensure_3p_tables():
    """Tạo toàn bộ bảng phục vụ chính sách 3P nếu chưa có (an toàn khi gọi lại nhiều lần)."""
    if st.session_state.get("_3p_da_khoi_tao"):
        return
    db, c = _conn(dict_cursor=False)
    try:
        c.execute("""
        CREATE TABLE IF NOT EXISTS tp_bsc_strategy_map (
            id SERIAL PRIMARY KEY,
            khia_canh TEXT NOT NULL,
            muc_tieu TEXT NOT NULL,
            chi_so_do_luong TEXT,
            chi_tieu TEXT,
            phong_ban_phu_trach TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_functional_matrix (
            id SERIAL PRIMARY KEY,
            phong_ban TEXT NOT NULL,
            nhom_chuc_nang TEXT NOT NULL,
            muc_do_tham_gia TEXT,
            bsc_id INTEGER REFERENCES tp_bsc_strategy_map(id) ON DELETE SET NULL,
            ghi_chu TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_job_description (
            id SERIAL PRIMARY KEY,
            ma_chuc_danh TEXT UNIQUE NOT NULL,
            ten_chuc_danh TEXT NOT NULL,
            phong_ban TEXT,
            bao_cao_cho TEXT,
            muc_tieu_cong_viec TEXT,
            nhiem_vu_chinh TEXT,
            yeu_cau_trinh_do TEXT,
            yeu_cau_kinh_nghiem TEXT,
            yeu_cau_ky_nang TEXT,
            dieu_kien_lam_viec TEXT,
            trang_thai TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_job_evaluation (
            id SERIAL PRIMARY KEY,
            ma_chuc_danh TEXT REFERENCES tp_job_description(ma_chuc_danh) ON DELETE CASCADE,
            diem_chi_tiet JSONB,
            tong_diem NUMERIC(6,2),
            ma_ngach TEXT,
            ten_ngach TEXT,
            nguoi_danh_gia TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_p1_salary_scale (
            id SERIAL PRIMARY KEY,
            ma_ngach TEXT NOT NULL,
            ten_ngach TEXT,
            bac INTEGER NOT NULL,
            he_so NUMERIC(6,3),
            luong_p1 NUMERIC(14,0) NOT NULL,
            ghi_chu TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(ma_ngach, bac)
        );

        CREATE TABLE IF NOT EXISTS tp_p2_competency (
            id SERIAL PRIMARY KEY,
            nhom_nang_luc TEXT NOT NULL,
            ten_nang_luc TEXT NOT NULL,
            mo_ta TEXT,
            ap_dung_chuc_danh TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_employee_p2_score (
            id SERIAL PRIMARY KEY,
            nhan_vien_id INTEGER NOT NULL,
            thang INTEGER NOT NULL,
            nam INTEGER NOT NULL,
            diem_nang_luc NUMERIC(4,2) NOT NULL,
            nguoi_danh_gia TEXT,
            ghi_chu TEXT,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(nhan_vien_id, thang, nam)
        );

        CREATE TABLE IF NOT EXISTS tp_p3_kpi_system (
            id SERIAL PRIMARY KEY,
            ma_kpi TEXT UNIQUE NOT NULL,
            ten_kpi TEXT NOT NULL,
            phong_ban TEXT,
            khia_canh_bsc TEXT,
            don_vi_tinh TEXT,
            trong_so NUMERIC(5,2),
            tan_suat TEXT DEFAULT 'Tháng',
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_employee_p3_score (
            id SERIAL PRIMARY KEY,
            nhan_vien_id INTEGER NOT NULL,
            thang INTEGER NOT NULL,
            nam INTEGER NOT NULL,
            ty_le_hoan_thanh NUMERIC(6,2) NOT NULL,
            ghi_chu TEXT,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(nhan_vien_id, thang, nam)
        );

        CREATE TABLE IF NOT EXISTS tp_employee_p1_assignment (
            id SERIAL PRIMARY KEY,
            nhan_vien_id INTEGER UNIQUE NOT NULL,
            ma_chuc_danh TEXT,
            ma_ngach TEXT NOT NULL,
            bac INTEGER NOT NULL,
            hieu_luc_tu DATE DEFAULT CURRENT_DATE,
            updated_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_employee_dependents (
            id SERIAL PRIMARY KEY,
            nhan_vien_id INTEGER NOT NULL,
            ho_ten TEXT NOT NULL,
            quan_he TEXT,
            ngay_sinh DATE,
            created_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_payroll_config (
            id SERIAL PRIMARY KEY,
            cau_hinh JSONB NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        );

        CREATE TABLE IF NOT EXISTS tp_policy_version (
            id SERIAL PRIMARY KEY,
            phien_ban INTEGER NOT NULL,
            ngay_xuat_ban TIMESTAMP DEFAULT NOW(),
            nguoi_xuat_ban TEXT,
            trang_thai TEXT DEFAULT 'published',
            thong_ke JSONB,
            ghi_chu TEXT
        );

        CREATE TABLE IF NOT EXISTS tp_payroll_period (
            id SERIAL PRIMARY KEY,
            thang INTEGER NOT NULL,
            nam INTEGER NOT NULL,
            trang_thai TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(thang, nam)
        );

        CREATE TABLE IF NOT EXISTS tp_payroll_detail (
            id SERIAL PRIMARY KEY,
            period_id INTEGER REFERENCES tp_payroll_period(id) ON DELETE CASCADE,
            nhan_vien_id INTEGER NOT NULL,
            p1_luong_vi_tri NUMERIC(14,0) DEFAULT 0,
            p2_luong_nang_luc NUMERIC(14,0) DEFAULT 0,
            p3_luong_hieu_qua NUMERIC(14,0) DEFAULT 0,
            phu_cap_chuc_vu NUMERIC(14,0) DEFAULT 0,
            phu_cap_tham_nien NUMERIC(14,0) DEFAULT 0,
            phu_cap_trach_nhiem NUMERIC(14,0) DEFAULT 0,
            phu_cap_khac NUMERIC(14,0) DEFAULT 0,
            tong_thu_nhap NUMERIC(14,0) DEFAULT 0,
            luong_dong_bh NUMERIC(14,0) DEFAULT 0,
            bhxh_nld NUMERIC(14,0) DEFAULT 0,
            bhyt_nld NUMERIC(14,0) DEFAULT 0,
            bhtn_nld NUMERIC(14,0) DEFAULT 0,
            doan_phi NUMERIC(14,0) DEFAULT 0,
            so_nguoi_phu_thuoc INTEGER DEFAULT 0,
            giam_tru_gia_canh NUMERIC(14,0) DEFAULT 0,
            thu_nhap_tinh_thue NUMERIC(14,0) DEFAULT 0,
            thue_tncn NUMERIC(14,0) DEFAULT 0,
            tong_khau_tru NUMERIC(14,0) DEFAULT 0,
            thuc_nhan NUMERIC(14,0) DEFAULT 0,
            chi_tiet JSONB,
            ngay_tinh TIMESTAMP DEFAULT NOW(),
            nguoi_tinh TEXT,
            da_gui_chat BOOLEAN DEFAULT FALSE,
            UNIQUE(period_id, nhan_vien_id)
        );
        """)
        db.commit()
        st.session_state["_3p_da_khoi_tao"] = True
    except Exception as e:
        db.rollback()
        st.error(f"Lỗi khởi tạo bảng dữ liệu chính sách 3P: {e}")
    finally:
        db.close()


# ---------- CRUD tiện ích ----------
def _fetch_all(table, order_by="id", where="", params=None):
    db, c = _conn()
    try:
        c.execute(f"SELECT * FROM {table} {where} ORDER BY {order_by}", params or ())
        return c.fetchall()
    finally:
        db.close()


def _insert(table, fields: dict):
    db, c = _conn(dict_cursor=False)
    try:
        cols = list(fields.keys())
        vals = [fields[k] for k in cols]
        placeholders = ", ".join(["%s"] * len(cols))
        c.execute(f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({placeholders}) RETURNING id", vals)
        new_id = c.fetchone()[0]
        db.commit()
        return new_id
    finally:
        db.close()


def _delete(table, row_id):
    db, c = _conn(dict_cursor=False)
    try:
        c.execute(f"DELETE FROM {table} WHERE id = %s", (row_id,))
        db.commit()
    finally:
        db.close()


# ============================================================================
# 2) NGHIỆP VỤ TÍNH TOÁN
# ============================================================================

def goi_y_ngach(tong_diem: float):
    for lo, hi, ma, ten in NGACH_GOI_Y:
        if lo <= tong_diem <= hi:
            return ma, ten
    return NGACH_GOI_Y[-1][2], NGACH_GOI_Y[-1][3]


def goi_y_khung_p1(luong_bac1_ngach1=6_500_000, so_bac=5, buoc_bac=0.06, buoc_ngach=0.18):
    """Sinh khung Thang lương P1 gợi ý (7 ngạch x so_bac bậc) bằng cấp số nhân —
    phương pháp phổ biến trong xây dựng thang bảng lương 3P."""
    rows = []
    for idx, (_, _, ma, ten) in enumerate(NGACH_GOI_Y):
        grade_min = luong_bac1_ngach1 * ((1 + buoc_ngach) ** idx)
        for bac in range(1, so_bac + 1):
            luong = grade_min * ((1 + buoc_bac) ** (bac - 1))
            he_so = round(luong / luong_bac1_ngach1, 3)
            rows.append({
                "ma_ngach": ma, "ten_ngach": ten, "bac": bac,
                "he_so": he_so, "luong_p1": int(round(luong / 1000) * 1000),
            })
    return rows


def tinh_he_so_p2(diem_tb, cfg):
    chuan = cfg.get("chuan_diem_nang_luc", 3.0)
    he_so = diem_tb / chuan if chuan else 1.0
    return max(cfg.get("he_so_p2_min", 0.7), min(cfg.get("he_so_p2_max", 1.3), he_so))


def tinh_ty_le_p3(ty_le_hoan_thanh_pct, cfg):
    """ty_le_hoan_thanh_pct: 0-150(%) -> trả về hệ số 0-1.5"""
    he_so = (ty_le_hoan_thanh_pct or 0) / 100.0
    return max(cfg.get("ty_le_p3_min", 0.0), min(cfg.get("ty_le_p3_max", 1.5), he_so))


def tinh_thue_tncn(thu_nhap_tinh_thue, bac_thue=None):
    """Tính thuế TNCN luỹ tiến từng phần. Trả về (tổng thuế, list chi tiết từng bậc)."""
    bac_thue = bac_thue or BAC_THUE_TNCN_MAC_DINH
    if thu_nhap_tinh_thue <= 0:
        return 0, []
    tong_thue = 0
    chi_tiet = []
    for bac in bac_thue:
        tu, den, ts = bac["tu"], bac["den"], bac["thue_suat"]
        if thu_nhap_tinh_thue <= tu:
            continue
        muc_tren = den if (den is not None and thu_nhap_tinh_thue > den) else thu_nhap_tinh_thue
        phan_chiu_thue = max(0, muc_tren - tu)
        if phan_chiu_thue <= 0:
            continue
        thue_bac = phan_chiu_thue * ts
        tong_thue += thue_bac
        chi_tiet.append({"tu": tu, "den": den, "thue_suat": ts, "phan_chiu_thue": phan_chiu_thue, "thue": thue_bac})
    return round(tong_thue), chi_tiet


def lay_cau_hinh_luong():
    rows = _fetch_all("tp_payroll_config", order_by="id DESC")
    if rows:
        cfg = dict(CAU_HINH_MAC_DINH)
        cfg.update(rows[0]["cau_hinh"] or {})
        return cfg
    # Chưa có -> seed mặc định
    _insert("tp_payroll_config", {"cau_hinh": json.dumps(CAU_HINH_MAC_DINH)})
    return dict(CAU_HINH_MAC_DINH)


def luu_cau_hinh_luong(cfg: dict):
    db, c = _conn(dict_cursor=False)
    try:
        c.execute("INSERT INTO tp_payroll_config (cau_hinh) VALUES (%s)", (json.dumps(cfg),))
        db.commit()
    finally:
        db.close()


def lay_gan_p1(nhan_vien_id):
    rows = _fetch_all("tp_employee_p1_assignment", where=f"WHERE nhan_vien_id = {int(nhan_vien_id)}")
    return rows[0] if rows else None


def lay_p1_luong(ma_ngach, bac):
    rows = _fetch_all("tp_p1_salary_scale", where=f"WHERE ma_ngach = '{ma_ngach}' AND bac = {int(bac)}")
    return rows[0]["luong_p1"] if rows else 0


def lay_diem_p2(nhan_vien_id, thang, nam):
    rows = _fetch_all("tp_employee_p2_score",
                       where=f"WHERE nhan_vien_id = {int(nhan_vien_id)} AND thang = {int(thang)} AND nam = {int(nam)}")
    return float(rows[0]["diem_nang_luc"]) if rows else None


def lay_ty_le_p3(nhan_vien_id, thang, nam):
    rows = _fetch_all("tp_employee_p3_score",
                       where=f"WHERE nhan_vien_id = {int(nhan_vien_id)} AND thang = {int(thang)} AND nam = {int(nam)}")
    return float(rows[0]["ty_le_hoan_thanh"]) if rows else None


def dem_nguoi_phu_thuoc(nhan_vien_id):
    rows = _fetch_all("tp_employee_dependents", where=f"WHERE nhan_vien_id = {int(nhan_vien_id)}")
    return len(rows)


def so_nam_cong_tac(ngay_vao_lam):
    if not ngay_vao_lam:
        return 0
    today = date.today()
    if hasattr(ngay_vao_lam, "year"):
        d = ngay_vao_lam
    else:
        return 0
    nam = today.year - d.year - ((today.month, today.day) < (d.month, d.day))
    return max(0, nam)


def _load_salary_module():
    """Nạp module công thức lương ĐANG ÁP DỤNG cho tenant hiện tại: ưu tiên file riêng
    'salary/salary_{ma_so_thue}.py' (đặt cùng cấp với app.py); nếu tenant chưa có file
    riêng, rơi về mặc định 'salary/salary_demo.py'.

    Đây là bản sao ĐỘC LẬP của _load_tenant_module_or_demo() trong app.py — không
    import trực tiếp từ app.py để tránh import vòng (app.py `import tinh_thu_nhap` ở
    mức module ngay khi khởi động; nếu tinh_thu_nhap.py quay lại `import app` sẽ chạy
    lại toàn bộ app.py giữa chừng và crash)."""
    import importlib
    tenant = st.session_state.get('tenant') or {}
    ma_so_thue = (tenant.get('ma_so_thue') or '').strip()

    module = None
    if ma_so_thue:
        try:
            module = importlib.import_module(f"salary.salary_{ma_so_thue}")
        except ModuleNotFoundError:
            module = None
        except Exception as e:
            print(f"Lỗi import salary.salary_{ma_so_thue}: {e}")
            module = None

    if module is None:
        try:
            module = importlib.import_module("salary.salary_demo")
        except Exception as e:
            print(f"Lỗi import salary.salary_demo (mặc định): {e}")
            module = None

    return module


def tinh_luong_nhan_vien(nv: dict, thang: int, nam: int, cfg: dict = None,
                          phu_cap_trach_nhiem=0, phu_cap_khac=0, ghi_chu=""):
    """Hàm lõi tính lương 3P đầy đủ cho 1 nhân viên trong 1 kỳ. Trả về dict chi tiết.

    LƯU Ý: đây là công thức 3P MẶC ĐỊNH thật sự đang chạy (được salary/salary_demo.py
    gọi lại qua tinh_luong()). Màn '💰 Tính thu nhập' (_ui_tinh_luong_thang() bên dưới)
    KHÔNG gọi thẳng hàm này nữa — nó gọi qua _load_salary_module() để tự động dùng
    đúng công thức riêng của tenant (nếu có), hoặc rơi về công thức 3P này làm mặc
    định. Hàm này vẫn giữ nguyên logic gốc, không đổi."""
    cfg = cfg or lay_cau_hinh_luong()
    nv_id = nv["id"]

    # ----- P1: lương vị trí -----
    gan = lay_gan_p1(nv_id)
    if gan:
        p1 = float(lay_p1_luong(gan["ma_ngach"], gan["bac"]) or 0)
        ma_ngach, bac = gan["ma_ngach"], gan["bac"]
    else:
        p1, ma_ngach, bac = 0.0, None, None

    # ----- P2: lương năng lực -----
    diem_p2 = lay_diem_p2(nv_id, thang, nam)
    if diem_p2 is None:
        diem_p2 = cfg.get("chuan_diem_nang_luc", 3.0)  # chưa đánh giá -> coi như đạt chuẩn
    he_so_p2 = tinh_he_so_p2(diem_p2, cfg)
    p2 = p1 * cfg.get("ty_le_quy_p2", 0.20) * he_so_p2

    # ----- P3: lương hiệu quả (KPI) -----
    ty_le_kpi = lay_ty_le_p3(nv_id, thang, nam)
    if ty_le_kpi is None:
        ty_le_kpi = 100.0  # chưa có KPI kỳ này -> mặc định 100% để không thiệt cho NLĐ
    he_so_p3 = tinh_ty_le_p3(ty_le_kpi, cfg)
    p3 = p1 * cfg.get("ty_le_quy_p3", 0.15) * he_so_p3

    # ----- Phụ cấp -----
    phu_cap_chuc_vu = float(nv.get("phu_cap_chuc_vu") or 0)
    so_nam = so_nam_cong_tac(nv.get("ngay_vao_lam"))
    so_nam_tinh = min(so_nam, cfg.get("phu_cap_tham_nien_toi_da_nam", 10))
    phu_cap_tham_nien = so_nam_tinh * cfg.get("phu_cap_tham_nien_nam", 0)

    tong_thu_nhap = p1 + p2 + p3 + phu_cap_chuc_vu + phu_cap_tham_nien + phu_cap_trach_nhiem + phu_cap_khac

    # ----- Khấu trừ BHXH/BHYT/BHTN -----
    luong_dong_bh = float(nv.get("luong_bao_hiem") or p1 or 0)
    tran_bhxh_bhyt = 20 * cfg.get("luong_co_so", 2_340_000)
    tran_bhtn = 20 * cfg.get("luong_toi_thieu_vung", 4_960_000)
    bh_xh_bhyt_base = min(luong_dong_bh, tran_bhxh_bhyt)
    bh_tn_base = min(luong_dong_bh, tran_bhtn)

    bhxh_nld = round(bh_xh_bhyt_base * cfg.get("bhxh_nld", 0.08))
    bhyt_nld = round(bh_xh_bhyt_base * cfg.get("bhyt_nld", 0.015))
    bhtn_nld = round(bh_tn_base * cfg.get("bhtn_nld", 0.01))
    doan_phi = round(tong_thu_nhap * cfg.get("doan_phi_nld", 0.01)) if cfg.get("ap_dung_doan_phi") else 0
    tong_khau_tru_bh = bhxh_nld + bhyt_nld + bhtn_nld + doan_phi

    # ----- Thuế TNCN -----
    so_nguoi_phu_thuoc = dem_nguoi_phu_thuoc(nv_id)
    giam_tru = cfg.get("giam_tru_ban_than", 11_000_000) + so_nguoi_phu_thuoc * cfg.get("giam_tru_nguoi_phu_thuoc", 4_400_000)
    thu_nhap_truoc_thue = tong_thu_nhap - tong_khau_tru_bh
    thu_nhap_tinh_thue = max(0, thu_nhap_truoc_thue - giam_tru)
    thue_tncn, chi_tiet_thue = tinh_thue_tncn(thu_nhap_tinh_thue, cfg.get("bac_thue_tncn"))

    tong_khau_tru = tong_khau_tru_bh + thue_tncn
    thuc_nhan = tong_thu_nhap - tong_khau_tru

    return {
        "nhan_vien_id": nv_id, "ma_nv": nv.get("ma_nv"), "ho_ten": nv.get("ho_ten"),
        "phong_ban": nv.get("phong_ban_lam_viec"), "chuc_vu": nv.get("chuc_vu"),
        "thang": thang, "nam": nam,
        "ma_ngach": ma_ngach, "bac": bac,
        "p1_luong_vi_tri": round(p1), "p2_luong_nang_luc": round(p2), "p3_luong_hieu_qua": round(p3),
        "diem_nang_luc": diem_p2, "he_so_p2": round(he_so_p2, 3),
        "ty_le_hoan_thanh_kpi": ty_le_kpi, "he_so_p3": round(he_so_p3, 3),
        "phu_cap_chuc_vu": round(phu_cap_chuc_vu), "phu_cap_tham_nien": round(phu_cap_tham_nien),
        "phu_cap_trach_nhiem": round(phu_cap_trach_nhiem), "phu_cap_khac": round(phu_cap_khac),
        "tong_thu_nhap": round(tong_thu_nhap),
        "luong_dong_bh": round(luong_dong_bh),
        "bhxh_nld": bhxh_nld, "bhyt_nld": bhyt_nld, "bhtn_nld": bhtn_nld, "doan_phi": doan_phi,
        "tong_khau_tru_bh": tong_khau_tru_bh,
        "so_nguoi_phu_thuoc": so_nguoi_phu_thuoc, "giam_tru_gia_canh": round(giam_tru),
        "thu_nhap_tinh_thue": round(thu_nhap_tinh_thue), "thue_tncn": thue_tncn,
        "chi_tiet_thue": chi_tiet_thue,
        "tong_khau_tru": round(tong_khau_tru), "thuc_nhan": round(thuc_nhan),
        "ghi_chu": ghi_chu,
    }


def lay_hoac_tao_ky_luong(thang, nam):
    rows = _fetch_all("tp_payroll_period", where=f"WHERE thang={int(thang)} AND nam={int(nam)}")
    if rows:
        return rows[0]["id"]
    return _insert("tp_payroll_period", {"thang": thang, "nam": nam})


def luu_bang_luong(period_id, breakdown: dict, nguoi_tinh=""):
    db, c = _conn(dict_cursor=False)
    try:
        c.execute("""
            INSERT INTO tp_payroll_detail (
                period_id, nhan_vien_id, p1_luong_vi_tri, p2_luong_nang_luc, p3_luong_hieu_qua,
                phu_cap_chuc_vu, phu_cap_tham_nien, phu_cap_trach_nhiem, phu_cap_khac,
                tong_thu_nhap, luong_dong_bh, bhxh_nld, bhyt_nld, bhtn_nld, doan_phi,
                so_nguoi_phu_thuoc, giam_tru_gia_canh, thu_nhap_tinh_thue, thue_tncn,
                tong_khau_tru, thuc_nhan, chi_tiet, nguoi_tinh
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (period_id, nhan_vien_id) DO UPDATE SET
                p1_luong_vi_tri=EXCLUDED.p1_luong_vi_tri, p2_luong_nang_luc=EXCLUDED.p2_luong_nang_luc,
                p3_luong_hieu_qua=EXCLUDED.p3_luong_hieu_qua, phu_cap_chuc_vu=EXCLUDED.phu_cap_chuc_vu,
                phu_cap_tham_nien=EXCLUDED.phu_cap_tham_nien, phu_cap_trach_nhiem=EXCLUDED.phu_cap_trach_nhiem,
                phu_cap_khac=EXCLUDED.phu_cap_khac, tong_thu_nhap=EXCLUDED.tong_thu_nhap,
                luong_dong_bh=EXCLUDED.luong_dong_bh, bhxh_nld=EXCLUDED.bhxh_nld, bhyt_nld=EXCLUDED.bhyt_nld,
                bhtn_nld=EXCLUDED.bhtn_nld, doan_phi=EXCLUDED.doan_phi,
                so_nguoi_phu_thuoc=EXCLUDED.so_nguoi_phu_thuoc, giam_tru_gia_canh=EXCLUDED.giam_tru_gia_canh,
                thu_nhap_tinh_thue=EXCLUDED.thu_nhap_tinh_thue, thue_tncn=EXCLUDED.thue_tncn,
                tong_khau_tru=EXCLUDED.tong_khau_tru, thuc_nhan=EXCLUDED.thuc_nhan,
                chi_tiet=EXCLUDED.chi_tiet, nguoi_tinh=EXCLUDED.nguoi_tinh, ngay_tinh=NOW()
        """, (
            period_id, breakdown["nhan_vien_id"], breakdown["p1_luong_vi_tri"], breakdown["p2_luong_nang_luc"],
            breakdown["p3_luong_hieu_qua"], breakdown["phu_cap_chuc_vu"], breakdown["phu_cap_tham_nien"],
            breakdown["phu_cap_trach_nhiem"], breakdown["phu_cap_khac"], breakdown["tong_thu_nhap"],
            breakdown["luong_dong_bh"], breakdown["bhxh_nld"], breakdown["bhyt_nld"], breakdown["bhtn_nld"],
            breakdown["doan_phi"], breakdown["so_nguoi_phu_thuoc"], breakdown["giam_tru_gia_canh"],
            breakdown["thu_nhap_tinh_thue"], breakdown["thue_tncn"], breakdown["tong_khau_tru"],
            breakdown["thuc_nhan"], json.dumps(breakdown, default=str), nguoi_tinh,
        ))
        db.commit()
    finally:
        db.close()


def chinh_sach_da_xuat_ban():
    rows = _fetch_all("tp_policy_version", where="WHERE trang_thai='published'", order_by="id DESC")
    return rows[0] if rows else None


def xuat_ban_chinh_sach(nguoi_xuat_ban):
    thong_ke = {
        "so_muc_tieu_bsc": len(_fetch_all("tp_bsc_strategy_map")),
        "so_dong_ma_tran_chuc_nang": len(_fetch_all("tp_functional_matrix")),
        "so_ban_mo_ta_cong_viec": len(_fetch_all("tp_job_description")),
        "so_danh_gia_gia_tri_cv": len(_fetch_all("tp_job_evaluation")),
        "so_dong_khung_p1": len(_fetch_all("tp_p1_salary_scale")),
        "so_nang_luc_p2": len(_fetch_all("tp_p2_competency")),
        "so_kpi_p3": len(_fetch_all("tp_p3_kpi_system")),
    }
    ban_truoc = _fetch_all("tp_policy_version", order_by="phien_ban DESC")
    phien_ban_moi = (ban_truoc[0]["phien_ban"] + 1) if ban_truoc else 1
    db, c = _conn(dict_cursor=False)
    try:
        c.execute("UPDATE tp_policy_version SET trang_thai='superseded' WHERE trang_thai='published'")
        c.execute("""INSERT INTO tp_policy_version (phien_ban, nguoi_xuat_ban, trang_thai, thong_ke)
                     VALUES (%s,%s,'published',%s)""", (phien_ban_moi, nguoi_xuat_ban, json.dumps(thong_ke)))
        db.commit()
    finally:
        db.close()
    return phien_ban_moi, thong_ke


# ============================================================================
# 3) XUẤT EXCEL / PDF
# ============================================================================

def export_bang_luong_excel(rows, thang, nam, ten_cty="") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"BangLuong_T{thang}_{nam}"

    thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    headers = ["STT", "Mã NV", "Họ tên", "Phòng ban", "P1 (Vị trí)", "P2 (Năng lực)", "P3 (Hiệu quả)",
               "Phụ cấp CV", "Phụ cấp TN", "Phụ cấp khác", "Tổng thu nhập", "BHXH", "BHYT", "BHTN",
               "Đoàn phí", "Giảm trừ GC", "TN tính thuế", "Thuế TNCN", "Thực nhận"]

    title = f"BẢNG LƯƠNG 3P — {ten_cty or ''} — Tháng {thang}/{nam}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, name="Times New Roman")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws["A2"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True, name="Times New Roman")
    ws["A2"].alignment = Alignment(horizontal="center")

    header_row = 4
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, name="Times New Roman", color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin

    data_row = header_row + 1
    for i, r in enumerate(rows):
        values = [
            i + 1, r.get("ma_nv"), r.get("ho_ten"), r.get("phong_ban"),
            r.get("p1_luong_vi_tri"), r.get("p2_luong_nang_luc"), r.get("p3_luong_hieu_qua"),
            r.get("phu_cap_chuc_vu"), r.get("phu_cap_tham_nien"),
            (r.get("phu_cap_trach_nhiem", 0) or 0) + (r.get("phu_cap_khac", 0) or 0),
            r.get("tong_thu_nhap"), r.get("bhxh_nld"), r.get("bhyt_nld"), r.get("bhtn_nld"),
            r.get("doan_phi"), r.get("giam_tru_gia_canh"), r.get("thu_nhap_tinh_thue"),
            r.get("thue_tncn"), r.get("thuc_nhan"),
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=data_row + i, column=col_idx, value=v)
            cell.font = Font(size=9, name="Times New Roman")
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col_idx <= 4 else "right")
            if col_idx >= 5 and isinstance(v, (int, float)):
                cell.number_format = "#,##0"

    footer_row = data_row + len(rows) + 1
    ws.cell(row=footer_row, column=1, value=f"Tổng cộng: {len(rows)} nhân viên").font = Font(bold=True, name="Times New Roman")
    ws.cell(row=footer_row + 2, column=1, value="Người lập bảng lương").font = Font(bold=True, name="Times New Roman")
    ws.cell(row=footer_row + 2, column=int(len(headers) / 2) + 1, value="Giám đốc phê duyệt").font = Font(bold=True, name="Times New Roman")

    widths = [5, 10, 20, 18] + [13] * (len(headers) - 4)
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _dang_ky_font_unicode():
    """Cố gắng đăng ký 1 font TTF hỗ trợ tiếng Việt cho reportlab; nếu không tìm thấy,
    trả về font mặc định (dấu tiếng Việt có thể không hiển thị đúng — nên cài thêm
    1 font .ttf hỗ trợ Unicode, ví dụ DejaVuSans, vào thư mục assets của dự án)."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os as _os
    ung_vien = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "assets/fonts/DejaVuSans.ttf",
    ]
    da_dang_ky = False
    for path in ung_vien:
        if _os.path.exists(path) and "Bold" not in path:
            try:
                pdfmetrics.registerFont(TTFont("VN-Unicode", path))
                da_dang_ky = True
            except Exception:
                pass
    return "VN-Unicode" if da_dang_ky else "Helvetica"


def export_phieu_luong_pdf(breakdown: dict, ten_cty="") -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    font_name = _dang_ky_font_unicode()
    styles = getSampleStyleSheet()
    style_title = ParagraphStyle("title", parent=styles["Title"], fontName=font_name, fontSize=14)
    style_normal = ParagraphStyle("normal", parent=styles["Normal"], fontName=font_name, fontSize=10)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    elements = []

    elements.append(Paragraph(f"{ten_cty or 'CÔNG TY'}", style_normal))
    elements.append(Paragraph(f"PHIẾU LƯƠNG THÁNG {breakdown['thang']}/{breakdown['nam']}", style_title))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        f"Họ tên: <b>{breakdown.get('ho_ten', '')}</b> &nbsp;&nbsp; Mã NV: {breakdown.get('ma_nv', '')} "
        f"&nbsp;&nbsp; Phòng ban: {breakdown.get('phong_ban', '')} &nbsp;&nbsp; Chức vụ: {breakdown.get('chuc_vu', '')}",
        style_normal))
    elements.append(Spacer(1, 10))

    def fmt(v):
        try:
            return f"{v:,.0f}"
        except Exception:
            return str(v)

    data = [
        ["Khoản mục", "Số tiền (VNĐ)"],
        ["P1 — Lương vị trí", fmt(breakdown["p1_luong_vi_tri"])],
        ["P2 — Lương năng lực", fmt(breakdown["p2_luong_nang_luc"])],
        ["P3 — Lương hiệu quả (KPI)", fmt(breakdown["p3_luong_hieu_qua"])],
        ["Phụ cấp chức vụ", fmt(breakdown["phu_cap_chuc_vu"])],
        ["Phụ cấp thâm niên", fmt(breakdown["phu_cap_tham_nien"])],
        ["Phụ cấp trách nhiệm", fmt(breakdown["phu_cap_trach_nhiem"])],
        ["Phụ cấp khác", fmt(breakdown["phu_cap_khac"])],
        ["TỔNG THU NHẬP", fmt(breakdown["tong_thu_nhap"])],
        ["BHXH (NLĐ đóng)", f"-{fmt(breakdown['bhxh_nld'])}"],
        ["BHYT (NLĐ đóng)", f"-{fmt(breakdown['bhyt_nld'])}"],
        ["BHTN (NLĐ đóng)", f"-{fmt(breakdown['bhtn_nld'])}"],
        ["Đoàn phí", f"-{fmt(breakdown['doan_phi'])}"],
        ["Giảm trừ gia cảnh", fmt(breakdown["giam_tru_gia_canh"])],
        ["Thu nhập tính thuế", fmt(breakdown["thu_nhap_tinh_thue"])],
        ["Thuế TNCN", f"-{fmt(breakdown['thue_tncn'])}"],
        ["THỰC NHẬN", fmt(breakdown["thuc_nhan"])],
    ]
    table = Table(data, colWidths=[9 * cm, 6 * cm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BACKGROUND", (0, 8), (-1, 8), colors.HexColor("#DCE6F1")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#DCE6F1")),
        ("FONTNAME", (0, 8), (-1, 8), font_name),
        ("FONTNAME", (0, -1), (-1, -1), font_name),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 16))
    elements.append(Paragraph(f"Ngày xuất phiếu: {datetime.now().strftime('%d/%m/%Y %H:%M')}", style_normal))

    doc.build(elements)
    return buf.getvalue()


# ============================================================================
# 4) GỬI PHIẾU LƯƠNG QUA CHAT NỘI BỘ
# ============================================================================

def gui_phieu_luong_qua_chat(nhan_vien_id, breakdown, sender_id):
    if chat_noi_bo is None:
        return False, "Chưa cấu hình module chat_noi_bo."
    if not sender_id:
        return False, "Tài khoản hiện tại chưa gắn với hồ sơ nhân viên nên không thể gửi chat."
    room_id = chat_noi_bo.create_private_room(sender_id, nhan_vien_id)
    if not room_id:
        return False, "Không tạo được phòng chat riêng với nhân viên này."
    ok = chat_noi_bo.send_payslip_message(
        room_id, sender_id, breakdown.get("ho_ten", ""), breakdown["thang"], breakdown["nam"],
        {
            "luong_co_ban": breakdown["p1_luong_vi_tri"],
            "phu_cap_chuc_vu": breakdown["phu_cap_chuc_vu"],
            "phu_cap_tnvk": breakdown["p2_luong_nang_luc"],
            "phu_cap_tnn": breakdown["p3_luong_hieu_qua"],
            "tong": breakdown["tong_thu_nhap"],
            "bhxh": breakdown["bhxh_nld"],
            "bhyt": breakdown["bhyt_nld"],
            "bhtn": breakdown["bhtn_nld"],
            "thuc_nhan": breakdown["thuc_nhan"],
        },
    )
    return ok, ("Đã gửi phiếu lương qua chat nội bộ." if ok else "Gửi phiếu lương thất bại.")


# ============================================================================
# 5) UI — LẤY DANH SÁCH NHÂN VIÊN (DÙNG CHUNG)
# ============================================================================

def _lay_ds_nhan_vien(active_only=True):
    db, c = _conn()
    try:
        if active_only:
            c.execute("""SELECT id, ma_nv, ho_ten, phong_ban_lam_viec, chuc_vu, chuc_danh_nghe,
                                he_so_luong, luong_bao_hiem, phu_cap_chuc_vu, ngay_vao_lam, trang_thai
                         FROM nhan_vien WHERE trang_thai IN ('DANG_LAM','THU_VIEC') ORDER BY ho_ten""")
        else:
            c.execute("""SELECT id, ma_nv, ho_ten, phong_ban_lam_viec, chuc_vu, chuc_danh_nghe,
                                he_so_luong, luong_bao_hiem, phu_cap_chuc_vu, ngay_vao_lam, trang_thai
                         FROM nhan_vien ORDER BY ho_ten""")
        return c.fetchall()
    finally:
        db.close()


def _lay_ds_phong_ban_thuc_te():
    db, c = _conn()
    try:
        c.execute("""SELECT DISTINCT phong_ban_lam_viec FROM nhan_vien
                     WHERE phong_ban_lam_viec IS NOT NULL AND phong_ban_lam_viec != ''
                     ORDER BY phong_ban_lam_viec""")
        rows = [r["phong_ban_lam_viec"] for r in c.fetchall()]
        return rows or list(CHL_PHONG_BAN_GOI_Y)
    finally:
        db.close()


# ============================================================================
# 6) UI — WIZARD THIẾT LẬP CHÍNH SÁCH 3P (chỉ admin)
# ============================================================================

def _seed_neu_trong(table, rows_goi_y, insert_fn):
    if not _fetch_all(table):
        for r in rows_goi_y:
            insert_fn(r)


def _ui_bsc():
    st.markdown("#### 🧭 Bước 1 — Bản đồ chiến lược (BSC)")
    st.caption("Xác định mục tiêu chiến lược theo 4 khía cạnh Balanced Scorecard — là nền tảng để "
               "liên kết Ma trận chức năng phòng ban và Hệ thống KPI P3 phía sau.")

    if st.button("✨ Nạp gợi ý BSC mặc định cho CHL", key="bsc_seed"):
        if _fetch_all("tp_bsc_strategy_map"):
            st.warning("Đã có dữ liệu BSC — chỉ nạp gợi ý khi bảng đang trống. Hãy xoá dữ liệu cũ nếu muốn nạp lại.")
        else:
            for kc, mt, cs, ct, pb in BSC_GOI_Y:
                _insert("tp_bsc_strategy_map", {
                    "khia_canh": kc, "muc_tieu": mt, "chi_so_do_luong": cs,
                    "chi_tieu": ct, "phong_ban_phu_trach": pb,
                })
            st.success(f"Đã nạp {len(BSC_GOI_Y)} mục tiêu chiến lược gợi ý.")
            st.rerun()

    with st.form("form_them_bsc", clear_on_submit=True):
        c1, c2 = st.columns(2)
        khia_canh = c1.selectbox("Khía cạnh BSC", BSC_KHIA_CANH)
        phong_ban = c2.selectbox("Phòng ban phụ trách", _lay_ds_phong_ban_thuc_te())
        muc_tieu = st.text_input("Mục tiêu chiến lược")
        c3, c4 = st.columns(2)
        chi_so = c3.text_input("Chỉ số đo lường (KPI cấp chiến lược)")
        chi_tieu = c4.text_input("Chỉ tiêu mục tiêu")
        if st.form_submit_button("➕ Thêm mục tiêu", type="primary"):
            if muc_tieu:
                _insert("tp_bsc_strategy_map", {
                    "khia_canh": khia_canh, "muc_tieu": muc_tieu, "chi_so_do_luong": chi_so,
                    "chi_tieu": chi_tieu, "phong_ban_phu_trach": phong_ban,
                })
                st.success("Đã thêm mục tiêu chiến lược.")
                st.rerun()
            else:
                st.warning("Vui lòng nhập mục tiêu chiến lược.")

    rows = _fetch_all("tp_bsc_strategy_map", order_by="khia_canh, id")
    if rows:
        for kc in BSC_KHIA_CANH:
            nhom = [r for r in rows if r["khia_canh"] == kc]
            if not nhom:
                continue
            st.markdown(f"**{kc}**")
            for r in nhom:
                cc = st.columns([4, 3, 2, 2, 1])
                cc[0].write(r["muc_tieu"])
                cc[1].write(r["chi_so_do_luong"] or "")
                cc[2].write(r["chi_tieu"] or "")
                cc[3].write(r["phong_ban_phu_trach"] or "")
                if cc[4].button("🗑️", key=f"del_bsc_{r['id']}"):
                    _delete("tp_bsc_strategy_map", r["id"])
                    st.rerun()
    else:
        st.info("Chưa có dữ liệu BSC. Bấm nút gợi ý ở trên hoặc tự thêm mục tiêu.")


def _ui_functional_matrix():
    st.markdown("#### 🧩 Bước 2 — Ma trận chức năng các phòng ban")
    st.caption("Xác định phòng ban nào chịu trách nhiệm chính cho nhóm chức năng/quy trình nào — "
               "làm cơ sở xây dựng Bản mô tả công việc chuẩn xác theo đúng cơ cấu tổ chức.")

    if st.button("✨ Nạp gợi ý Ma trận chức năng cho CHL", key="fm_seed"):
        if _fetch_all("tp_functional_matrix"):
            st.warning("Đã có dữ liệu — chỉ nạp gợi ý khi bảng đang trống.")
        else:
            for pb, cn, muc in FUNCTIONAL_MATRIX_GOI_Y:
                _insert("tp_functional_matrix", {"phong_ban": pb, "nhom_chuc_nang": cn, "muc_do_tham_gia": muc})
            st.success(f"Đã nạp {len(FUNCTIONAL_MATRIX_GOI_Y)} dòng ma trận chức năng gợi ý.")
            st.rerun()

    with st.form("form_them_fm", clear_on_submit=True):
        c1, c2 = st.columns(2)
        phong_ban = c1.selectbox("Phòng ban", _lay_ds_phong_ban_thuc_te())
        muc_do = c2.selectbox("Mức độ tham gia", CAP_THAM_GIA)
        nhom_cn = st.text_input("Nhóm chức năng / Quy trình")
        ghi_chu = st.text_input("Ghi chú (tuỳ chọn)")
        if st.form_submit_button("➕ Thêm dòng", type="primary"):
            if nhom_cn:
                _insert("tp_functional_matrix", {
                    "phong_ban": phong_ban, "nhom_chuc_nang": nhom_cn,
                    "muc_do_tham_gia": muc_do, "ghi_chu": ghi_chu,
                })
                st.rerun()

    rows = _fetch_all("tp_functional_matrix", order_by="phong_ban, id")
    if rows:
        import pandas as pd
        df = pd.DataFrame([{"Phòng ban": r["phong_ban"], "Nhóm chức năng": r["nhom_chuc_nang"],
                             "Mức độ tham gia": r["muc_do_tham_gia"], "Ghi chú": r["ghi_chu"]} for r in rows])
        st.dataframe(df, width='stretch', hide_index=True)
        del_id = st.selectbox("Chọn dòng để xoá (theo ID)", ["--"] + [str(r["id"]) for r in rows], key="fm_del_sel")
        if del_id != "--" and st.button("🗑️ Xoá dòng đã chọn", key="fm_del_btn"):
            _delete("tp_functional_matrix", int(del_id))
            st.rerun()
    else:
        st.info("Chưa có dữ liệu Ma trận chức năng.")


def _ui_jd():
    st.markdown("#### 📋 Bước 3 — Bản mô tả công việc (JD)")
    ds_pb = _lay_ds_phong_ban_thuc_te()

    with st.expander("➕ Thêm / Cập nhật bản mô tả công việc", expanded=not bool(_fetch_all("tp_job_description"))):
        with st.form("form_jd", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            ma_cd = c1.text_input("Mã chức danh *", placeholder="VD: KD-CV01")
            ten_cd = c2.text_input("Tên chức danh *", placeholder="VD: Chuyên viên Kinh doanh")
            phong_ban = c3.selectbox("Phòng ban", ds_pb)
            bao_cao_cho = st.text_input("Báo cáo cho chức danh")
            muc_tieu = st.text_area("Mục tiêu công việc", height=70)
            nhiem_vu = st.text_area("Nhiệm vụ chính (mỗi dòng 1 nhiệm vụ)", height=100)
            c4, c5 = st.columns(2)
            trinh_do = c4.text_area("Yêu cầu trình độ", height=70)
            kinh_nghiem = c5.text_area("Yêu cầu kinh nghiệm", height=70)
            c6, c7 = st.columns(2)
            ky_nang = c6.text_area("Yêu cầu kỹ năng", height=70)
            dieu_kien = c7.text_area("Điều kiện làm việc", height=70)
            if st.form_submit_button("💾 Lưu bản mô tả công việc", type="primary"):
                if ma_cd and ten_cd:
                    db, c = _conn(dict_cursor=False)
                    try:
                        c.execute("""
                            INSERT INTO tp_job_description (ma_chuc_danh, ten_chuc_danh, phong_ban, bao_cao_cho,
                                muc_tieu_cong_viec, nhiem_vu_chinh, yeu_cau_trinh_do, yeu_cau_kinh_nghiem,
                                yeu_cau_ky_nang, dieu_kien_lam_viec)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT (ma_chuc_danh) DO UPDATE SET
                                ten_chuc_danh=EXCLUDED.ten_chuc_danh, phong_ban=EXCLUDED.phong_ban,
                                bao_cao_cho=EXCLUDED.bao_cao_cho, muc_tieu_cong_viec=EXCLUDED.muc_tieu_cong_viec,
                                nhiem_vu_chinh=EXCLUDED.nhiem_vu_chinh, yeu_cau_trinh_do=EXCLUDED.yeu_cau_trinh_do,
                                yeu_cau_kinh_nghiem=EXCLUDED.yeu_cau_kinh_nghiem,
                                yeu_cau_ky_nang=EXCLUDED.yeu_cau_ky_nang,
                                dieu_kien_lam_viec=EXCLUDED.dieu_kien_lam_viec, updated_at=NOW()
                        """, (ma_cd.strip().upper(), ten_cd, phong_ban, bao_cao_cho, muc_tieu, nhiem_vu,
                              trinh_do, kinh_nghiem, ky_nang, dieu_kien))
                        db.commit()
                        st.success(f"Đã lưu JD cho chức danh {ten_cd}.")
                        st.rerun()
                    finally:
                        db.close()
                else:
                    st.warning("Vui lòng nhập Mã chức danh và Tên chức danh.")

    rows = _fetch_all("tp_job_description", order_by="phong_ban, ten_chuc_danh")
    if rows:
        for r in rows:
            with st.expander(f"📄 {r['ma_chuc_danh']} — {r['ten_chuc_danh']} ({r['phong_ban'] or ''})"):
                st.write(f"**Báo cáo cho:** {r['bao_cao_cho'] or '—'}")
                st.write(f"**Mục tiêu công việc:** {r['muc_tieu_cong_viec'] or '—'}")
                st.write(f"**Nhiệm vụ chính:**\n\n{r['nhiem_vu_chinh'] or '—'}")
                st.write(f"**Yêu cầu trình độ:** {r['yeu_cau_trinh_do'] or '—'}")
                st.write(f"**Yêu cầu kinh nghiệm:** {r['yeu_cau_kinh_nghiem'] or '—'}")
                st.write(f"**Yêu cầu kỹ năng:** {r['yeu_cau_ky_nang'] or '—'}")
                st.write(f"**Điều kiện làm việc:** {r['dieu_kien_lam_viec'] or '—'}")
                if st.button("🗑️ Xoá JD này", key=f"del_jd_{r['id']}"):
                    _delete("tp_job_description", r["id"])
                    st.rerun()
    else:
        st.info("Chưa có Bản mô tả công việc nào. Hãy thêm ít nhất các chức danh chủ chốt trước khi Đánh giá giá trị công việc.")


def _ui_job_evaluation():
    st.markdown("#### ⚖️ Bước 4 — Đánh giá giá trị công việc")
    st.caption("Chấm điểm 1–5 cho từng yếu tố theo trọng số. Hệ thống tự tính tổng điểm (thang 100) "
               "và gợi ý xếp Ngạch (Job Grade) tương ứng.")

    jds = _fetch_all("tp_job_description", order_by="ten_chuc_danh")
    if not jds:
        st.warning("Cần có ít nhất 1 Bản mô tả công việc (Bước 3) trước khi đánh giá giá trị công việc.")
        return

    with st.form("form_job_eval"):
        options = {f"{r['ma_chuc_danh']} — {r['ten_chuc_danh']}": r["ma_chuc_danh"] for r in jds}
        chon = st.selectbox("Chọn chức danh cần đánh giá", list(options.keys()))
        st.markdown("**Chấm điểm từng yếu tố (1 = thấp nhất, 5 = cao nhất):**")
        diem = {}
        for ten_yt, trong_so in YEU_TO_DANH_GIA_CONG_VIEC:
            diem[ten_yt] = st.slider(f"{ten_yt}  (trọng số {trong_so*100:.0f}%)", 1, 5, 3, key=f"je_{ten_yt}")
        nguoi_dg = st.text_input("Người đánh giá", value=st.session_state.get("ho_ten_dang_nhap", ""))
        if st.form_submit_button("📊 Tính điểm & Lưu kết quả đánh giá", type="primary"):
            tong_diem = sum(diem[ten] * trong_so for ten, trong_so in YEU_TO_DANH_GIA_CONG_VIEC) * 20
            ma_ngach, ten_ngach = goi_y_ngach(tong_diem)
            _insert("tp_job_evaluation", {
                "ma_chuc_danh": options[chon], "diem_chi_tiet": json.dumps(diem),
                "tong_diem": round(tong_diem, 2), "ma_ngach": ma_ngach, "ten_ngach": ten_ngach,
                "nguoi_danh_gia": nguoi_dg,
            })
            st.success(f"Tổng điểm: {tong_diem:.1f}/100 → Gợi ý xếp: **{ten_ngach}**")
            st.rerun()

    rows = _fetch_all("tp_job_evaluation", order_by="tong_diem DESC")
    if rows:
        import pandas as pd
        df = pd.DataFrame([{
            "Chức danh": r["ma_chuc_danh"], "Tổng điểm": float(r["tong_diem"]),
            "Ngạch gợi ý": r["ten_ngach"], "Người đánh giá": r["nguoi_danh_gia"],
        } for r in rows])
        st.dataframe(df, width='stretch', hide_index=True)


def _ui_p1():
    st.markdown("#### 💵 Bước 5 — Gợi ý Khung/Thang lương P1")
    st.caption("Hệ thống tự sinh khung Ngạch x Bậc theo phương pháp cấp số nhân — có thể chỉnh sửa "
               "thủ công từng dòng sau khi sinh, hoặc tự nhập tay toàn bộ.")

    with st.form("form_sinh_p1"):
        c1, c2, c3, c4 = st.columns(4)
        luong_bac1 = c1.number_input("Lương bậc 1 - Ngạch 1 (VNĐ)", min_value=1_000_000, value=6_500_000, step=100_000)
        so_bac = c2.number_input("Số bậc / ngạch", min_value=2, max_value=10, value=5)
        buoc_bac = c3.number_input("Bước tăng giữa các bậc (%)", min_value=1.0, max_value=20.0, value=6.0, step=0.5) / 100
        buoc_ngach = c4.number_input("Bước tăng giữa các ngạch (%)", min_value=5.0, max_value=50.0, value=18.0, step=1.0) / 100
        if st.form_submit_button("✨ Sinh khung lương P1 gợi ý", type="primary"):
            if _fetch_all("tp_p1_salary_scale"):
                st.warning("Đã có khung lương P1. Hãy xoá dữ liệu cũ trước khi sinh lại (để tránh ghi đè trùng ngạch/bậc).")
            else:
                rows = goi_y_khung_p1(luong_bac1, int(so_bac), buoc_bac, buoc_ngach)
                for r in rows:
                    _insert("tp_p1_salary_scale", r)
                st.success(f"Đã sinh {len(rows)} dòng khung lương P1 (7 ngạch x {so_bac} bậc).")
                st.rerun()

    rows = _fetch_all("tp_p1_salary_scale", order_by="ma_ngach, bac")
    if rows:
        import pandas as pd
        df = pd.DataFrame([{"Ngạch": r["ma_ngach"], "Tên ngạch": r["ten_ngach"], "Bậc": r["bac"],
                             "Hệ số": float(r["he_so"]), "Lương P1 (VNĐ)": int(r["luong_p1"])} for r in rows])
        edited = st.data_editor(df, width='stretch', hide_index=True, num_rows="fixed",
                                 disabled=["Ngạch", "Tên ngạch", "Bậc", "Hệ số"], key="p1_editor")
        if st.button("💾 Lưu chỉnh sửa lương P1", key="p1_save"):
            db, c = _conn(dict_cursor=False)
            try:
                for i, row in edited.iterrows():
                    c.execute("UPDATE tp_p1_salary_scale SET luong_p1=%s WHERE ma_ngach=%s AND bac=%s",
                              (int(row["Lương P1 (VNĐ)"]), row["Ngạch"], int(row["Bậc"])))
                db.commit()
                st.success("Đã cập nhật khung lương P1.")
            finally:
                db.close()
        if st.button("🗑️ Xoá toàn bộ khung lương P1 để sinh lại", key="p1_clear"):
            db, c = _conn(dict_cursor=False)
            try:
                c.execute("DELETE FROM tp_p1_salary_scale")
                db.commit()
            finally:
                db.close()
            st.rerun()
    else:
        st.info("Chưa có khung lương P1. Bấm 'Sinh khung lương P1 gợi ý' ở trên.")


def _ui_p2():
    st.markdown("#### 🧠 Bước 6 — Gợi ý Khung năng lực P2")
    st.caption("Danh mục năng lực cốt lõi / quản lý / chuyên môn theo khối công việc — mỗi năng lực "
               "đánh giá theo thang điểm 1–5 (mô tả cấp độ chuẩn hoá bên dưới).")

    with st.expander("ℹ️ Thang mô tả cấp độ năng lực dùng chung (1–5)"):
        for cap, mo_ta in P2_MO_TA_CAP_DO.items():
            st.write(f"**Cấp {cap}:** {mo_ta}")

    if st.button("✨ Nạp gợi ý Khung năng lực P2 cho CHL", key="p2_seed"):
        if _fetch_all("tp_p2_competency"):
            st.warning("Đã có dữ liệu — chỉ nạp gợi ý khi bảng đang trống.")
        else:
            for nhom, ten, mota in P2_NHOM_GOI_Y:
                _insert("tp_p2_competency", {"nhom_nang_luc": nhom, "ten_nang_luc": ten, "mo_ta": mota})
            st.success(f"Đã nạp {len(P2_NHOM_GOI_Y)} năng lực gợi ý.")
            st.rerun()

    with st.form("form_them_p2", clear_on_submit=True):
        c1, c2 = st.columns(2)
        nhom = c1.text_input("Nhóm năng lực", placeholder="VD: Cốt lõi (toàn công ty)")
        ten = c2.text_input("Tên năng lực")
        mota = st.text_area("Mô tả năng lực", height=60)
        ap_dung = st.text_input("Áp dụng cho chức danh/phòng ban (tuỳ chọn)")
        if st.form_submit_button("➕ Thêm năng lực", type="primary"):
            if ten:
                _insert("tp_p2_competency", {"nhom_nang_luc": nhom, "ten_nang_luc": ten,
                                              "mo_ta": mota, "ap_dung_chuc_danh": ap_dung})
                st.rerun()

    rows = _fetch_all("tp_p2_competency", order_by="nhom_nang_luc, id")
    if rows:
        nhom_hien_tai = None
        for r in rows:
            if r["nhom_nang_luc"] != nhom_hien_tai:
                st.markdown(f"**{r['nhom_nang_luc']}**")
                nhom_hien_tai = r["nhom_nang_luc"]
            cc = st.columns([3, 5, 1])
            cc[0].write(r["ten_nang_luc"])
            cc[1].write(r["mo_ta"] or "")
            if cc[2].button("🗑️", key=f"del_p2_{r['id']}"):
                _delete("tp_p2_competency", r["id"])
                st.rerun()
    else:
        st.info("Chưa có Khung năng lực P2.")


def _ui_p3():
    st.markdown("#### 🎯 Bước 7 — Gợi ý Hệ thống KPI P3")
    st.caption("KPI theo phòng ban, liên kết với khía cạnh BSC — dùng để tính hệ số hiệu quả (P3) hàng kỳ.")

    if st.button("✨ Nạp gợi ý Hệ thống KPI cho CHL", key="p3_seed"):
        if _fetch_all("tp_p3_kpi_system"):
            st.warning("Đã có dữ liệu — chỉ nạp gợi ý khi bảng đang trống.")
        else:
            for ma, ten, pb, kc, dvt, ts, tan_suat in P3_KPI_GOI_Y:
                _insert("tp_p3_kpi_system", {
                    "ma_kpi": ma, "ten_kpi": ten, "phong_ban": pb, "khia_canh_bsc": kc,
                    "don_vi_tinh": dvt, "trong_so": ts, "tan_suat": tan_suat,
                })
            st.success(f"Đã nạp {len(P3_KPI_GOI_Y)} KPI gợi ý.")
            st.rerun()

    with st.form("form_them_kpi", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        ma_kpi = c1.text_input("Mã KPI *")
        ten_kpi = c2.text_input("Tên KPI *")
        phong_ban = c3.selectbox("Phòng ban áp dụng", _lay_ds_phong_ban_thuc_te())
        c4, c5, c6 = st.columns(3)
        khia_canh = c4.selectbox("Khía cạnh BSC", BSC_KHIA_CANH)
        dvt = c5.text_input("Đơn vị tính")
        trong_so = c6.number_input("Trọng số (%)", min_value=1, max_value=100, value=20) / 100
        tan_suat = st.selectbox("Tần suất đo lường", ["Tháng", "Quý", "Năm"])
        if st.form_submit_button("➕ Thêm KPI", type="primary"):
            if ma_kpi and ten_kpi:
                _insert("tp_p3_kpi_system", {
                    "ma_kpi": ma_kpi.strip().upper(), "ten_kpi": ten_kpi, "phong_ban": phong_ban,
                    "khia_canh_bsc": khia_canh, "don_vi_tinh": dvt, "trong_so": trong_so, "tan_suat": tan_suat,
                })
                st.rerun()

    rows = _fetch_all("tp_p3_kpi_system", order_by="phong_ban, id")
    if rows:
        import pandas as pd
        df = pd.DataFrame([{"Mã KPI": r["ma_kpi"], "Tên KPI": r["ten_kpi"], "Phòng ban": r["phong_ban"],
                             "BSC": r["khia_canh_bsc"], "ĐVT": r["don_vi_tinh"],
                             "Trọng số": f"{float(r['trong_so'])*100:.0f}%", "Tần suất": r["tan_suat"]} for r in rows])
        st.dataframe(df, width='stretch', hide_index=True)
        tong_ts = {}
        for r in rows:
            tong_ts.setdefault(r["phong_ban"], 0)
            tong_ts[r["phong_ban"]] += float(r["trong_so"])
        for pb, ts in tong_ts.items():
            if abs(ts - 1.0) > 0.01:
                st.warning(f"⚠️ Tổng trọng số KPI của **{pb}** hiện là {ts*100:.0f}% (nên = 100%).")
    else:
        st.info("Chưa có Hệ thống KPI P3.")


def _ui_publish():
    st.markdown("#### 🚀 Bước 8 — Xuất bản chính sách lương 3P")
    hien_tai = chinh_sach_da_xuat_ban()
    thong_ke_hien_tai = {
        "Mục tiêu BSC": len(_fetch_all("tp_bsc_strategy_map")),
        "Dòng Ma trận chức năng": len(_fetch_all("tp_functional_matrix")),
        "Bản mô tả công việc": len(_fetch_all("tp_job_description")),
        "Đánh giá giá trị công việc": len(_fetch_all("tp_job_evaluation")),
        "Dòng Khung lương P1": len(_fetch_all("tp_p1_salary_scale")),
        "Năng lực P2": len(_fetch_all("tp_p2_competency")),
        "KPI P3": len(_fetch_all("tp_p3_kpi_system")),
    }
    cols = st.columns(len(thong_ke_hien_tai))
    for col, (k, v) in zip(cols, thong_ke_hien_tai.items()):
        col.metric(k, v)

    thieu = [k for k, v in thong_ke_hien_tai.items() if v == 0]
    if thieu:
        st.warning("⚠️ Còn thiếu dữ liệu ở: " + ", ".join(thieu) + ". Vẫn có thể xuất bản, nhưng nên hoàn thiện trước.")

    if hien_tai:
        st.success(f"✅ Chính sách 3P đang áp dụng: **Phiên bản {hien_tai['phien_ban']}**, "
                   f"xuất bản {hien_tai['ngay_xuat_ban'].strftime('%d/%m/%Y %H:%M')} bởi {hien_tai['nguoi_xuat_ban']}.")
    else:
        st.info("Chưa có phiên bản chính sách 3P nào được xuất bản. Chức năng 'Tính thu nhập' cho toàn công ty "
                "sẽ chờ đến khi có ít nhất 1 phiên bản được xuất bản.")

    if st.button("📢 Xuất bản chính sách lương 3P (chốt phiên bản mới)", type="primary"):
        nguoi = st.session_state.get("ho_ten_dang_nhap") or st.session_state.get("username") or "admin"
        phien_ban, tk = xuat_ban_chinh_sach(nguoi)
        st.success(f"🎉 Đã xuất bản chính sách lương 3P — Phiên bản {phien_ban}.")
        st.balloons()
        st.rerun()

    lich_su = _fetch_all("tp_policy_version", order_by="phien_ban DESC")
    if lich_su:
        st.markdown("**Lịch sử phiên bản:**")
        import pandas as pd
        df = pd.DataFrame([{"Phiên bản": r["phien_ban"], "Ngày xuất bản": r["ngay_xuat_ban"].strftime("%d/%m/%Y %H:%M"),
                             "Người xuất bản": r["nguoi_xuat_ban"], "Trạng thái": r["trang_thai"]} for r in lich_su])
        st.dataframe(df, width='stretch', hide_index=True)


def _ui_setup_3p():
    st.markdown("### 🧭 Quy trình thiết lập chính sách lương 3P")
    st.caption("Giúp Quý doanh nghiệp xây dựng Chính sách lương 3P từ con số 0: Thực hiện tuần tự 8 bước - mỗi bước có gợi ý dữ liệu mẫu riêng theo cơ cấu tổ chức của doanh nghiệp, admin có thể chỉnh sửa tự do trước khi Xuất bản."
               "Nếu Quý doanh nghiệp áp dụng chính sách tiền lương khác - vui lòng liên hệ và cung cấp nội dung chi tiết để được tùy chỉnh chức năng đúng với thực tế.")
    buoc = st.radio(
        "Chọn bước thiết lập:",
        ["1️⃣ Bản đồ chiến lược (BSC)", "2️⃣ Ma trận chức năng", "3️⃣ Bản mô tả công việc (JD)",
         "4️⃣ Đánh giá giá trị công việc", "5️⃣ Khung lương P1", "6️⃣ Khung năng lực P2",
         "7️⃣ Hệ thống KPI P3", "8️⃣ Xuất bản chính sách"],
        horizontal=False, key="buoc_setup_3p",
    )
    st.divider()
    if buoc.startswith("1️⃣"):
        _ui_bsc()
    elif buoc.startswith("2️⃣"):
        _ui_functional_matrix()
    elif buoc.startswith("3️⃣"):
        _ui_jd()
    elif buoc.startswith("4️⃣"):
        _ui_job_evaluation()
    elif buoc.startswith("5️⃣"):
        _ui_p1()
    elif buoc.startswith("6️⃣"):
        _ui_p2()
    elif buoc.startswith("7️⃣"):
        _ui_p3()
    elif buoc.startswith("8️⃣"):
        _ui_publish()


# ============================================================================
# 7) UI — CẤU HÌNH LƯƠNG & THUẾ
# ============================================================================

def _ui_payroll_config():
    st.markdown("### ⚙️ Cấu hình lương & thuế")
    st.caption("Các tỷ lệ mặc định tham khảo theo quy định hiện hành — vui lòng rà soát và cập nhật "
               "định kỳ theo văn bản pháp luật mới nhất hoặc chính sách riêng của công ty.")
    cfg = lay_cau_hinh_luong()

    with st.form("form_cau_hinh_luong"):
        st.markdown("**Bảo hiểm bắt buộc (phần Người lao động đóng)**")
        c1, c2, c3 = st.columns(3)
        bhxh = c1.number_input("Tỷ lệ BHXH (%)", 0.0, 20.0, cfg["bhxh_nld"] * 100, step=0.5) / 100
        bhyt = c2.number_input("Tỷ lệ BHYT (%)", 0.0, 10.0, cfg["bhyt_nld"] * 100, step=0.5) / 100
        bhtn = c3.number_input("Tỷ lệ BHTN (%)", 0.0, 5.0, cfg["bhtn_nld"] * 100, step=0.5) / 100

        c4, c5 = st.columns(2)
        ap_dung_doan_phi = c4.checkbox("Áp dụng khấu trừ Đoàn phí công đoàn", value=cfg.get("ap_dung_doan_phi", False))
        doan_phi = c5.number_input("Tỷ lệ đoàn phí (%)", 0.0, 5.0, cfg["doan_phi_nld"] * 100, step=0.5) / 100

        st.markdown("**Trần đóng bảo hiểm**")
        c6, c7 = st.columns(2)
        luong_co_so = c6.number_input("Lương cơ sở (VNĐ) — trần BHXH/BHYT = 20 lần", min_value=0,
                                       value=int(cfg["luong_co_so"]), step=10_000)
        luong_vung = c7.number_input("Lương tối thiểu vùng (VNĐ) — trần BHTN = 20 lần", min_value=0,
                                      value=int(cfg["luong_toi_thieu_vung"]), step=10_000)

        st.markdown("**Giảm trừ gia cảnh (Thuế TNCN)**")
        c8, c9 = st.columns(2)
        gt_ban_than = c8.number_input("Giảm trừ bản thân (VNĐ/tháng)", min_value=0,
                                       value=int(cfg["giam_tru_ban_than"]), step=100_000)
        gt_phu_thuoc = c9.number_input("Giảm trừ mỗi người phụ thuộc (VNĐ/tháng)", min_value=0,
                                        value=int(cfg["giam_tru_nguoi_phu_thuoc"]), step=100_000)

        st.markdown("**Quỹ P2 / P3 và điều chỉnh theo hiệu quả**")
        c10, c11 = st.columns(2)
        ty_le_p2 = c10.number_input("Quỹ P2 mục tiêu (% của P1)", 0.0, 100.0, cfg["ty_le_quy_p2"] * 100, step=1.0) / 100
        ty_le_p3 = c11.number_input("Quỹ P3 mục tiêu (% của P1)", 0.0, 100.0, cfg["ty_le_quy_p3"] * 100, step=1.0) / 100
        c12, c13, c14 = st.columns(3)
        chuan_diem = c12.number_input("Điểm năng lực chuẩn (thang 1-5)", 1.0, 5.0, cfg["chuan_diem_nang_luc"], step=0.1)
        he_so_p2_min = c13.number_input("Hệ số P2 tối thiểu", 0.0, 2.0, cfg["he_so_p2_min"], step=0.05)
        he_so_p2_max = c14.number_input("Hệ số P2 tối đa", 0.0, 2.0, cfg["he_so_p2_max"], step=0.05)

        st.markdown("**Phụ cấp thâm niên công ty (ngoài quy định nhà nước)**")
        c15, c16 = st.columns(2)
        pc_tn_nam = c15.number_input("Mức phụ cấp / năm công tác (VNĐ)", min_value=0,
                                      value=int(cfg["phu_cap_tham_nien_nam"]), step=10_000)
        pc_tn_max = c16.number_input("Số năm tối đa tính phụ cấp thâm niên", min_value=0, max_value=40,
                                      value=int(cfg["phu_cap_tham_nien_toi_da_nam"]))

        if st.form_submit_button("💾 Lưu cấu hình", type="primary"):
            cfg_moi = dict(cfg)
            cfg_moi.update({
                "bhxh_nld": bhxh, "bhyt_nld": bhyt, "bhtn_nld": bhtn,
                "ap_dung_doan_phi": ap_dung_doan_phi, "doan_phi_nld": doan_phi,
                "luong_co_so": luong_co_so, "luong_toi_thieu_vung": luong_vung,
                "giam_tru_ban_than": gt_ban_than, "giam_tru_nguoi_phu_thuoc": gt_phu_thuoc,
                "ty_le_quy_p2": ty_le_p2, "ty_le_quy_p3": ty_le_p3,
                "chuan_diem_nang_luc": chuan_diem, "he_so_p2_min": he_so_p2_min, "he_so_p2_max": he_so_p2_max,
                "phu_cap_tham_nien_nam": pc_tn_nam, "phu_cap_tham_nien_toi_da_nam": pc_tn_max,
            })
            luu_cau_hinh_luong(cfg_moi)
            st.success("Đã lưu cấu hình lương & thuế.")
            st.rerun()

    with st.expander("📊 Biểu thuế TNCN luỹ tiến từng phần đang áp dụng"):
        import pandas as pd
        bac_thue = cfg.get("bac_thue_tncn", BAC_THUE_TNCN_MAC_DINH)
        df = pd.DataFrame([{
            "Từ (VNĐ)": f"{b['tu']:,.0f}",
            "Đến (VNĐ)": f"{b['den']:,.0f}" if b["den"] else "Trở lên",
            "Thuế suất": f"{b['thue_suat']*100:.0f}%",
        } for b in bac_thue])
        st.dataframe(df, width='stretch', hide_index=True)
        st.caption("Biểu thuế theo Điều 22 Luật Thuế TNCN 2007 (sửa đổi 2012). Để chỉnh sửa biểu thuế, "
                   "sửa trực tiếp hằng số BAC_THUE_TNCN_MAC_DINH trong mã nguồn hoặc mở rộng UI này.")


# ============================================================================
# 8) UI — GÁN NHÂN VIÊN VÀO KHUNG 3P (P1 / P2 / P3 hàng kỳ)
# ============================================================================

def _ui_gan_nhan_vien():
    st.markdown("### 👥 Gán nhân viên vào khung 3P")
    tab1, tab2, tab3, tab4 = st.tabs(["📍 Gán Ngạch/Bậc (P1)", "🧠 Nhập điểm năng lực (P2)",
                                        "🎯 Nhập tỷ lệ hoàn thành KPI (P3)", "👪 Người phụ thuộc"])

    p1_rows = _fetch_all("tp_p1_salary_scale", order_by="ma_ngach, bac")
    nv_list = _lay_ds_nhan_vien()

    with tab1:
        if not p1_rows:
            st.warning("Chưa có Khung lương P1 — hãy thiết lập ở Bước 5 trước.")
        else:
            nv_options = {f"{nv['ma_nv']} - {nv['ho_ten']} ({nv['phong_ban_lam_viec'] or ''})": nv["id"] for nv in nv_list}
            with st.form("form_gan_p1"):
                chon_nv = st.selectbox("Chọn nhân viên", list(nv_options.keys()), key="gan_p1_nv")
                ma_chuc_danh = st.text_input("Mã chức danh (tuỳ chọn, tham chiếu JD)")
                ngach_options = sorted(set(r["ma_ngach"] for r in p1_rows))
                ma_ngach = st.selectbox("Ngạch", ngach_options)
                bac_options = sorted(set(r["bac"] for r in p1_rows if r["ma_ngach"] == ma_ngach))
                bac = st.selectbox("Bậc", bac_options)
                if st.form_submit_button("💾 Gán / Cập nhật", type="primary"):
                    nv_id = nv_options[chon_nv]
                    db, c = _conn(dict_cursor=False)
                    try:
                        c.execute("""
                            INSERT INTO tp_employee_p1_assignment (nhan_vien_id, ma_chuc_danh, ma_ngach, bac)
                            VALUES (%s,%s,%s,%s)
                            ON CONFLICT (nhan_vien_id) DO UPDATE SET
                                ma_chuc_danh=EXCLUDED.ma_chuc_danh, ma_ngach=EXCLUDED.ma_ngach,
                                bac=EXCLUDED.bac, updated_at=NOW()
                        """, (nv_id, ma_chuc_danh, ma_ngach, bac))
                        db.commit()
                        st.success("Đã gán vị trí lương P1 cho nhân viên.")
                        st.rerun()
                    finally:
                        db.close()

            gan_rows = _fetch_all("tp_employee_p1_assignment")
            if gan_rows:
                nv_map = {nv["id"]: nv for nv in nv_list}
                import pandas as pd
                df = pd.DataFrame([{
                    "Mã NV": nv_map.get(r["nhan_vien_id"], {}).get("ma_nv", r["nhan_vien_id"]),
                    "Họ tên": nv_map.get(r["nhan_vien_id"], {}).get("ho_ten", ""),
                    "Ngạch": r["ma_ngach"], "Bậc": r["bac"],
                    "Lương P1": f"{lay_p1_luong(r['ma_ngach'], r['bac']):,.0f}",
                } for r in gan_rows])
                st.dataframe(df, width='stretch', hide_index=True)

    with tab2:
        c1, c2 = st.columns(2)
        thang = c1.number_input("Tháng", 1, 12, datetime.now().month, key="p2_thang")
        nam = c2.number_input("Năm", 2020, 2100, datetime.now().year, key="p2_nam")
        nv_options2 = {f"{nv['ma_nv']} - {nv['ho_ten']}": nv["id"] for nv in nv_list}
        with st.form("form_diem_p2"):
            chon = st.selectbox("Chọn nhân viên", list(nv_options2.keys()), key="p2_nv")
            diem = st.slider("Điểm năng lực trung bình kỳ này (thang 1-5)", 1.0, 5.0, 3.0, step=0.1)
            ghi_chu = st.text_input("Ghi chú đánh giá")
            if st.form_submit_button("💾 Lưu điểm năng lực", type="primary"):
                nv_id = nv_options2[chon]
                db, c = _conn(dict_cursor=False)
                try:
                    c.execute("""
                        INSERT INTO tp_employee_p2_score (nhan_vien_id, thang, nam, diem_nang_luc, ghi_chu)
                        VALUES (%s,%s,%s,%s,%s)
                        ON CONFLICT (nhan_vien_id, thang, nam) DO UPDATE SET
                            diem_nang_luc=EXCLUDED.diem_nang_luc, ghi_chu=EXCLUDED.ghi_chu, updated_at=NOW()
                    """, (nv_id, thang, nam, diem, ghi_chu))
                    db.commit()
                    st.success("Đã lưu điểm năng lực P2.")
                    st.rerun()
                finally:
                    db.close()

    with tab3:
        c1, c2 = st.columns(2)
        thang3 = c1.number_input("Tháng", 1, 12, datetime.now().month, key="p3_thang")
        nam3 = c2.number_input("Năm", 2020, 2100, datetime.now().year, key="p3_nam")
        nv_options3 = {f"{nv['ma_nv']} - {nv['ho_ten']}": nv["id"] for nv in nv_list}
        with st.form("form_ty_le_p3"):
            chon3 = st.selectbox("Chọn nhân viên", list(nv_options3.keys()), key="p3_nv")
            ty_le = st.slider("Tỷ lệ hoàn thành KPI kỳ này (%)", 0, 150, 100)
            ghi_chu3 = st.text_input("Ghi chú KPI", key="p3_ghichu")
            if st.form_submit_button("💾 Lưu tỷ lệ hoàn thành KPI", type="primary"):
                nv_id = nv_options3[chon3]
                db, c = _conn(dict_cursor=False)
                try:
                    c.execute("""
                        INSERT INTO tp_employee_p3_score (nhan_vien_id, thang, nam, ty_le_hoan_thanh, ghi_chu)
                        VALUES (%s,%s,%s,%s,%s)
                        ON CONFLICT (nhan_vien_id, thang, nam) DO UPDATE SET
                            ty_le_hoan_thanh=EXCLUDED.ty_le_hoan_thanh, ghi_chu=EXCLUDED.ghi_chu, updated_at=NOW()
                    """, (nv_id, thang3, nam3, ty_le, ghi_chu3))
                    db.commit()
                    st.success("Đã lưu tỷ lệ hoàn thành KPI.")
                    st.rerun()
                finally:
                    db.close()

    with tab4:
        nv_options4 = {f"{nv['ma_nv']} - {nv['ho_ten']}": nv["id"] for nv in nv_list}
        with st.form("form_phu_thuoc", clear_on_submit=True):
            chon4 = st.selectbox("Chọn nhân viên", list(nv_options4.keys()), key="pt_nv")
            c1, c2 = st.columns(2)
            ho_ten_pt = c1.text_input("Họ tên người phụ thuộc")
            quan_he = c2.selectbox("Quan hệ", ["Con", "Vợ/Chồng", "Cha/Mẹ", "Khác"])
            ngay_sinh_pt = st.date_input("Ngày sinh", value=None, format="DD/MM/YYYY")
            if st.form_submit_button("➕ Thêm người phụ thuộc", type="primary"):
                if ho_ten_pt:
                    nv_id = nv_options4[chon4]
                    _insert("tp_employee_dependents", {
                        "nhan_vien_id": nv_id, "ho_ten": ho_ten_pt, "quan_he": quan_he,
                        "ngay_sinh": ngay_sinh_pt,
                    })
                    st.rerun()

        rows_pt = _fetch_all("tp_employee_dependents")
        if rows_pt:
            nv_map = {nv["id"]: nv for nv in nv_list}
            import pandas as pd
            df = pd.DataFrame([{
                "Mã NV": nv_map.get(r["nhan_vien_id"], {}).get("ma_nv", r["nhan_vien_id"]),
                "Người phụ thuộc": r["ho_ten"], "Quan hệ": r["quan_he"],
                "Ngày sinh": r["ngay_sinh"].strftime("%d/%m/%Y") if r["ngay_sinh"] else "",
            } for r in rows_pt])
            st.dataframe(df, width='stretch', hide_index=True)


# ============================================================================
# 9) UI — TÍNH THU NHẬP THÁNG & BẢNG LƯƠNG
# ============================================================================

def _ui_tinh_luong_thang():
    st.markdown("### 💰 Tính thu nhập & Bảng lương tháng")

    chinh_sach = chinh_sach_da_xuat_ban()
    if not chinh_sach:
        st.error("⛔ Chưa có chính sách lương 3P nào được **Xuất bản**. Vui lòng liên hệ Admin để hoàn thiện "
                  "và xuất bản chính sách 3P trước khi tính thu nhập.")
        return
    st.caption(f"Đang áp dụng chính sách 3P — Phiên bản {chinh_sach['phien_ban']} "
               f"(xuất bản {chinh_sach['ngay_xuat_ban'].strftime('%d/%m/%Y')}).")

    cfg = lay_cau_hinh_luong()
    ten_cty = ""
    tenant = st.session_state.get("tenant")
    if tenant:
        ten_cty = tenant.get("ten_cty", "")

    c1, c2, c3 = st.columns(3)
    thang = c1.selectbox("Tháng", list(range(1, 13)), index=datetime.now().month - 1)
    nam = c2.number_input("Năm", 2020, 2100, datetime.now().year)
    pham_vi = c3.radio("Phạm vi tính", ["Toàn công ty", "Theo phòng ban", "1 nhân viên"], horizontal=True)

    nv_list = _lay_ds_nhan_vien()
    if pham_vi == "Theo phòng ban":
        ds_pb = _lay_ds_phong_ban_thuc_te()
        pb_chon = st.selectbox("Chọn phòng ban", ds_pb)
        nv_list = [nv for nv in nv_list if nv["phong_ban_lam_viec"] == pb_chon]
    elif pham_vi == "1 nhân viên":
        nv_options = {f"{nv['ma_nv']} - {nv['ho_ten']}": nv for nv in nv_list}
        chon = st.selectbox("Chọn nhân viên", list(nv_options.keys()))
        nv_list = [nv_options[chon]]

    if not nv_list:
        st.warning("Không có nhân viên phù hợp trong phạm vi đã chọn.")
        return

    salary_module = _load_salary_module()
    if salary_module and hasattr(salary_module, 'tinh_luong'):
        _mst_hien_tai_luong = ((st.session_state.get('tenant') or {}).get('ma_so_thue') or '').strip()
        if _mst_hien_tai_luong and salary_module.__name__ == f"salary.salary_{_mst_hien_tai_luong}":
            st.caption(f"💡 Đang dùng công thức lương RIÊNG của công ty bạn: `salary/salary_{_mst_hien_tai_luong}.py`")
        else:
            st.caption("💡 Đang dùng công thức lương mặc định: `salary/salary_demo.py`")

    if st.button(f"🧮 Tính lương 3P — Tháng {thang}/{nam} ({len(nv_list)} nhân viên)", type="primary"):
        ket_qua = []
        chua_gan_p1 = []
        for nv in nv_list:
            if not lay_gan_p1(nv["id"]):
                chua_gan_p1.append(nv["ho_ten"])
            if salary_module and hasattr(salary_module, 'tinh_luong'):
                bd = salary_module.tinh_luong(dict(nv), thang, nam, cfg)
            else:
                # An toàn dự phòng: nếu vì lý do gì đó không nạp được cả file riêng lẫn
                # salary_demo.py (VD thiếu thư mục salary/ trên môi trường triển khai),
                # vẫn tính lương được bằng công thức 3P gốc thay vì crash cả màn hình.
                bd = tinh_luong_nhan_vien(dict(nv), thang, nam, cfg)
            ket_qua.append(bd)
        st.session_state["_ketqua_luong_3p"] = ket_qua
        st.session_state["_ketqua_luong_ky"] = (thang, nam)
        if chua_gan_p1:
            st.warning("⚠️ Các nhân viên sau CHƯA được gán Ngạch/Bậc P1 (lương P1 = 0), vào tab "
                       "'Gán nhân viên vào khung 3P' để gán: " + ", ".join(chua_gan_p1))
        st.success(f"Đã tính lương cho {len(ket_qua)} nhân viên.")

    ket_qua = st.session_state.get("_ketqua_luong_3p")
    ky_luu = st.session_state.get("_ketqua_luong_ky")
    if not ket_qua or ky_luu != (thang, nam):
        return

    import pandas as pd
    df = pd.DataFrame([{
        "Mã NV": r["ma_nv"], "Họ tên": r["ho_ten"], "Phòng ban": r["phong_ban"],
        "P1": r["p1_luong_vi_tri"], "P2": r["p2_luong_nang_luc"], "P3": r["p3_luong_hieu_qua"],
        "Phụ cấp": r["phu_cap_chuc_vu"] + r["phu_cap_tham_nien"] + r["phu_cap_trach_nhiem"] + r["phu_cap_khac"],
        "Tổng thu nhập": r["tong_thu_nhap"], "Khấu trừ BH": r["tong_khau_tru_bh"],
        "Thuế TNCN": r["thue_tncn"], "Thực nhận": r["thuc_nhan"],
    } for r in ket_qua])
    st.dataframe(df, width='stretch', hide_index=True)

    tong_qtl = sum(r["tong_thu_nhap"] for r in ket_qua)
    tong_tn = sum(r["thuc_nhan"] for r in ket_qua)
    m1, m2, m3 = st.columns(3)
    m1.metric("Tổng quỹ thu nhập", f"{tong_qtl:,.0f} đ")
    m2.metric("Tổng khấu trừ (BH + Thuế)", f"{(tong_qtl - tong_tn):,.0f} đ")
    m3.metric("Tổng thực nhận", f"{tong_tn:,.0f} đ")

    st.divider()
    st.markdown("#### 💾 Lưu / Xuất / Gửi bảng lương")
    cA, cB, cC, cD = st.columns(4)

    if cA.button("💾 Lưu bảng lương vào hệ thống"):
        period_id = lay_hoac_tao_ky_luong(thang, nam)
        nguoi_tinh = st.session_state.get("ho_ten_dang_nhap") or st.session_state.get("username") or ""
        for r in ket_qua:
            luu_bang_luong(period_id, r, nguoi_tinh)
        st.success(f"Đã lưu {len(ket_qua)} dòng bảng lương tháng {thang}/{nam}.")

    excel_bytes = export_bang_luong_excel(ket_qua, thang, nam, ten_cty)
    cB.download_button("📊 Tải Excel bảng lương", data=excel_bytes,
                        file_name=f"BangLuong_3P_T{thang}_{nam}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with cC.popover("📄 Tải phiếu lương PDF (từng NV)"):
        nv_pdf_options = {f"{r['ma_nv']} - {r['ho_ten']}": r for r in ket_qua}
        chon_pdf = st.selectbox("Chọn nhân viên", list(nv_pdf_options.keys()), key="pdf_nv_chon")
        if st.button("Tạo PDF", key="pdf_tao"):
            try:
                pdf_bytes = export_phieu_luong_pdf(nv_pdf_options[chon_pdf], ten_cty)
                st.download_button("⬇️ Tải xuống phiếu lương PDF", data=pdf_bytes,
                                    file_name=f"PhieuLuong_{nv_pdf_options[chon_pdf]['ma_nv']}_T{thang}_{nam}.pdf",
                                    mime="application/pdf", key="pdf_dl")
            except ModuleNotFoundError:
                st.error("Cần cài thư viện `reportlab` để xuất PDF: pip install reportlab")

    with cD.popover("💬 Gửi phiếu lương qua Chat nội bộ"):
        gui_tat_ca = st.checkbox("Gửi cho tất cả nhân viên trong bảng", value=False)
        nv_chat_options = {f"{r['ma_nv']} - {r['ho_ten']}": r for r in ket_qua}
        chon_chat = None
        if not gui_tat_ca:
            chon_chat = st.selectbox("Chọn nhân viên", list(nv_chat_options.keys()), key="chat_nv_chon")
        if st.button("Gửi", key="chat_gui"):
            sender_id = st.session_state.get("nhan_vien_id")
            targets = list(nv_chat_options.values()) if gui_tat_ca else [nv_chat_options[chon_chat]]
            thanh_cong, that_bai = 0, 0
            for r in targets:
                ok, msg = gui_phieu_luong_qua_chat(r["nhan_vien_id"], r, sender_id)
                thanh_cong += 1 if ok else 0
                that_bai += 0 if ok else 1
            if that_bai == 0:
                st.success(f"Đã gửi phiếu lương qua chat nội bộ cho {thanh_cong} nhân viên.")
            else:
                st.warning(f"Gửi thành công {thanh_cong}, thất bại {that_bai}. "
                           "(Thất bại thường do nhân viên chưa có tài khoản chat / chưa liên kết id.)")


# ============================================================================
# 10) ENTRY POINT — gọi từ app.py
# ============================================================================

def show_tinh_thu_nhap():
    ensure_3p_tables()
    st.markdown("# 💰 Tính thu nhập (Chính sách lương 3P)")
    st.caption("Pay for Position (P1) • Pay for Person (P2) • Pay for Performance (P3)")

    role = st.session_state.get("role")
    if role == "admin":
        tab1, tab2, tab3, tab4 = st.tabs([
            "🧭 Thiết lập chính sách 3P", "⚙️ Cấu hình lương & thuế",
            "👥 Gán nhân viên vào khung 3P", "💰 Tính thu nhập & Bảng lương",
        ])
        with tab1:
            _ui_setup_3p()
        with tab2:
            _ui_payroll_config()
        with tab3:
            _ui_gan_nhan_vien()
        with tab4:
            _ui_tinh_luong_thang()
    else:
        _ui_tinh_luong_thang()
