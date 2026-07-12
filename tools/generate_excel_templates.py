"""
generate_excel_templates.py
============================
Sinh các file Excel MẪU CHUẨN để khách hàng mới (onboarding) điền dữ liệu của họ,
sau đó dùng import_from_excel.py để nạp thẳng vào Database.

CÁCH DÙNG:
    python generate_excel_templates.py                # sinh tất cả các bảng trong SCHEMA_CONFIG
    python generate_excel_templates.py nhan_vien       # chỉ sinh 1 bảng cụ thể

QUY TẮC THIẾT KẾ MẪU (áp dụng cho MỌI bảng, không ngoại lệ):
    - Dòng 1 (Header): tên cột KHÔNG ĐƯỢC ĐỔI (import script khớp theo đúng tên cột này).
      Cột bắt buộc có dữ liệu -> chữ ĐỎ, đậm, có dấu *.
      Cột không bắt buộc -> chữ đen, đậm.
    - Dòng 2: mô tả ngắn / định dạng mong muốn cho từng cột (chữ nghiêng, xám).
    - Dòng 3: 1 dòng VÍ DỤ mẫu (dữ liệu thật hợp lệ, khách xoá đi trước khi điền dữ liệu thật).
    - Từ dòng 4 trở đi: khách hàng điền dữ liệu.
    - Cấu trúc cột bị KHOÁ (sheet protection): khách không thể thêm/xoá/đổi tên cột,
      không sửa được Header/Mô tả - chỉ nhập được vào vùng dữ liệu.
    - Cột nào có danh sách giá trị cố định (VD: Giới tính, Trạng thái...) -> có Data Validation
      dropdown để tránh sai chính tả khi nhập liệu.

THÊM BẢNG MỚI: chỉ cần thêm 1 entry vào SCHEMA_CONFIG bên dưới theo đúng format có sẵn,
KHÔNG cần sửa phần code phía dưới.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "excel_templates")

# ==============================================================================
# SCHEMA_CONFIG — định nghĩa cấu trúc chuẩn của từng bảng.
# Mỗi cột: (ten_cot_trong_db, nhan_hien_thi, bat_buoc(bool), mo_ta_dinh_dang, gia_tri_vi_du, danh_sach_lua_chon_hoac_None)
# ==============================================================================
SCHEMA_CONFIG = {

    "nhan_vien": {
        "sheet_title": "Nhân viên",
        "note": "Mỗi dòng = 1 nhân viên. Mã NV sẽ được hệ thống tự sinh khi import — KHÔNG cần điền.",
        "columns": [
            ("ho_ten", "Họ và tên", True, "Chữ, viết hoa đầu từ", "Nguyễn Văn A", None),
            ("gioi_tinh", "Giới tính", True, "Chọn trong danh sách", "Nam", ["Nam", "Nữ"]),
            ("ngay_sinh", "Ngày sinh", True, "dd/mm/yyyy", "15/03/1990", None),
            ("so_cccd", "Số CCCD/CMND", True, "Số, đủ 9 hoặc 12 số", "038090001234", None),
            ("ngay_cap_cccd", "Ngày cấp CCCD", False, "dd/mm/yyyy", "10/05/2021", None),
            ("noi_cap_cccd", "Nơi cấp CCCD", False, "Chữ", "Cục CS QLHC về TTXH", None),
            ("nguyen_quan", "Nguyên quán", False, "Chữ", "Quảng Bình", None),
            ("thuong_tru", "Địa chỉ thường trú", False, "Chữ", "123 Đường ABC, TP Đồng Hới", None),
            ("dien_thoai", "Số điện thoại", True, "10 số, bắt đầu bằng 0", "0912345678", None),
            ("email", "Email cá nhân", False, "email@domain.com", "nguyenvana@gmail.com", None),
            ("email_lien_he", "Email liên hệ (nhận OTP)", False, "email@domain.com — dùng để reset mật khẩu", "nguyenvana@gmail.com", None),
            ("chuc_danh_nghe", "Chức danh nghề", False, "Theo danh mục Chức danh đã cấu hình", "Nhân viên kế toán", None),
            ("chuc_vu", "Chức vụ", False, "Theo danh mục Chức vụ đã cấu hình (để trống nếu là NV thường)", "Trưởng phòng", None),
            ("phong_ban_lam_viec", "Phòng ban", True, "Theo danh mục Phòng ban đã cấu hình", "Phòng Kế toán", None),
            ("ngay_vao_lam", "Ngày vào làm", True, "dd/mm/yyyy", "01/06/2020", None),
            ("loai_hop_dong", "Loại hợp đồng", True, "Theo danh mục Loại HĐ đã cấu hình", "Không xác định thời hạn", None),
            ("so_hdld", "Số hợp đồng lao động", False, "Chữ/số", "05/2020/HĐLĐ-CHL", None),
            ("ngay_ky_hd", "Ngày ký hợp đồng", False, "dd/mm/yyyy", "01/06/2020", None),
            ("ngay_ket_thuc", "Ngày kết thúc HĐ (nếu có thời hạn)", False, "dd/mm/yyyy — để trống nếu không xác định thời hạn", "", None),
            ("trang_thai", "Trạng thái", True, "Chọn trong danh sách", "DANG_LAM", ["DANG_LAM", "THU_VIEC", "NGHI_VIEC"]),
            ("trinh_do", "Trình độ học vấn", False, "Theo danh mục Trình độ đã cấu hình", "Đại học", None),
            ("quoc_tich", "Quốc tịch", False, "Chữ", "Việt Nam", None),
            ("dan_toc", "Dân tộc", False, "Chữ", "Kinh", None),
            ("he_so_luong", "Hệ số lương", False, "Số thập phân", "2.34", None),
            ("phu_cap_chuc_vu", "Phụ cấp chức vụ (VNĐ)", False, "Số nguyên", "500000", None),
            ("phu_cap_tnvk", "Phụ cấp trách nhiệm việc khó (VNĐ)", False, "Số nguyên", "0", None),
            ("phu_cap_tnn", "Phụ cấp thâm niên nghề (VNĐ)", False, "Số nguyên", "0", None),
            ("luong_bao_hiem", "Lương đóng bảo hiểm (VNĐ)", False, "Số nguyên", "6500000", None),
            ("ma_so_bhxh", "Mã số BHXH", False, "Số", "0123456789", None),
            ("nhom_bhxh", "Nhóm BHXH", False, "Chọn trong danh sách", "Văn phòng", ["Văn phòng", "Lao động trực tiếp"]),
            ("noi_lam_viec", "Nơi làm việc", False, "Chữ", "Cảng Hòn La", None),
            ("so_tai_khoan_nh", "Số tài khoản ngân hàng", False, "Số", "0071001234567", None),
            ("chi_nhanh_nh", "Ngân hàng / Chi nhánh", False, "Chữ", "Vietcombank Quảng Bình", None),
        ],
    },

    "ung_vien": {
        "sheet_title": "Ứng viên",
        "note": "Mỗi dòng = 1 hồ sơ ứng viên đang trong quá trình tuyển dụng. Mã UV tự sinh khi import.",
        "columns": [
            ("ho_ten", "Họ và tên", True, "Chữ", "Trần Thị B", None),
            ("vi_tri_du_tuyen", "Vị trí dự tuyển", True, "Theo danh mục Vị trí dự tuyển (Ứng viên)", "Nhân viên kế toán", None),
            ("dien_thoai", "Số điện thoại", True, "10 số", "0987654321", None),
            ("ngay_sinh", "Ngày sinh", False, "dd/mm/yyyy", "20/07/1998", None),
            ("gioi_tinh", "Giới tính", False, "Chọn trong danh sách", "Nữ", ["Nam", "Nữ"]),
            ("ngay_vao_lam", "Ngày dự kiến vào làm", False, "dd/mm/yyyy", "01/08/2026", None),
            ("luong_bao_hiem", "Mức lương thoả thuận (VNĐ)", False, "Số nguyên", "7000000", None),
        ],
    },

    "danh_muc_phong_ban": {
        "sheet_title": "Phòng ban",
        "note": "Danh mục Phòng ban của doanh nghiệp.",
        "columns": [
            ("ten_phong_ban", "Tên phòng ban", True, "Chữ, không trùng lặp", "Phòng Kế toán", None),
        ],
    },

    "danh_muc_loai_hop_dong": {
        "sheet_title": "Loại hợp đồng",
        "note": "Danh mục các loại hợp đồng lao động doanh nghiệp đang áp dụng.",
        "columns": [
            ("ten_loai_hd", "Tên loại hợp đồng", True, "Chữ, không trùng lặp", "Hợp đồng thời vụ", None),
        ],
    },

    "danh_muc_trinh_do_hoc_van": {
        "sheet_title": "Trình độ học vấn",
        "note": "Danh mục trình độ học vấn/bằng cấp.",
        "columns": [
            ("ten_trinh_do", "Tên trình độ", True, "Chữ, không trùng lặp", "Đại học", None),
        ],
    },

    "chuc_vu_danh_muc": {
        "sheet_title": "Chức vụ",
        "note": "Danh mục Chức vụ (Trưởng phòng, Tổ trưởng, Phó phòng...).",
        "columns": [
            ("ten_chuc_vu", "Tên chức vụ", True, "Chữ, không trùng lặp", "Trưởng phòng", None),
        ],
    },

    "vi_tri_cong_tac": {
        "sheet_title": "Chức danh Nhân viên",
        "note": "Danh mục Chức danh nghề áp dụng cho Nhân viên (khác với Vị trí dự tuyển của Ứng viên).",
        "columns": [
            ("ten_vi_tri", "Tên chức danh", True, "Chữ, không trùng lặp", "Nhân viên kế toán", None),
        ],
    },

    "chuc_danh_ung_vien": {
        "sheet_title": "Vị trí dự tuyển (Ứng viên)",
        "note": "Danh mục Vị trí dự tuyển áp dụng riêng cho Ứng viên (độc lập với Chức danh Nhân viên).",
        "columns": [
            ("ten_chuc_danh", "Tên vị trí dự tuyển", True, "Chữ, không trùng lặp", "Nhân viên kinh doanh", None),
        ],
    },
}

# ==============================================================================
# Styling constants
# ==============================================================================
FONT_NAME = "Arial"
RED_REQUIRED = Font(name=FONT_NAME, bold=True, color="C00000", size=11)
BLACK_OPTIONAL = Font(name=FONT_NAME, bold=True, color="000000", size=11)
DESC_FONT = Font(name=FONT_NAME, italic=True, color="808080", size=9)
EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, color="1F4E78", size=10)
HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def _build_sheet(ws, table_key, cfg):
    cols = cfg["columns"]
    n_cols = len(cols)

    # ----- Dòng ghi chú tổng quan (dòng 0, ẩn phía trên header bằng cách chèn ở row 1 rồi đẩy header xuống) -----
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    note_cell = ws.cell(row=1, column=1, value=f"📋 {cfg['sheet_title']} — {cfg['note']}  |  "
                                                  f"Cột màu ĐỎ có dấu (*) = BẮT BUỘC phải nhập.")
    note_cell.font = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    note_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    HEADER_ROW = 2
    DESC_ROW = 3
    EXAMPLE_ROW = 4
    DATA_START_ROW = 5

    for idx, (col_key, label, required, desc, example, choices) in enumerate(cols, start=1):
        col_letter = get_column_letter(idx)

        # Header
        header_text = f"{label} (*)" if required else label
        hc = ws.cell(row=HEADER_ROW, column=idx, value=header_text)
        hc.font = RED_REQUIRED if required else BLACK_OPTIONAL
        hc.fill = HEADER_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = THIN_BORDER

        # Mô tả định dạng
        dc = ws.cell(row=DESC_ROW, column=idx, value=desc)
        dc.font = DESC_FONT
        dc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dc.border = THIN_BORDER

        # Dòng ví dụ
        ec = ws.cell(row=EXAMPLE_ROW, column=idx, value=example)
        ec.font = EXAMPLE_FONT
        ec.fill = EXAMPLE_FILL
        ec.border = THIN_BORDER

        # Data validation dropdown nếu có danh sách lựa chọn cố định
        if choices:
            dv = DataValidation(type="list", formula1=f'"{",".join(choices)}"', allow_blank=not required)
            dv.error = f"Vui lòng chọn 1 giá trị trong danh sách: {', '.join(choices)}"
            dv.errorTitle = "Giá trị không hợp lệ"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{DATA_START_ROW}:{col_letter}{DATA_START_ROW + 500}")

        ws.column_dimensions[col_letter].width = max(16, len(label) + 4)

    ws.row_dimensions[HEADER_ROW].height = 34
    ws.row_dimensions[DESC_ROW].height = 26
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # ----- Khoá cấu trúc: header/mô tả/ví dụ bị khoá, vùng dữ liệu (500 dòng) được mở để nhập -----
    ws.protection.sheet = True
    ws.protection.password = "kendu2026"  # đổi mật khẩu này nếu muốn, chỉ để chặn sửa nhầm, KHÔNG phải bảo mật cao
    for r in (1, HEADER_ROW, DESC_ROW, EXAMPLE_ROW):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).protection = ws.cell(row=r, column=c).protection.copy(locked=True)
    for r in range(DATA_START_ROW, DATA_START_ROW + 500):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.protection = cell.protection.copy(locked=False)
            cell.border = THIN_BORDER

    ws.sheet_view.showGridLines = True


def generate_template(table_key):
    if table_key not in SCHEMA_CONFIG:
        raise ValueError(f"Không tìm thấy cấu hình cho bảng '{table_key}' trong SCHEMA_CONFIG")
    cfg = SCHEMA_CONFIG[table_key]
    wb = Workbook()
    ws = wb.active
    ws.title = cfg["sheet_title"][:31]
    _build_sheet(ws, table_key, cfg)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, f"mau_{table_key}.xlsx")
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    import sys
    targets = sys.argv[1:] or list(SCHEMA_CONFIG.keys())
    for t in targets:
        path = generate_template(t)
        print(f"✅ Đã tạo: {path}")
