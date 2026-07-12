"""
import_from_excel.py
======================
Đọc file Excel đã điền (theo đúng mẫu do generate_excel_templates.py sinh ra) và nạp
dữ liệu vào Database — dùng khi onboarding khách hàng mới.

CÁCH DÙNG:
    python import_from_excel.py <ten_bang> <duong_dan_file.xlsx> [--dry-run]

    VD:
    python import_from_excel.py nhan_vien excel_templates/mau_nhan_vien.xlsx --dry-run
    python import_from_excel.py nhan_vien /path/to/khach_A_da_dien.xlsx

    --dry-run : chỉ kiểm tra + báo lỗi, KHÔNG ghi vào DB (luôn chạy thử bước này trước).

YÊU CẦU MÔI TRƯỜNG:
    Biến môi trường DATABASE_URL (connection string Postgres/Supabase của đúng tenant
    cần import), VD:
        export DATABASE_URL="postgresql://user:pass@host:5432/postgres"

NGUYÊN TẮC:
    - Không sửa cấu trúc cột trong file Excel (đã bị khoá) -> script luôn đọc đúng vị trí.
    - Dòng 1-4 của mỗi sheet là note/header/mô tả/ví dụ -> dữ liệu thật bắt đầu từ dòng 5.
    - Dòng nào thiếu cột bắt buộc (*) sẽ bị BÁO LỖI và bỏ qua, không làm dừng cả file
      (các dòng hợp lệ khác vẫn được nạp bình thường).
    - Bảng nhan_vien: tự sinh mã NV theo đúng quy tắc app đang dùng (C001, C002...).
    - Các bảng danh mục đơn giản (phòng ban, chức vụ...): dùng ON CONFLICT DO NOTHING,
      import trùng tên sẽ tự bỏ qua thay vì báo lỗi.
"""
import sys
import os
import re
from datetime import datetime

import psycopg2
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))
from generate_excel_templates import SCHEMA_CONFIG  # tái sử dụng đúng schema với file mẫu

DATA_START_ROW = 5  # khớp với generate_excel_templates.py

# Các bảng chỉ có 1 cột tên (danh mục đơn giản) -> xử lý theo nhánh riêng, đơn giản hơn
BANG_DANH_MUC_DON_GIAN = {
    "danh_muc_phong_ban", "danh_muc_loai_hop_dong", "danh_muc_trinh_do_hoc_van",
    "chuc_vu_danh_muc", "vi_tri_cong_tac", "chuc_danh_ung_vien",
}


def parse_ngay(gia_tri):
    """Chấp nhận dd/mm/yyyy (text) hoặc datetime (Excel tự nhận diện ngày)."""
    if gia_tri is None or gia_tri == "":
        return None
    if isinstance(gia_tri, datetime):
        return gia_tri.date()
    s = str(gia_tri).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return "INVALID"


def doc_du_lieu(sheet, cols):
    """Đọc toàn bộ dòng dữ liệu (từ DATA_START_ROW) thành list[dict], theo đúng thứ tự cột trong SCHEMA_CONFIG."""
    rows_out = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue  # bỏ qua dòng trống
        record = {}
        for (col_key, label, required, desc, example, choices), value in zip(cols, row):
            record[col_key] = value.strip() if isinstance(value, str) else value
        rows_out.append((row_idx, record))
    return rows_out


def kiem_tra_dong(record, cols):
    """Trả về list lỗi (chuỗi mô tả) cho 1 dòng dữ liệu. Rỗng = hợp lệ."""
    loi = []
    for col_key, label, required, desc, example, choices in cols:
        val = record.get(col_key)
        if required and (val is None or str(val).strip() == ""):
            loi.append(f"Thiếu '{label}' (bắt buộc)")
            continue
        if choices and val not in (None, "") and str(val) not in choices:
            loi.append(f"'{label}' = '{val}' không nằm trong danh sách hợp lệ: {', '.join(choices)}")
        if col_key.startswith("ngay_") and val not in (None, ""):
            if parse_ngay(val) == "INVALID":
                loi.append(f"'{label}' = '{val}' không đúng định dạng ngày (dd/mm/yyyy)")
    return loi


def import_danh_muc_don_gian(conn, table_key, sheet, cols, dry_run):
    cot_ten = cols[0][0]  # tên cột duy nhất, VD 'ten_phong_ban'
    du_lieu = doc_du_lieu(sheet, cols)
    ok, loi_list, trung = 0, [], 0
    if dry_run:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
            else:
                ok += 1
        return ok, loi_list, trung
    with conn.cursor() as c:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
                continue
            gia_tri = str(record[cot_ten]).strip()
            c.execute(
                f"INSERT INTO {table_key} ({cot_ten}) VALUES (%s) ON CONFLICT DO NOTHING",
                (gia_tri,)
            )
            if c.rowcount == 0:
                trung += 1
            else:
                ok += 1
    conn.commit()
    return ok, loi_list, trung


def import_nhan_vien(conn, sheet, cols, dry_run):
    du_lieu = doc_du_lieu(sheet, cols)
    ok, loi_list = 0, []
    if dry_run:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
            else:
                ok += 1
        return ok, loi_list, 0
    with conn.cursor() as c:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
                continue

            # Sinh mã NV tự động theo đúng quy tắc app đang dùng: C001, C002, ...
            c.execute("SELECT COALESCE(MAX(CAST(SUBSTRING(ma_nv FROM 2) AS INTEGER)), 0)+1 FROM nhan_vien WHERE ma_nv LIKE 'C%'")
            so_moi = c.fetchone()[0]
            ma_nv = f"C{so_moi:03d}"
            c.execute("SELECT COALESCE(MAX(STT),0)+1 FROM nhan_vien")
            stt_moi = c.fetchone()[0]

            cot_ten_list = [ck for ck, *_ in cols]
            gia_tri_list = []
            for col_key, *_ in cols:
                v = record.get(col_key)
                if col_key.startswith("ngay_"):
                    v = parse_ngay(v)
                    if v == "INVALID":
                        v = None
                gia_tri_list.append(v)

            cols_sql = ", ".join(["stt", "ma_nv"] + cot_ten_list)
            placeholders = ", ".join(["%s"] * (2 + len(cot_ten_list)))
            c.execute(
                f"INSERT INTO nhan_vien ({cols_sql}) VALUES ({placeholders})",
                [stt_moi, ma_nv] + gia_tri_list
            )
            ok += 1
    conn.commit()
    return ok, loi_list, 0


def import_ung_vien(conn, sheet, cols, dry_run):
    du_lieu = doc_du_lieu(sheet, cols)
    ok, loi_list = 0, []
    if dry_run:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
            else:
                ok += 1
        return ok, loi_list, 0
    with conn.cursor() as c:
        for row_idx, record in du_lieu:
            loi = kiem_tra_dong(record, cols)
            if loi:
                loi_list.append((row_idx, loi))
                continue
            cot_ten_list = [ck for ck, *_ in cols]
            gia_tri_list = []
            for col_key, *_ in cols:
                v = record.get(col_key)
                if col_key.startswith("ngay_"):
                    v = parse_ngay(v)
                    if v == "INVALID":
                        v = None
                gia_tri_list.append(v)
            cols_sql = ", ".join(cot_ten_list)
            placeholders = ", ".join(["%s"] * len(cot_ten_list))
            c.execute(
                f"INSERT INTO ung_vien ({cols_sql}, trang_thai) VALUES ({placeholders}, 'CHO_DUYET') RETURNING id",
                gia_tri_list
            )
            new_id = c.fetchone()[0]
            c.execute("UPDATE ung_vien SET ma_uv=%s WHERE id=%s", (f"UV{new_id:04d}", new_id))
            ok += 1
    conn.commit()
    return ok, loi_list, 0


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    table_key = sys.argv[1]
    file_path = sys.argv[2]
    dry_run = "--dry-run" in sys.argv

    if table_key not in SCHEMA_CONFIG:
        print(f"❌ Không hỗ trợ bảng '{table_key}'. Các bảng hỗ trợ: {', '.join(SCHEMA_CONFIG.keys())}")
        sys.exit(1)

    db_url = os.environ.get("DATABASE_URL")
    if not dry_run and not db_url:
        print("❌ Chưa thiết lập biến môi trường DATABASE_URL. Xem hướng dẫn ở đầu file.")
        sys.exit(1)

    cfg = SCHEMA_CONFIG[table_key]
    cols = cfg["columns"]

    wb = load_workbook(file_path, data_only=True)
    sheet = wb[cfg["sheet_title"][:31]] if cfg["sheet_title"][:31] in wb.sheetnames else wb.active

    conn = psycopg2.connect(db_url) if not dry_run else None
    try:
        if table_key == "nhan_vien":
            ok, loi_list, trung = import_nhan_vien(conn, sheet, cols, dry_run)
        elif table_key == "ung_vien":
            ok, loi_list, trung = import_ung_vien(conn, sheet, cols, dry_run)
        elif table_key in BANG_DANH_MUC_DON_GIAN:
            ok, loi_list, trung = import_danh_muc_don_gian(conn, table_key, sheet, cols, dry_run)
        else:
            print(f"❌ Bảng '{table_key}' chưa có logic import riêng — cần bổ sung.")
            sys.exit(1)
    finally:
        if conn:
            conn.close()

    print(f"\n{'🔍 [DRY-RUN — CHƯA GHI DB]' if dry_run else '✅ [ĐÃ GHI DB]'} Kết quả import bảng '{table_key}':")
    print(f"  ✅ Hợp lệ: {ok} dòng")
    if trung:
        print(f"  ⚠️  Trùng (bỏ qua): {trung} dòng")
    if loi_list:
        print(f"  ❌ Lỗi: {len(loi_list)} dòng")
        for row_idx, loi in loi_list:
            print(f"     - Dòng {row_idx}: {'; '.join(loi)}")
    else:
        print("  🎉 Không có dòng nào lỗi.")


if __name__ == "__main__":
    main()
